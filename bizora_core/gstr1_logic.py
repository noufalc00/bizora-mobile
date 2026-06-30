"""
GSTR-1 Report Logic Module.
Handles data extraction, categorization, and export for GSTR-1 GST compliance reports.
"""

import csv
import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from db import Database
from bizora_core.gst_compliance import (
    classify_invoice,
    is_interstate,
    normalized_gstin,
    place_of_supply_label,
    state_code,
    tax_totals_for_supply,
)


class GSTR1Logic:
    """Logic for generating GSTR-1 reports for GST compliance."""

    def __init__(self, db: Optional[Database] = None):
        self.db = db or Database()

    def generate_gstr1_report(self, company_id: int, from_date: str, to_date: str) -> Dict:
        """
        Generate complete GSTR-1 report for the specified date range.
        
        Args:
            company_id: Company ID
            from_date: Start date in YYYY-MM-DD format
            to_date: End date in YYYY-MM-DD format
            
        Returns:
            Dictionary containing all GSTR-1 sections
        """
        # Extract sales data
        sales_data = self._extract_sales_data(company_id, from_date, to_date)
        
        # Extract sales returns (credit notes)
        return_data = self._extract_sales_returns(company_id, from_date, to_date)
        
        # Get company details
        company_details = self._get_company_details(company_id)
        
        # Categorize sales
        b2b_data = self._categorize_b2b(sales_data, company_details)
        b2cl_data = self._categorize_b2cl(sales_data, company_details)
        b2cs_data = self._categorize_b2cs(sales_data, company_details)
        
        # Categorize returns
        b2b_returns = self._categorize_b2b_returns(return_data, company_details)
        b2cl_returns = self._categorize_b2cl_returns(return_data, company_details)
        
        # Generate additional sections
        hsn_summary = self._generate_hsn_summary(company_id, from_date, to_date)
        doc_summary = self._generate_document_summary(sales_data, return_data)
        
        # Compile report
        report = {
            "gstin": company_details.get('gstin', ''),
            "period": self._format_period(from_date),
            "b2b": b2b_data,
            "b2cl": b2cl_data,
            "b2cs": b2cs_data,
            "cdn": {
                "b2b": b2b_returns,
                "b2cl": b2cl_returns,
            },
            "hsn": hsn_summary,
            "doc_issue": doc_summary,
            "summary": self._calculate_summary(b2b_data, b2cl_data, b2cs_data)
        }
        
        return report

    def _extract_sales_data(self, company_id: int, from_date: str, to_date: str) -> List[Dict]:
        """Extract sales data with party details and items."""
        ph = self.db._get_placeholder()
        query = f"""
            SELECT 
                s.id,
                s.invoice_number,
                s.invoice_date,
                s.party_id,
                s.grand_total,
                s.sub_total,
                s.discount_total,
                s.tax_total,
                s.round_off,
                s.nature,
                s.form_of_sale,
                p.name as party_name,
                p.gstin as party_gstin,
                p.state as party_state,
                COALESCE(item_totals.taxable_value, 0) as taxable_value,
                COALESCE(item_totals.cgst_total, 0) as cgst_total,
                COALESCE(item_totals.sgst_total, 0) as sgst_total,
                COALESCE(item_totals.igst_total, 0) as igst_total,
                COALESCE(item_totals.cess_total, 0) as cess_total,
                COALESCE(item_totals.item_tax_total, 0) as item_tax_total
            FROM sales s
            LEFT JOIN parties p ON s.party_id = p.id
            LEFT JOIN (
                SELECT
                    si.sale_id as sale_id,
                    SUM(COALESCE(si.net_value, 0)) as taxable_value,
                    SUM(COALESCE(si.cgst_amount, 0)) as cgst_total,
                    SUM(COALESCE(si.sgst_amount, 0)) as sgst_total,
                    SUM(COALESCE(si.igst_amount, 0)) as igst_total,
                    SUM(COALESCE(si.cess_amount, 0)) as cess_total,
                    SUM(COALESCE(si.tax_amount, 0)) as item_tax_total
                FROM sales_items si
                GROUP BY si.sale_id
            ) item_totals ON item_totals.sale_id = s.id
            WHERE s.company_id = {ph}
                AND s.invoice_date BETWEEN {ph} AND {ph}
                AND COALESCE(s.status, 'Active') <> 'Voided'
            ORDER BY s.invoice_date, s.id
        """
        results = self.db.execute_query(query, (company_id, from_date, to_date))

        for sale in results:
            if not self._to_float(sale.get('tax_total')) and sale.get('item_tax_total'):
                sale['tax_total'] = sale.get('item_tax_total')
            if not self._to_float(sale.get('taxable_value')):
                sale['taxable_value'] = (
                    self._to_float(sale.get('sub_total'))
                    - self._to_float(sale.get('discount_total'))
                )
            self._apply_supply_tax_mapping(sale, self._get_company_state(company_id))

        return results

    def _extract_sales_returns(self, company_id: int, from_date: str, to_date: str) -> List[Dict]:
        """Extract sales return data (credit notes)."""
        ph = self.db._get_placeholder()
        query = f"""
            SELECT 
                sr.id,
                sr.return_no as invoice_number,
                sr.return_date as invoice_date,
                sr.party_id,
                sr.grand_total,
                sr.sub_total,
                sr.discount_total,
                sr.tax_total,
                sr.round_off,
                sr.nature,
                p.name as party_name,
                p.gstin as party_gstin,
                p.state as party_state
            FROM sales_returns sr
            LEFT JOIN parties p ON sr.party_id = p.id
            WHERE sr.company_id = {ph}
                AND sr.return_date BETWEEN {ph} AND {ph}
            ORDER BY sr.return_date, sr.id
        """
        results = self.db.execute_query(query, (company_id, from_date, to_date))
        
        # Calculate taxable_value and tax breakdown for each return.
        for ret in results:
            sub_total = self._to_float(ret.get('sub_total'))
            discount = self._to_float(ret.get('discount_total'))
            ret['taxable_value'] = sub_total - discount
            ret['cess_total'] = self._to_float(ret.get('cess_total'))
            self._apply_supply_tax_mapping(ret, self._get_company_state(company_id))
        
        return results

    def _get_company_details(self, company_id: int) -> Dict:
        """Get company GSTIN and state details."""
        ph = self.db._get_placeholder()
        query = f"""
            SELECT gstin, state, business_name
            FROM companies
            WHERE id = {ph}
        """
        result = self.db.execute_query(query, (company_id,))
        return result[0] if result else {}

    def _get_company_state(self, company_id: int) -> str:
        """Get company state for interstate calculation."""
        details = self._get_company_details(company_id)
        return (details.get('state') or '').strip().lower()

    def _categorize_b2b(self, sales_data: List[Dict], company_details: Dict) -> List[Dict]:
        """
        Categorize B2B sales (registered customers with GSTIN).
        Table 4 of GSTR-1.
        """
        b2b_list = []
        company_state = company_details.get('state', '').strip()
        
        for sale in sales_data:
            gstin = normalized_gstin(sale.get('party_gstin'))
            
            # B2B is strictly for structurally valid GSTIN customers.
            if not gstin:
                continue
                
            place_of_supply = self._determine_place_of_supply(sale, company_details)
            is_interstate = self._is_interstate(place_of_supply, company_state)
            igst, cgst, sgst, cess = self._tax_totals(sale, is_interstate)
            
            # Calculate tax rate
            taxable = self._to_float(sale.get('taxable_value'))
            tax = igst + cgst + sgst + cess
            rate = round((tax / taxable) * 100, 2) if taxable > 0 else 0.0
            
            b2b_list.append({
                "ctin": gstin,
                "cstin": gstin[:2],
                "inv": {
                    "inum": sale.get('invoice_number', ''),
                    "idt": self._format_date(sale.get('invoice_date')),
                    "val": self._to_float(sale.get('grand_total')),
                    "pos": place_of_supply,
                    "rchrg": "N",
                    "itms": [{
                        "num": 1,
                        "itm_det": {
                            "txval": taxable,
                            "rt": rate,
                            "iamt": igst,
                            "camt": cgst,
                            "samt": sgst,
                            "csamt": cess,
                        }
                    }]
                }
            })
            
        return b2b_list

    def _categorize_b2cl(self, sales_data: List[Dict], company_details: Dict) -> List[Dict]:
        """
        Categorize B2CL sales (unregistered interstate > 2.5L).
        Table 5 of GSTR-1.
        """
        b2cl_list = []
        company_state = company_details.get('state', '').strip()
        
        for sale in sales_data:
            gstin = normalized_gstin(sale.get('party_gstin'))
            place_of_supply = self._determine_place_of_supply(sale, company_details)
            grand_total = self._to_float(sale.get('grand_total'))
            
            # B2CL: No GSTIN, different state, > 2.5L
            if gstin:
                continue
            if not self._is_interstate(place_of_supply, company_state):
                continue
            if grand_total <= 250000:
                continue

            # Calculate tax rate
            taxable = self._to_float(sale.get('taxable_value'))
            igst, _cgst, _sgst, cess = self._tax_totals(sale, True)
            tax = igst + cess
            rate = round((tax / taxable) * 100, 2) if taxable > 0 else 0.0
            
            b2cl_list.append({
                "pos": place_of_supply,
                "inv": {
                    "inum": sale.get('invoice_number', ''),
                    "idt": self._format_date(sale.get('invoice_date')),
                    "val": grand_total,
                    "itms": [{
                        "num": 1,
                        "itm_det": {
                            "txval": taxable,
                            "rt": rate,
                            "iamt": igst,
                            "camt": 0.0,
                            "samt": 0.0,
                            "csamt": cess,
                        }
                    }]
                }
            })
            
        return b2cl_list

    def _categorize_b2cs(self, sales_data: List[Dict], company_details: Dict) -> List[Dict]:
        """
        Categorize B2CS sales (unregistered or <= 2.5L interstate).
        Table 7 of GSTR-1 - aggregated by place of supply and tax rate.
        """
        # Aggregate by (place_of_supply, tax_rate)
        aggregation = {}
        company_state = company_details.get('state', '').strip()
        
        for sale in sales_data:
            gstin = normalized_gstin(sale.get('party_gstin'))
            grand_total = self._to_float(sale.get('grand_total'))
            
            # Skip B2B (has GSTIN)
            if gstin:
                continue
                
            # Determine place of supply
            place_of_supply = self._determine_place_of_supply(sale, company_details)
            is_inter = self._is_interstate(place_of_supply, company_state)
            if is_inter and grand_total > 250000:
                continue
            
            # Calculate tax rate
            taxable = self._to_float(sale.get('taxable_value'))
            igst, cgst, sgst, cess = self._tax_totals(sale, is_inter)
            tax = igst + cgst + sgst + cess
            rate = round((tax / taxable) * 100, 2) if taxable > 0 else 0.0
            
            key = (place_of_supply, rate)
            
            if key not in aggregation:
                aggregation[key] = {
                    "pos": place_of_supply,
                    "typ": "OE",
                    "txval": 0.0,
                    "iamt": 0.0,
                    "camt": 0.0,
                    "samt": 0.0,
                    "csamt": 0.0,
                    "rt": rate,
                    "inv_count": 0
                }
            
            # Aggregate values
            agg = aggregation[key]
            agg['txval'] += taxable
            agg['iamt'] += igst
            agg['camt'] += cgst
            agg['samt'] += sgst
            agg['csamt'] += cess
            agg['inv_count'] += 1
            
        return list(aggregation.values())

    def _categorize_b2b_returns(self, return_data: List[Dict], company_details: Dict) -> List[Dict]:
        """Categorize B2B credit notes (Table 9B)."""
        # Similar to B2B but for returns
        b2b_returns = []
        company_state = company_details.get('state', '').strip()
        
        for ret in return_data:
            gstin = normalized_gstin(ret.get('party_gstin'))
            
            if not gstin:
                continue
                
            place_of_supply = self._determine_place_of_supply(ret, company_details)
            is_inter = self._is_interstate(place_of_supply, company_state)
            
            taxable = self._to_float(ret.get('taxable_value'))
            igst, cgst, sgst, cess = self._tax_totals(ret, is_inter)
            tax = igst + cgst + sgst + cess
            rate = round((tax / taxable) * 100, 2) if taxable > 0 else 0.0
            
            b2b_returns.append({
                "ctin": gstin,
                "nt": [{
                    "ntty": "C",  # Credit Note
                    "ntnum": ret.get('invoice_number', ''),
                    "ntdt": self._format_date(ret.get('invoice_date')),
                    "val": self._to_float(ret.get('grand_total')),
                    "pos": place_of_supply,
                    "itms": [{
                        "num": 1,
                        "itm_det": {
                            "txval": taxable,
                            "rt": rate,
                            "iamt": igst,
                            "camt": cgst,
                            "samt": sgst,
                            "csamt": cess,
                        }
                    }]
                }]
            })
            
        return b2b_returns

    def _categorize_b2cl_returns(self, return_data: List[Dict], company_details: Dict) -> List[Dict]:
        """Categorize B2CL credit notes."""
        # Similar to B2CL but for returns
        b2cl_returns = []
        company_state = company_details.get('state', '').strip()
        
        for ret in return_data:
            gstin = normalized_gstin(ret.get('party_gstin'))
            place_of_supply = self._determine_place_of_supply(ret, company_details)
            grand_total = self._to_float(ret.get('grand_total'))
            
            if gstin:
                continue
            if not self._is_interstate(place_of_supply, company_state):
                continue
            if grand_total <= 250000:
                continue
            
            taxable = self._to_float(ret.get('taxable_value'))
            igst, _cgst, _sgst, cess = self._tax_totals(ret, True)
            tax = igst + cess
            rate = round((tax / taxable) * 100, 2) if taxable > 0 else 0.0
            
            b2cl_returns.append({
                "pos": place_of_supply,
                "nt": [{
                    "ntty": "C",
                    "ntnum": ret.get('invoice_number', ''),
                    "ntdt": self._format_date(ret.get('invoice_date')),
                    "val": grand_total,
                    "itms": [{
                        "num": 1,
                        "itm_det": {
                            "txval": taxable,
                            "rt": rate,
                            "iamt": igst,
                            "camt": 0.0,
                            "samt": 0.0,
                            "csamt": cess,
                        }
                    }]
                }]
            })
            
        return b2cl_returns

    def _generate_hsn_summary(self, company_id: int, from_date: str, to_date: str) -> List[Dict]:
        """
        Generate HSN Summary (Table 12 of GSTR-1).
        Groups item rows strictly by HSN code and UQC.
        """
        ph = self.db._get_placeholder()
        query = f"""
            SELECT 
                COALESCE(NULLIF(si.hsn, ''), NULLIF(p.hsn, '')) as hsn,
                COALESCE(NULLIF(si.unit, ''), NULLIF(p.unit, ''), 'PCS') as uqc,
                SUM(COALESCE(si.quantity, 0)) as total_qty,
                SUM(COALESCE(si.net_value, 0)) as total_taxable_value,
                SUM(COALESCE(si.igst_amount, 0)) as total_igst,
                SUM(COALESCE(si.cgst_amount, 0)) as total_cgst,
                SUM(COALESCE(si.sgst_amount, 0)) as total_sgst,
                SUM(COALESCE(si.cess_amount, 0)) as total_cess
            FROM sales_items si
            INNER JOIN sales s ON si.sale_id = s.id
            LEFT JOIN products p ON si.product_id = p.id
            WHERE s.company_id = {ph}
                AND s.invoice_date BETWEEN {ph} AND {ph}
                AND COALESCE(s.status, 'Active') <> 'Voided'
                AND COALESCE(NULLIF(si.hsn, ''), NULLIF(p.hsn, '')) IS NOT NULL
                AND COALESCE(NULLIF(si.hsn, ''), NULLIF(p.hsn, '')) != ''
            GROUP BY
                COALESCE(NULLIF(si.hsn, ''), NULLIF(p.hsn, '')),
                COALESCE(NULLIF(si.unit, ''), NULLIF(p.unit, ''), 'PCS')
            ORDER BY
                COALESCE(NULLIF(si.hsn, ''), NULLIF(p.hsn, '')),
                COALESCE(NULLIF(si.unit, ''), NULLIF(p.unit, ''), 'PCS')
        """
        results = self.db.execute_query(query, (company_id, from_date, to_date))

        hsn_list = []
        
        for row in results:
            hsn = row.get('hsn', '')
            taxable = self._to_float(row.get('total_taxable_value'))
            igst = self._to_float(row.get('total_igst'))
            cgst = self._to_float(row.get('total_cgst'))
            sgst = self._to_float(row.get('total_sgst'))
            cess = self._to_float(row.get('total_cess'))
            tax = igst + cgst + sgst + cess
            rate = round((tax / taxable) * 100, 2) if taxable > 0 else 0.0

            hsn_list.append({
                "hsn": hsn,
                "desc": "",
                "uqc": str(row.get('uqc') or 'PCS').upper(),
                "qty": self._to_float(row.get('total_qty')),
                "val": taxable,
                "iamt": igst,
                "camt": cgst,
                "samt": sgst,
                "csamt": cess,
                "rt": rate
            })
            
        return hsn_list

    def _generate_document_summary(self, sales_data: List[Dict], return_data: List[Dict]) -> Dict:
        """
        Generate Document Summary (Table 13 of GSTR-1).
        Reports invoice sequence, count, and cancelled count.
        """
        if not sales_data:
            return {
                "doc_det": []
            }
        
        # Sort by invoice number
        sorted_sales = sorted(sales_data, key=lambda x: x.get('invoice_number', ''))
        
        start_inv = sorted_sales[0].get('invoice_number', '') if sorted_sales else ''
        end_inv = sorted_sales[-1].get('invoice_number', '') if sorted_sales else ''
        total_count = len(sales_data)
        
        # Count cancelled (assuming cancelled invoices have a status field or similar)
        # For now, we'll set cancelled to 0
        cancelled_count = 0
        
        return {
            "doc_det": [{
                "doc_num": start_inv,
                "doc_dt": self._format_date(sorted_sales[0].get('invoice_date')) if sorted_sales else '',
                "doc_typ": "INV",
                "from_nm": start_inv,
                "to_nm": end_inv,
                "totnum": total_count,
                "cancel_num": cancelled_count,
                "net_doc": total_count - cancelled_count
            }]
        }

    def _calculate_summary(self, b2b_data: List, b2cl_data: List, b2cs_data: List) -> Dict:
        """Calculate summary totals for verification."""
        total_taxable = 0.0
        total_tax = 0.0
        
        # Sum B2B
        for b2b in b2b_data:
            for inv in b2b.get('inv', {}).get('itms', []):
                det = inv.get('itm_det', {})
                total_taxable += det.get('txval', 0)
                total_tax += det.get('iamt', 0) + det.get('camt', 0) + det.get('samt', 0) + det.get('csamt', 0)
        
        # Sum B2CL
        for b2cl in b2cl_data:
            for inv in b2cl.get('inv', {}).get('itms', []):
                det = inv.get('itm_det', {})
                total_taxable += det.get('txval', 0)
                total_tax += det.get('iamt', 0) + det.get('csamt', 0)
        
        # Sum B2CS
        for b2cs in b2cs_data:
            total_taxable += b2cs.get('txval', 0)
            total_tax += b2cs.get('iamt', 0) + b2cs.get('camt', 0) + b2cs.get('samt', 0) + b2cs.get('csamt', 0)
        
        return {
            "total_taxable_value": round(total_taxable, 2),
            "total_tax": round(total_tax, 2),
            "total_invoice_count": len(b2b_data) + len(b2cl_data) + sum(b.get('inv_count', 0) for b in b2cs_data)
        }

    def export_to_json(self, report: Dict, filepath: str) -> bool:
        """Export GSTR-1 report to JSON file for portal upload."""
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2, default=str)
            return True
        except Exception as e:
            print(f"Error exporting JSON: {e}")
            return False

    def export_to_excel(self, report: Dict, filepath: str) -> bool:
        """Export portal-ready GSTR-1 sections to Excel or CSV files."""
        try:
            sheets = self.build_portal_export_rows(report)
            if filepath.lower().endswith(".csv"):
                self._export_sections_to_csv_files(sheets, filepath)
                return True

            try:
                import pandas as pd

                with pd.ExcelWriter(filepath) as writer:
                    for sheet_name, rows in sheets.items():
                        pd.DataFrame(rows).to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False,
                        )
            except ImportError:
                self._export_sections_with_openpyxl(sheets, filepath)
            return True
        except ImportError:
            print("Install pandas or openpyxl to export GSTR-1 Excel files.")
            return False
        except Exception as e:
            print(f"Error exporting Excel: {e}")
            return False

    def build_portal_export_rows(self, report: Dict) -> Dict[str, List[Dict[str, Any]]]:
        """Build portal-ready section rows with GSTR-1 offline tool headers."""
        return {
            "b2b": self._b2b_export_rows(report.get("b2b", [])),
            "b2cl": self._b2cl_export_rows(report.get("b2cl", [])),
            "b2cs": self._b2cs_export_rows(report.get("b2cs", [])),
            "hsn": self._hsn_export_rows(report.get("hsn", [])),
        }

    def _b2b_export_rows(self, b2b_data: List[Dict]) -> List[Dict[str, Any]]:
        """Flatten B2B report data using offline-tool column headers."""
        rows = []
        for b2b in b2b_data:
            inv = b2b.get("inv", {})
            for item in inv.get("itms", []):
                det = item.get("itm_det", {})
                rows.append({
                    "GSTIN/UIN of Recipient": b2b.get("ctin", ""),
                    "Invoice Number": inv.get("inum", ""),
                    "Invoice Date": inv.get("idt", ""),
                    "Invoice Value": inv.get("val", 0.0),
                    "Place Of Supply": inv.get("pos", ""),
                    "Reverse Charge": inv.get("rchrg", "N"),
                    "Invoice Type": "Regular",
                    "Rate": det.get("rt", 0.0),
                    "Taxable Value": det.get("txval", 0.0),
                    "Cess Amount": det.get("csamt", 0.0),
                })
        return rows

    def _b2cl_export_rows(self, b2cl_data: List[Dict]) -> List[Dict[str, Any]]:
        """Flatten B2CL report data using offline-tool column headers."""
        rows = []
        for b2cl in b2cl_data:
            inv = b2cl.get("inv", {})
            for item in inv.get("itms", []):
                det = item.get("itm_det", {})
                rows.append({
                    "Invoice Number": inv.get("inum", ""),
                    "Invoice Date": inv.get("idt", ""),
                    "Invoice Value": inv.get("val", 0.0),
                    "Place Of Supply": b2cl.get("pos", ""),
                    "Rate": det.get("rt", 0.0),
                    "Taxable Value": det.get("txval", 0.0),
                    "Cess Amount": det.get("csamt", 0.0),
                    "E-Commerce GSTIN": "",
                })
        return rows

    def _b2cs_export_rows(self, b2cs_data: List[Dict]) -> List[Dict[str, Any]]:
        """Flatten B2CS report data using offline-tool column headers."""
        rows = []
        for b2cs in b2cs_data:
            rows.append({
                "Type": b2cs.get("typ", "OE"),
                "Place Of Supply": b2cs.get("pos", ""),
                "Rate": b2cs.get("rt", 0.0),
                "Taxable Value": b2cs.get("txval", 0.0),
                "Cess Amount": b2cs.get("csamt", 0.0),
                "E-Commerce GSTIN": "",
            })
        return rows

    def _hsn_export_rows(self, hsn_data: List[Dict]) -> List[Dict[str, Any]]:
        """Flatten HSN summary using offline-tool column headers."""
        rows = []
        for hsn in hsn_data:
            rows.append({
                "HSN": hsn.get("hsn", ""),
                "Description": hsn.get("desc", ""),
                "UQC": hsn.get("uqc", "PCS"),
                "Total Quantity": hsn.get("qty", 0.0),
                "Total Value": (
                    self._to_float(hsn.get("val"))
                    + self._to_float(hsn.get("iamt"))
                    + self._to_float(hsn.get("camt"))
                    + self._to_float(hsn.get("samt"))
                    + self._to_float(hsn.get("csamt"))
                ),
                "Rate": hsn.get("rt", 0.0),
                "Taxable Value": hsn.get("val", 0.0),
                "Integrated Tax Amount": hsn.get("iamt", 0.0),
                "Central Tax Amount": hsn.get("camt", 0.0),
                "State/UT Tax Amount": hsn.get("samt", 0.0),
                "Cess Amount": hsn.get("csamt", 0.0),
            })
        return rows

    def _export_sections_to_csv_files(self, sheets: Dict[str, List[Dict[str, Any]]], filepath: str) -> None:
        """Write one CSV file per required GSTR-1 section."""
        base, _ext = os.path.splitext(filepath)
        for section, rows in sheets.items():
            section_path = f"{base}_{section}.csv"
            headers = list(rows[0].keys()) if rows else self._empty_export_headers(section)
            with open(section_path, "w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=headers)
                writer.writeheader()
                writer.writerows(rows)

    def _export_sections_with_openpyxl(self, sheets: Dict[str, List[Dict[str, Any]]], filepath: str) -> None:
        """Write portal-ready sheets without pandas when openpyxl is available."""
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        for sheet_name, rows in sheets.items():
            ws = wb.create_sheet(sheet_name)
            headers = list(rows[0].keys()) if rows else self._empty_export_headers(sheet_name)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            for row_index, row_data in enumerate(rows, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row_index, column=col, value=row_data.get(header, ""))
        wb.save(filepath)

    def _empty_export_headers(self, section: str) -> List[str]:
        """Return headers when a portal export section has no rows."""
        headers = {
            "b2b": [
                "GSTIN/UIN of Recipient", "Invoice Number", "Invoice Date",
                "Invoice Value", "Place Of Supply", "Reverse Charge",
                "Invoice Type", "Rate", "Taxable Value", "Cess Amount",
            ],
            "b2cl": [
                "Invoice Number", "Invoice Date", "Invoice Value",
                "Place Of Supply", "Rate", "Taxable Value",
                "Cess Amount", "E-Commerce GSTIN",
            ],
            "b2cs": [
                "Type", "Place Of Supply", "Rate", "Taxable Value",
                "Cess Amount", "E-Commerce GSTIN",
            ],
            "hsn": [
                "HSN", "Description", "UQC", "Total Quantity", "Total Value",
                "Rate", "Taxable Value", "Integrated Tax Amount",
                "Central Tax Amount", "State/UT Tax Amount", "Cess Amount",
            ],
        }
        return headers.get(section, [])

    def _create_b2b_sheet(self, wb, b2b_data):
        """Create B2B sheet in Excel workbook."""
        ws = wb.create_sheet("B2B")
        headers = ["GSTIN/UIN", "Invoice No", "Invoice Date", "Invoice Value", "Place of Supply", 
                   "Reverse Charge", "Invoice Type", "Tax Rate", "Taxable Value", "CGST", "SGST", "IGST", "CESS"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        row = 2
        for b2b in b2b_data:
            inv = b2b.get('inv', {})
            for item in inv.get('itms', []):
                det = item.get('itm_det', {})
                ws.cell(row, 1, b2b.get('ctin', ''))
                ws.cell(row, 2, inv.get('inum', ''))
                ws.cell(row, 3, inv.get('idt', ''))
                ws.cell(row, 4, inv.get('val', 0))
                ws.cell(row, 5, inv.get('pos', ''))
                ws.cell(row, 6, inv.get('rchrg', 'N'))
                ws.cell(row, 7, 'R')  # Regular
                ws.cell(row, 8, det.get('rt', 0))
                ws.cell(row, 9, det.get('txval', 0))
                ws.cell(row, 10, det.get('iamt', 0))
                ws.cell(row, 11, det.get('samt', 0))
                ws.cell(row, 12, 0)  # IGST not split in B2B
                ws.cell(row, 13, det.get('csamt', 0))
                row += 1

    def _create_b2cl_sheet(self, wb, b2cl_data):
        """Create B2CL sheet in Excel workbook."""
        ws = wb.create_sheet("B2CL")
        headers = ["Invoice No", "Invoice Date", "Invoice Value", "Place of Supply", "Tax Rate",
                   "Taxable Value", "IGST", "CESS"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for b2cl in b2cl_data:
            inv = b2cl.get('inv', {})
            for item in inv.get('itms', []):
                det = item.get('itm_det', {})
                ws.cell(row, 1, inv.get('inum', ''))
                ws.cell(row, 2, inv.get('idt', ''))
                ws.cell(row, 3, inv.get('val', 0))
                ws.cell(row, 4, b2cl.get('pos', ''))
                ws.cell(row, 5, det.get('rt', 0))
                ws.cell(row, 6, det.get('txval', 0))
                ws.cell(row, 7, 0)  # IGST handled differently
                ws.cell(row, 8, det.get('csamt', 0))
                row += 1

    def _create_b2cs_sheet(self, wb, b2cs_data):
        """Create B2CS sheet in Excel workbook."""
        ws = wb.create_sheet("B2CS")
        headers = ["Type", "Place of Supply", "Tax Rate", "Taxable Value", "CGST", "SGST", "CESS", "Total Value", "Invoice Count"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for b2cs in b2cs_data:
            ws.cell(row, 1, b2cs.get('typ', ''))
            ws.cell(row, 2, b2cs.get('pos', ''))
            ws.cell(row, 3, b2cs.get('rt', 0))
            ws.cell(row, 4, b2cs.get('txval', 0))
            ws.cell(row, 5, b2cs.get('iamt', 0))
            ws.cell(row, 6, b2cs.get('samt', 0))
            ws.cell(row, 7, b2cs.get('csamt', 0))
            ws.cell(row, 8, b2cs.get('txval', 0) + b2cs.get('iamt', 0) + b2cs.get('samt', 0) + b2cs.get('csamt', 0))
            ws.cell(row, 9, b2cs.get('inv_count', 0))
            row += 1

    def _create_hsn_sheet(self, wb, hsn_data):
        """Create HSN Summary sheet in Excel workbook."""
        ws = wb.create_sheet("HSN")
        headers = ["HSN Code", "Description", "UQC", "Total Quantity", "Total Taxable Value", 
                   "IGST", "CGST", "SGST", "CESS", "Tax Rate"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for hsn in hsn_data:
            ws.cell(row, 1, hsn.get('hsn', ''))
            ws.cell(row, 2, hsn.get('desc', ''))
            ws.cell(row, 3, hsn.get('uqc', ''))
            ws.cell(row, 4, hsn.get('qty', 0))
            ws.cell(row, 5, hsn.get('val', 0))
            ws.cell(row, 6, hsn.get('iamt', 0))
            ws.cell(row, 7, hsn.get('camt', 0))
            ws.cell(row, 8, hsn.get('samt', 0))
            ws.cell(row, 9, hsn.get('camt', 0))
            ws.cell(row, 10, hsn.get('rt', 0))
            row += 1

    def _create_doc_sheet(self, wb, doc_data):
        """Create Document Summary sheet in Excel workbook."""
        ws = wb.create_sheet("Doc Issue")
        headers = ["Document Type", "From No", "To No", "Total Count", "Cancelled Count", "Net Count"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for doc in doc_data.get('doc_det', []):
            ws.cell(row, 1, 'Invoice')
            ws.cell(row, 2, doc.get('from_nm', ''))
            ws.cell(row, 3, doc.get('to_nm', ''))
            ws.cell(row, 4, doc.get('totnum', 0))
            ws.cell(row, 5, doc.get('cancel_num', 0))
            ws.cell(row, 6, doc.get('net_doc', 0))
            row += 1

    # Helper methods
    def _determine_place_of_supply(self, record: Dict, company_details: Dict) -> str:
        """Determine place of supply for a transaction."""
        party_state = (record.get('party_state') or record.get('state') or '').strip()
        return place_of_supply_label(party_state, company_details.get('state', ''))

    def _is_interstate(self, place_of_supply: str, company_state: str) -> bool:
        """Check if transaction is interstate."""
        return is_interstate(place_of_supply, company_state)

    def _apply_supply_tax_mapping(self, record: Dict, company_state: str) -> None:
        """Normalize stored tax totals according to POS versus home state."""
        place_of_supply = self._determine_place_of_supply(
            record,
            {"state": company_state},
        )
        interstate = self._is_interstate(place_of_supply, company_state)
        igst, cgst, sgst, cess = self._tax_totals(record, interstate)
        record["place_of_supply"] = place_of_supply
        record["igst_total"] = igst
        record["cgst_total"] = cgst
        record["sgst_total"] = sgst
        record["cess_total"] = cess
        record["tax_total"] = igst + cgst + sgst + cess

    def _tax_totals(self, record: Dict, interstate: bool) -> tuple[float, float, float, float]:
        """Return IGST, CGST, SGST, and CESS for a record."""
        return tax_totals_for_supply(record, interstate)

    def classify_sale(self, sale: Dict, company_details: Dict) -> str:
        """Classify a sale record for tests and worker-side report generation."""
        pos = self._determine_place_of_supply(sale, company_details)
        return classify_invoice(
            sale.get("party_gstin"),
            pos,
            company_details.get("state", ""),
            sale.get("grand_total"),
        )

    def _format_date(self, date_str: str) -> str:
        """Format date to DD-MM-YYYY for GSTR-1."""
        if not date_str:
            return ''
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            return dt.strftime('%d-%m-%Y')
        except:
            return date_str

    def _format_period(self, date_str: str) -> str:
        """Format period to MMYYYY for GSTR-1 (e.g., 042026 for April 2026)."""
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            return dt.strftime('%m%Y')
        except:
            return ''

    def _to_float(self, value) -> float:
        """Convert value to float safely."""
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0
