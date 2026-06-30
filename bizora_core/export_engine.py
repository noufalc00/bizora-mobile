"""
Export Engine

Centralized export functionality for PDF and Excel generation.
All export logic is consolidated here to maintain separation of concerns.
UI files should only handle QFileDialog and pass data to this engine.
"""

from typing import List, Dict, Any, Optional
import os


class ExportEngine:
    """Centralized export engine for PDF and Excel generation."""
    
    def __init__(self, db=None):
        """Initialize the export engine.
        
        Args:
            db: Database connection instance
        """
        self.db = db
        self._openpyxl_available = False
        self._reportlab_available = False
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Check if required dependencies are available."""
        try:
            import openpyxl
            self._openpyxl_available = True
        except ImportError:
            self._openpyxl_available = False
        
        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            self._reportlab_available = True
        except ImportError:
            self._reportlab_available = False
    
    def is_excel_available(self) -> bool:
        """Check if Excel export is available."""
        return self._openpyxl_available
    
    def is_pdf_available(self) -> bool:
        """Check if PDF export is available."""
        return self._reportlab_available
    
    def _clean_currency_text(self, text: Any) -> str:
        """
        Clean currency text by replacing ₹ symbol with 'Rs. ' to prevent rendering issues.
        
        Args:
            text: Input text (string, number, or None)
            
        Returns:
            Cleaned string with ₹ replaced by Rs.
        """
        if text is None:
            return ""
        text_str = str(text)
        # Replace ₹ symbol with Rs. (with space after)
        return text_str.replace('₹', 'Rs. ')
    
    def _safe_float(self, value: Any, default: float = 0.0) -> float:
        """Safely convert value to float, handling empty strings and invalid formats."""
        if value is None or value == '':
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _number_to_words(self, num: int) -> str:
        """
        Convert a number to words (Indian numbering system).
        
        Args:
            num: Integer to convert
            
        Returns:
            Number in words
        """
        if num == 0:
            return "Zero"
        
        ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
                'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
                'Seventeen', 'Eighteen', 'Nineteen']
        tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
        
        def convert_less_than_thousand(n):
            if n == 0:
                return ""
            elif n < 20:
                return ones[n]
            elif n < 100:
                return tens[n // 10] + (" " + ones[n % 10] if n % 10 != 0 else "")
            else:
                return ones[n // 100] + " Hundred" + (" " + convert_less_than_thousand(n % 100) if n % 100 != 0 else "")
        
        if num < 1000:
            return convert_less_than_thousand(num)
        
        # Indian numbering system: Lakhs and Crores
        result = ""
        
        # Crores (1 crore = 10 million)
        if num >= 10000000:
            crores = num // 10000000
            result += convert_less_than_thousand(crores) + " Crore"
            num %= 10000000
            if num > 0:
                result += " "
        
        # Lakhs (1 lakh = 100,000)
        if num >= 100000:
            lakhs = num // 100000
            result += convert_less_than_thousand(lakhs) + " Lakh"
            num %= 100000
            if num > 0:
                result += " "
        
        # Thousands
        if num >= 1000:
            thousands = num // 1000
            result += convert_less_than_thousand(thousands) + " Thousand"
            num %= 1000
            if num > 0:
                result += " "
        
        # Less than thousand
        if num > 0:
            result += convert_less_than_thousand(num)
        
        return result.strip()
    
    def _get_company_header(self, company_id: int) -> Dict[str, str]:
        """
        Fetch company details from database for PDF header.
        
        Args:
            company_id: The ID of the company to fetch details for
            
        Returns:
            Dictionary containing company name, address, gstin, phone, email, state, pincode
        """
        if not self.db or not company_id:
            return {
                'name': 'Company Name',
                'address': '',
                'city': '',
                'state': '',
                'pincode': '',
                'phone': '',
                'email': '',
                'gstin': '',
                'print_phone': True,
                'print_email': True
            }
        
        try:
            ph = self.db._get_placeholder()
            result = self.db.execute_query(
                f"SELECT business_name, address, state, pincode, phone_number, email, gstin, print_phone, print_email FROM companies WHERE id = {ph}",
                (company_id,)
            )
            
            if result and len(result) > 0:
                row = result[0]
                return {
                    'name': row.get('business_name', 'Company Name'),
                    'address': row.get('address', ''),
                    'city': '',  # City may be part of address or separate field
                    'state': row.get('state', ''),
                    'pincode': row.get('pincode', ''),
                    'phone': row.get('phone_number', ''),
                    'email': row.get('email', ''),
                    'gstin': row.get('gstin', ''),
                    'print_phone': row.get('print_phone', 1) == 1,
                    'print_email': row.get('print_email', 1) == 1
                }
            else:
                return {
                    'name': 'Company Name',
                    'address': '',
                    'city': '',
                    'state': '',
                    'pincode': '',
                    'phone': '',
                    'email': '',
                    'gstin': '',
                    'print_phone': True,
                    'print_email': True
                }
        except Exception as e:
            print(f"[ExportEngine] Error fetching company header: {e}")
            return {
                'name': 'Company Name',
                'address': '',
                'city': '',
                'state': '',
                'pincode': '',
                'phone': '',
                'email': '',
                'gstin': '',
                'print_phone': True,
                'print_email': True
            }
    
    def export_table_to_excel(self, title: str, headers: List[str], data: List[List[Any]], filepath: str) -> Dict[str, Any]:
        """
        Export table data to Excel file.
        
        Args:
            title: Title for the Excel sheet
            headers: List of column headers
            data: List of rows (each row is a list of values)
            filepath: Full path to save the Excel file
            
        Returns:
            Dict with success status and message
        """
        if not self._openpyxl_available:
            return {
                'success': False,
                'error': 'openpyxl is not available. Install it with: pip install openpyxl'
            }
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title
            
            # Define styles
            header_fill = PatternFill("solid", fgColor="1e3a5f")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Write data with alternating row colors
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = border
                    
                    # Alternating row colors
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="f3f4f6")
            
            # Auto-adjust column widths
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                column = ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)]
                
                # Check header length
                if headers[col_idx - 1]:
                    max_length = len(str(headers[col_idx - 1]))
                
                # Check data length
                for row_idx in range(2, len(data) + 2):
                    if row_idx - 2 < len(data) and col_idx - 1 < len(data[row_idx - 2]):
                        cell_value = data[row_idx - 2][col_idx - 1]
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                
                # Set width (with some padding)
                column.width = min(max_length + 2, 50)
            
            wb.save(filepath)
            
            return {
                'success': True,
                'message': f'Excel file saved successfully to {filepath}'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Failed to export Excel: {str(e)}'
            }
    
    def export_table_to_pdf(self, title: str, headers: List[str], data: List[List[Any]], filepath: str, company_id: Optional[int] = None) -> Dict[str, Any]:
        """
        Export table data to PDF file with company header.
        
        Args:
            title: Title for the PDF document
            headers: List of column headers
            data: List of rows (each row is a list of values)
            filepath: Full path to save the PDF file
            company_id: Optional company ID for header
            
        Returns:
            Dict with success status and message
        """
        if not self._reportlab_available:
            return {
                'success': False,
                'error': 'reportlab is not available. Install it with: pip install reportlab'
            }
        
        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = []
            
            # Add company header if company_id provided
            if company_id:
                company_header = self._get_company_header(company_id)
                
                # Build full address line
                address_parts = []
                if company_header['address']:
                    address_parts.append(company_header['address'])
                if company_header['state']:
                    address_parts.append(company_header['state'])
                if company_header['pincode']:
                    address_parts.append(company_header['pincode'])
                full_address = ", ".join(address_parts) if address_parts else ""
                
                # Build contact line (respect print flags)
                contact_parts = []
                if company_header['phone'] and company_header.get('print_phone', True):
                    contact_parts.append(f"Phone: {company_header['phone']}")
                if company_header['email'] and company_header.get('print_email', True):
                    contact_parts.append(f"Email: {company_header['email']}")
                contact_line = " | ".join(contact_parts) if contact_parts else ""
                
                # Line 1: Company Name (Helvetica-Bold, Size: 16, Centered)
                name_style = styles['Heading1']
                name_style.alignment = 1  # Center
                name_style.fontName = 'Helvetica-Bold'
                name_style.fontSize = 16
                elements.append(Paragraph(company_header['name'], name_style))
                
                # Line 2: Full Address (Helvetica, Size: 10, Centered)
                if full_address:
                    address_style = styles['Normal']
                    address_style.alignment = 1  # Center
                    address_style.fontName = 'Helvetica'
                    address_style.fontSize = 10
                    elements.append(Paragraph(full_address, address_style))
                
                # Line 3: Contact Line (Helvetica, Size: 10, Centered)
                if contact_line:
                    contact_style = styles['Normal']
                    contact_style.alignment = 1  # Center
                    contact_style.fontName = 'Helvetica'
                    contact_style.fontSize = 10
                    elements.append(Paragraph(contact_line, contact_style))
                
                # Line 4: GSTIN (Helvetica-Bold, Size: 10, Centered)
                if company_header['gstin']:
                    gstin_style = styles['Normal']
                    gstin_style.alignment = 1  # Center
                    gstin_style.fontName = 'Helvetica-Bold'
                    gstin_style.fontSize = 10
                    elements.append(Paragraph(f"GSTIN: {company_header['gstin']}", gstin_style))
                
                # Add spacer and horizontal line
                elements.append(Spacer(1, 0.15 * inch))
                
                # Draw horizontal line using a table with a single line
                line_table = Table([['']], colWidths=[7.5 * inch])
                line_table.setStyle(TableStyle([
                    ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
                ]))
                elements.append(line_table)
                
                elements.append(Spacer(1, 0.2 * inch))
            
            # Prepare data with headers - clean currency symbols
            table_data = [[self._clean_currency_text(h) for h in headers]]
            for row in data:
                cleaned_row = [self._clean_currency_text(cell) for cell in row]
                table_data.append(cleaned_row)
            
            # Create table
            table = Table(table_data)
            
            # Define table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ])
            
            table.setStyle(table_style)
            
            # Add title
            title_style = styles['Heading2']
            title_style.alignment = 1  # Center
            title_paragraph = Paragraph(title, title_style)
            elements.append(title_paragraph)
            elements.append(Paragraph("<br/><br/>", styles['Normal']))
            
            elements.append(table)
            
            doc.build(elements)
            
            return {
                'success': True,
                'message': f'PDF file saved successfully to {filepath}'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Failed to export PDF: {str(e)}'
            }
    
    def export_voucher_invoice_pdf(self, voucher_data: Dict[str, Any], filepath: str, company_id: Optional[int] = None) -> Dict[str, Any]:
        """
        Export voucher/invoice to PDF file with dynamic title based on voucher type and company header.
        
        Args:
            voucher_data: Dictionary containing voucher information
            filepath: Full path to save the PDF file
            company_id: Optional company ID for header (if not provided, uses voucher_data company_details)
            
        Returns:
            Dict with success status and message
        """
        if not self._reportlab_available:
            return {
                'success': False,
                'error': 'reportlab is not available. Install it with: pip install reportlab'
            }
        
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []
            
            # Get company header from database if company_id provided
            if company_id:
                company_header = self._get_company_header(company_id)
                
                # Build full address line
                address_parts = []
                if company_header['address']:
                    address_parts.append(company_header['address'])
                if company_header['state']:
                    address_parts.append(company_header['state'])
                if company_header['pincode']:
                    address_parts.append(company_header['pincode'])
                full_address = ", ".join(address_parts) if address_parts else ""
                
                # Build contact line (respect print flags)
                contact_parts = []
                if company_header['phone'] and company_header.get('print_phone', True):
                    contact_parts.append(f"Phone: {company_header['phone']}")
                if company_header['email'] and company_header.get('print_email', True):
                    contact_parts.append(f"Email: {company_header['email']}")
                contact_line = " | ".join(contact_parts) if contact_parts else ""
                
                # Line 1: Company Name (Helvetica-Bold, Size: 18, Centered, Dark color)
                name_style = styles['Heading1']
                name_style.alignment = 1  # Center
                name_style.fontName = 'Helvetica-Bold'
                name_style.fontSize = 18
                name_style.textColor = colors.HexColor('#0f172a')
                elements.append(Paragraph(company_header['name'], name_style))
                
                # Line 2: Full Address (Helvetica, Size: 9, Centered, Grey)
                if full_address:
                    address_style = styles['Normal']
                    address_style.alignment = 1  # Center
                    address_style.fontName = 'Helvetica'
                    address_style.fontSize = 9
                    address_style.textColor = colors.HexColor('#64748b')
                    elements.append(Paragraph(full_address, address_style))
                
                # Line 3: Contact Line (Helvetica, Size: 9, Centered, Grey)
                if contact_line:
                    contact_style = styles['Normal']
                    contact_style.alignment = 1  # Center
                    contact_style.fontName = 'Helvetica'
                    contact_style.fontSize = 9
                    contact_style.textColor = colors.HexColor('#64748b')
                    elements.append(Paragraph(contact_line, contact_style))
                
                # Line 4: GSTIN (Helvetica-Bold, Size: 9, Centered, Dark Grey)
                if company_header['gstin']:
                    gstin_style = styles['Normal']
                    gstin_style.alignment = 1  # Center
                    gstin_style.fontName = 'Helvetica-Bold'
                    gstin_style.fontSize = 9
                    gstin_style.textColor = colors.HexColor('#475569')
                    elements.append(Paragraph(f"GSTIN: {company_header['gstin']}", gstin_style))
                
                # Add spacer and horizontal line (light grey)
                elements.append(Spacer(1, 0.2 * inch))
                
                # Draw horizontal line using a table with a single line
                line_table = Table([['']], colWidths=[5.5 * inch])
                line_table.setStyle(TableStyle([
                    ('LINEABOVE', (0, 0), (-1, 0), 1, colors.HexColor('#cbd5e1')),
                ]))
                elements.append(line_table)
                
                elements.append(Spacer(1, 0.2 * inch))
            else:
                # Fallback to company_details in voucher_data (for backward compatibility)
                company_name = voucher_data.get('company_name', 'Company Name')
                company_address = voucher_data.get('company_address', '')
                company_gst = voucher_data.get('company_gst', '')
                
                elements.append(Paragraph(f"<b>{company_name}</b>", styles['Heading1']))
                if company_address:
                    elements.append(Paragraph(company_address, styles['Normal']))
                if company_gst:
                    elements.append(Paragraph(f"GSTIN: {company_gst}", styles['Normal']))
                elements.append(Spacer(1, 0.3 * inch))
            
            # Voucher details
            voucher_type = voucher_data.get('voucher_type', 'Invoice')
            voucher_no = voucher_data.get('voucher_no', '')
            voucher_date = voucher_data.get('voucher_date', '')
            
            # Dynamic title based on voucher type
            title_map = {
                'Sales': 'TAX INVOICE',
                'SALES': 'TAX INVOICE',
                'sale': 'TAX INVOICE',
                'Purchase': 'PURCHASE BILL',
                'PURCHASE': 'PURCHASE BILL',
                'purchase': 'PURCHASE BILL',
                'Sales Return': 'CREDIT NOTE',
                'SALES RETURN': 'CREDIT NOTE',
                'sales_return': 'CREDIT NOTE',
                'Purchase Return': 'DEBIT NOTE',
                'PURCHASE RETURN': 'DEBIT NOTE',
                'purchase_return': 'DEBIT NOTE',
                'Quotation': 'QUOTATION',
                'QUOTATION': 'QUOTATION',
                'quotation': 'QUOTATION',
                'Estimate': 'ESTIMATE',
                'ESTIMATE': 'ESTIMATE',
                'estimate': 'ESTIMATE',
            }
            
            document_title = title_map.get(voucher_type, voucher_type.upper())
            
            # Document title
            title_style = styles['Heading2']
            title_style.alignment = 1  # Center
            title_style.fontName = 'Helvetica-Bold'
            title_style.fontSize = 12
            title_style.textColor = colors.HexColor('#0f172a')
            elements.append(Paragraph(document_title, title_style))
            elements.append(Spacer(1, 0.15 * inch))
            
            # Sub-Header: 2-column table for Billed To and Invoice Details
            party_name = voucher_data.get('party_name', '')
            billed_to_text = f"Billed To:\n{party_name}" if party_name else "Billed To:"
            invoice_details_text = f"Invoice No: {voucher_no}\nDate: {voucher_date}"
            
            sub_header_data = [
                [billed_to_text, invoice_details_text]
            ]
            
            sub_header_table = Table(sub_header_data, colWidths=[3.5*inch, 2.5*inch])
            sub_header_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#475569')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ]))
            
            elements.append(sub_header_table)
            elements.append(Spacer(1, 0.3 * inch))
            
            # Line items
            items = voucher_data.get('items', [])
            if items:
                item_headers = ['S.No', 'Item Description', 'Quantity', 'Rate', 'Amount']
                item_data = [[self._clean_currency_text(h) for h in item_headers]]
                
                for idx, item in enumerate(items, 1):
                    # Handle both dictionary and list-based item data
                    if isinstance(item, dict):
                        item_name = item.get('name', item.get('item_name', item.get('product_name', '')))
                        quantity = self._safe_float(item.get('quantity'))
                        rate = self._safe_float(item.get('rate'))
                        # Calculate line-item amount as qty * rate (NOT the voucher grand total)
                        amount = quantity * rate
                    else:
                        # Fallback for list-based data
                        item_name = str(item[0]) if len(item) > 0 else ''
                        quantity = self._safe_float(item[1] if len(item) > 1 else 0)
                        rate = self._safe_float(item[2] if len(item) > 2 else 0)
                        # Calculate line-item amount as qty * rate
                        amount = quantity * rate
                    
                    item_data.append([
                        str(idx),
                        self._clean_currency_text(item_name),
                        self._clean_currency_text(quantity),
                        self._clean_currency_text(rate),
                        self._clean_currency_text(amount)
                    ])
                
                item_table = Table(item_data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1*inch, 1.5*inch])
                item_table.setStyle(TableStyle([
                    # Header row - modern light blue
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#334155')),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    # Data rows - clean white with light borders
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                    # Column-by-column alignment (both header and data rows)
                    # Column 0 (S.No): CENTER
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    # Column 1 (Item Description): LEFT
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    # Column 2 (Quantity): CENTER
                    ('ALIGN', (2, 0), (2, -1), 'CENTER'),
                    # Column 3 (Rate): RIGHT
                    ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                    # Column 4 (Amount): RIGHT
                    ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
                    # Right padding for Rate and Amount columns
                    ('RIGHTPADDING', (3, 0), (4, -1), 10),
                    # Minimal borders - only horizontal lines
                    ('LINEABOVE', (0, 1), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                    ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#cbd5e1')),
                    ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor('#cbd5e1')),
                    # Alternate row backgrounds (very subtle)
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafbfc')]),
                ]))
                
                elements.append(item_table)
                elements.append(Spacer(1, 0.3 * inch))
            
            # Totals
            total_amount = self._safe_float(voucher_data.get('total_amount'))
            tax_amount = self._safe_float(voucher_data.get('tax_amount'))
            grand_total = self._safe_float(voucher_data.get('grand_total'))
            if grand_total == 0:
                grand_total = total_amount + tax_amount
            tax_breakdown = voucher_data.get('tax_breakdown', [])
            
            # Build total data with tax breakdown
            total_data = [
                ['Sub Total', self._clean_currency_text(str(total_amount))]
            ]
            
            # Add individual tax components if available
            if tax_breakdown:
                for tax_item in tax_breakdown:
                    tax_name = tax_item.get('name', 'Tax')
                    tax_amt = self._safe_float(tax_item.get('amount'))
                    if tax_amt > 0:
                        total_data.append([tax_name, self._clean_currency_text(str(tax_amt))])
            else:
                # Fallback to single tax line if no breakdown
                if tax_amount > 0:
                    total_data.append(['Tax', self._clean_currency_text(str(tax_amount))])
            
            # Grand Total (emphasized)
            total_data.append(['Grand Total', self._clean_currency_text(str(grand_total))])
            
            total_table = Table(total_data, colWidths=[5*inch, 1.5*inch])
            total_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -2), 9),
                ('VALIGN', (0, 0), (-1, -2), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -2), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -2), 4),
                # Column-by-column alignment
                # Column 0 (Labels): RIGHT to align with Amount column
                ('ALIGN', (0, 0), (0, -2), 'RIGHT'),
                # Column 1 (Values): RIGHT to match Amount column alignment
                ('ALIGN', (1, 0), (1, -2), 'RIGHT'),
                # Right padding for values column to match item grid
                ('RIGHTPADDING', (1, 0), (1, -2), 10),
                # Grand Total row - emphasized
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 11),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#0f172a')),
                ('ALIGN', (0, -1), (0, -1), 'RIGHT'),
                ('ALIGN', (1, -1), (1, -1), 'RIGHT'),
                ('VALIGN', (0, -1), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, -1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
                ('RIGHTPADDING', (1, -1), (1, -1), 10),
                # Minimal borders
                ('LINEABOVE', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ]))
            
            elements.append(total_table)
            elements.append(Spacer(1, 0.2 * inch))
            
            # Footer: Amount in words
            grand_total = voucher_data.get('grand_total', 0)
            try:
                gt_float = float(grand_total)
                amount_words = f"Amount in words: {self._number_to_words(int(gt_float))} Rupees Only"
            except:
                amount_words = f"Amount in words: {self._clean_currency_text(str(grand_total))}"
            
            amount_words_style = styles['Normal']
            amount_words_style.fontName = 'Helvetica-Oblique'
            amount_words_style.fontSize = 9
            amount_words_style.textColor = colors.HexColor('#64748b')
            elements.append(Paragraph(amount_words, amount_words_style))

            if str(voucher_type).strip().lower() in ("quotation", "estimate"):
                terms_style = styles['Normal']
                terms_style.fontName = 'Helvetica'
                terms_style.fontSize = 9
                terms_style.textColor = colors.black
                elements.append(Spacer(1, 0.2 * inch))
                elements.append(Paragraph(
                    "<b>Terms:</b> This document is a quotation/estimate for reference only. "
                    "It is not a final invoice and does not confirm sale, payment, ledger posting, or stock movement.",
                    terms_style,
                ))

            elements.append(Spacer(1, 0.6 * inch))
            
            # Sign-off section
            signoff_data = [
                ['', 'Authorized Signatory']
            ]
            signoff_table = Table(signoff_data, colWidths=[4*inch, 2*inch])
            signoff_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#475569')),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
            ]))
            elements.append(signoff_table)
            
            doc.build(elements)
            
            return {
                'success': True,
                'message': f'{document_title} PDF saved successfully to {filepath}'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Failed to export invoice PDF: {str(e)}'
            }

    def export_quotation_pdf(self, voucher_data: Dict[str, Any], filepath: str, company_id: Optional[int] = None) -> Dict[str, Any]:
        """Export a quotation/estimate PDF without changing Sales Entry output."""
        quotation_data = dict(voucher_data or {})
        quotation_data['voucher_type'] = quotation_data.get('voucher_type') or 'Quotation'
        return self.export_voucher_invoice_pdf(quotation_data, filepath, company_id)
