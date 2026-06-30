import random
import time
from pathlib import Path

from PySide6.QtWidgets import *
from PySide6.QtCore import QEvent, Qt, QTimer, QModelIndex, QSize
from PySide6.QtGui import QTextCursor, QColor, QIcon, QStandardItem, QStandardItemModel
from config import active_company_manager
from db import Database
from bizora_core.product_logic import ProductLogic
from bizora_core.product_settings_logic import (
    DEFAULT_ENTER_JUMP_FIELDS,
    get_product_page_settings,
)
from ui import theme
from ui.book_report_common import report_detail_dialog_style
from ui.checkbox_style import create_checkbox
from ui.product_settings_dialog import ProductSettingsDialog
from ui.table_header_utils import apply_read_only_report_table_selection
from ui.date_formats import configure_qdate_edit, format_display_date, qdate_to_db, qdate_to_display, db_to_qdate
from ui.ui_memory import UiMemoryMixin

_APP_ROOT = Path(__file__).resolve().parent.parent
_SETTINGS_ICON = _APP_ROOT / "assets" / "icons" / "settings.svg"
PDF_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
try:
    from docx import Document
    from docx.shared import Pt
    WORD_AVAILABLE = True
except ImportError:
    WORD_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ProductsWidget(UiMemoryMixin, QWidget):

    def __init__(self, db=None):
        super().__init__()
        self.db = db or Database()
        self.product_logic = ProductLogic(self.db)
        self.current_product_id = None
        self.products_data = []
        self.visible_products = []
        self.pdf_available = PDF_AVAILABLE
        self.word_available = WORD_AVAILABLE
        self.excel_available = EXCEL_AVAILABLE
        self.source_context = None
        self.source_purchase_widget = None
        self.source_purchase_row = None
        self.page_settings = {
            "allow_duplicate": False,
            "show_name_list": False,
            "enter_jump_fields": list(DEFAULT_ENTER_JUMP_FIELDS),
        }
        self._name_completer = None
        self._name_completer_model = None
        self._name_completer_signals_connected = False
        self._name_search_timer = QTimer(self)
        self._name_search_timer.setSingleShot(True)
        self._name_search_timer.setInterval(200)
        self._name_search_timer.timeout.connect(self._do_name_search)
        self.setup_ui()
        self.load_products()
        self.clear_form()
        self._load_product_page_settings()
        self._apply_product_page_settings()
        self._init_ui_memory()

    def setup_ui(self):
        self.setStyleSheet(theme.master_page_background_style())
        layout = QVBoxLayout(self)
        title = QLabel('Products / Services')
        title.setStyleSheet(theme.master_page_title_style(24))
        layout.addWidget(title)
        nav_layout = QHBoxLayout()
        nav_layout.setContentsMargins(0, 10, 0, 10)
        self.entry_btn = QPushButton('Product Entry')
        self.entry_btn.setStyleSheet(theme.master_nav_primary_button_style())
        self.entry_btn.clicked.connect(self.show_entry_page)
        self.list_btn = QPushButton('Product List')
        self.list_btn.setStyleSheet(theme.master_nav_secondary_button_style())
        self.list_btn.clicked.connect(self.show_list_page)
        nav_layout.addWidget(self.entry_btn)
        nav_layout.addWidget(self.list_btn)
        nav_layout.addStretch()
        self.settings_btn = QPushButton('Settings')
        self.settings_btn.setToolTip('Product / Service Settings')
        self.settings_btn.setFixedHeight(34)
        self.settings_btn.setMinimumWidth(112)
        if _SETTINGS_ICON.exists():
            self.settings_btn.setIcon(QIcon(str(_SETTINGS_ICON)))
            self.settings_btn.setIconSize(QSize(20, 20))
        self.settings_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.settings_btn.setStyleSheet(theme.master_settings_3d_button_style())
        self.settings_btn.clicked.connect(self.open_product_settings)
        nav_layout.addWidget(self.settings_btn)
        layout.addLayout(nav_layout)
        self.stack_widget = QStackedWidget()
        self.entry_page = self.create_entry_page()
        self.list_page = self.create_list_page()
        self.stack_widget.addWidget(self.entry_page)
        self.stack_widget.addWidget(self.list_page)
        layout.addWidget(self.stack_widget)

    def _apply_nav_styles(self, active: str) -> None:
        self.entry_btn.setStyleSheet(theme.master_nav_primary_button_style() if active == 'entry' else theme.master_nav_secondary_button_style())
        self.list_btn.setStyleSheet(theme.master_nav_primary_button_style() if active == 'list' else theme.master_nav_secondary_button_style())

    def refresh_theme(self) -> None:
        """Re-apply theme-aware styles after a global theme change."""
        self.setStyleSheet(theme.master_page_background_style())
        active = 'list' if self.stack_widget.currentWidget() == self.list_page else 'entry'
        self._apply_nav_styles(active)
        if hasattr(self, 'settings_btn'):
            self.settings_btn.setStyleSheet(theme.master_settings_3d_button_style())

    def open_product_settings(self) -> None:
        """Open the Product / Service settings dialog."""
        try:
            dialog = ProductSettingsDialog(parent=self, db=self.db)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self._load_product_page_settings()
                self._apply_product_page_settings()
        except Exception as exc:
            QMessageBox.critical(self, 'Settings', f'Failed to open product settings:\n{exc}')

    def _load_product_page_settings(self) -> None:
        """Load page settings for the active company."""
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                self.page_settings = {
                    "allow_duplicate": False,
                    "show_name_list": False,
                    "enter_jump_fields": list(DEFAULT_ENTER_JUMP_FIELDS),
                }
                return
            self.page_settings = get_product_page_settings(self.db, active_company['id'])
        except Exception:
            self.page_settings = {
                "allow_duplicate": False,
                "show_name_list": False,
                "enter_jump_fields": list(DEFAULT_ENTER_JUMP_FIELDS),
            }

    def _apply_product_page_settings(self) -> None:
        """Apply loaded settings to live widgets."""
        if self.page_settings.get('show_name_list'):
            self._setup_name_completer()
        else:
            self._teardown_name_completer()

    def _get_field_widget_map(self) -> dict:
        """Map setting field keys to entry-page widgets."""
        return {
            'name': self.name_input,
            'barcode': self.barcode_input,
            'hsn': self.hsn_input,
            'color': self.color_input,
            'size': self.size_input,
            'unit': self.unit,
            'category': self.category,
            'purchase_rate': self.purchase_rate,
            'sale_price': self.sale_price,
            'wholesale_rate': self.wholesale_rate,
            'mrp': self.mrp,
            'cgst': self.cgst,
            'sgst': self.sgst,
            'igst': self.igst,
            'cess': self.cess,
            'reorder_level': self.reorder_level,
            'description': self.description,
            'qty': self.qty,
        }

    def _get_enter_field_order(self) -> list:
        """Build the active Enter/Escape field order from page settings."""
        field_map = self._get_field_widget_map()
        enabled_keys = self.page_settings.get('enter_jump_fields') or list(DEFAULT_ENTER_JUMP_FIELDS)
        field_order = []
        for field_key in enabled_keys:
            if field_key == 'barcode' and self.auto_barcode.isChecked():
                continue
            widget = field_map.get(field_key)
            if widget is not None:
                field_order.append(widget)
        if not field_order:
            field_order = [self.name_input]
        return field_order

    def _disconnect_name_completer_signals(self) -> None:
        """Detach live completer signal hooks without destroying the debounce timer."""
        try:
            self._name_search_timer.stop()
        except Exception:
            pass
        if self._name_completer_signals_connected:
            try:
                self.name_input.textChanged.disconnect(self._on_name_completer_text_changed)
            except (TypeError, RuntimeError):
                pass
            self._name_completer_signals_connected = False
        if self._name_completer is not None:
            try:
                self._name_completer.activated[QModelIndex].disconnect(
                    self._on_name_suggestion_selected
                )
            except (TypeError, RuntimeError):
                pass

    def _teardown_name_completer(self) -> None:
        """Remove the product-name suggestion completer."""
        try:
            self._disconnect_name_completer_signals()
            if self.name_input.completer():
                old_completer = self.name_input.completer()
                self.name_input.setCompleter(None)
                old_completer.deleteLater()
            self._name_completer = None
            self._name_completer_model = None
        except Exception:
            pass

    def _on_name_completer_text_changed(self, _text: str) -> None:
        """Debounce name typing before querying product suggestions."""
        self._name_search_timer.start()

    def _do_name_search(self) -> None:
        """Populate the name suggestion list from the active company catalog."""
        model = self._name_completer_model
        completer = self._name_completer
        if model is None or completer is None:
            return
        text = self.name_input.text().strip()
        model.clear()
        if not text:
            return
        active_company = active_company_manager.get_active_company()
        if not active_company:
            return
        try:
            matches = self.db.search_products_limited(active_company['id'], text, 100)
        except Exception:
            matches = []
        text_lower = text.lower()
        seen_names: set[str] = set()
        for product in matches:
            product_name = str(product.get('name') or '').strip()
            if not product_name:
                continue
            name_key = product_name.lower()
            if name_key in seen_names:
                continue
            if not name_key.startswith(text_lower):
                continue
            seen_names.add(name_key)
            model.appendRow(QStandardItem(product_name))
        if model.rowCount() > 0:
            completer.complete()

    def _setup_name_completer(self) -> None:
        """Attach a debounced product-name suggestion list under the name field."""
        self._teardown_name_completer()
        try:
            completer = QCompleter(self)
            completer.setCaseSensitivity(Qt.CaseInsensitive)
            completer.setFilterMode(Qt.MatchStartsWith)
            completer.setCompletionMode(QCompleter.PopupCompletion)

            model = QStandardItemModel()
            completer.setModel(model)
            self._name_completer = completer
            self._name_completer_model = model

            self.name_input.textChanged.connect(self._on_name_completer_text_changed)
            self._name_completer_signals_connected = True
            self.name_input.setCompleter(completer)
            theme.apply_completer_popup_theme(completer)
            completer.activated[QModelIndex].connect(self._on_name_suggestion_selected)
        except Exception:
            self._teardown_name_completer()

    def _on_name_suggestion_selected(self, model_index: QModelIndex) -> None:
        """Fill only the suggested name for a new product — do not load an existing row."""
        try:
            selected_name = str(model_index.data(Qt.DisplayRole) or '').strip()
            if not selected_name:
                return
            self.name_input.blockSignals(True)
            self.name_input.setText(selected_name)
            self.name_input.blockSignals(False)
            # Stay in new-product mode so quantity applies to the new barcode only.
            self.current_product_id = None
            if hasattr(self, 'save_btn'):
                self.save_btn.setText('Save')
            if self.auto_barcode.isChecked():
                self.generate_sequential_barcode()
            field_order = self._get_enter_field_order()
            if self.name_input in field_order:
                name_index = field_order.index(self.name_input)
                if name_index < len(field_order) - 1:
                    QTimer.singleShot(0, lambda: self.focus_and_force_select(field_order[name_index + 1]))
        except Exception:
            pass

    def show_entry_page(self, clear_form=True):
        """Switch to Product Entry page."""
        self.stack_widget.setCurrentWidget(self.entry_page)
        self._apply_nav_styles('entry')
        if clear_form:
            self.clear_form()

    def show_list_page(self):
        """Switch to Product List page."""
        self.stack_widget.setCurrentWidget(self.list_page)
        self._apply_nav_styles('list')
        self.load_products()

    def label(self, text):
        lbl = QLabel(text)
        lbl.setStyleSheet(theme.master_label_style())
        return lbl

    def create_entry_page(self):
        container = QFrame()
        container.setStyleSheet(theme.master_panel_frame_style())
        outer_layout = QVBoxLayout(container)
        outer_layout.setContentsMargins(16, 14, 16, 14)
        outer_layout.setSpacing(8)
        input_style = theme.sales_compact_input_style()
        label_style = theme.sales_micro_label_style()
        section_style = theme.master_form_section_divider_style()
        lc = theme.legacy_colors()
        percent_label_style = f"\n            QLabel {{\n                color: {lc['success']};\n                font-size: 11px;\n                font-weight: 600;\n                background: transparent;\n                border: none;\n                padding: 0px;\n                margin: 0px;\n                min-width: 44px;\n                max-width: 44px;\n            }}\n        "
        LABEL_W = 110
        INPUT_W = 150
        WIDE_W = 360
        FIELD_H = 26
        left_align = Qt.AlignLeft | Qt.AlignVCenter

        def make_label(text):
            lbl = QLabel(text)
            lbl.setStyleSheet(label_style)
            return lbl

        def make_section(text):
            lbl = QLabel(text)
            lbl.setStyleSheet(section_style)
            return lbl

        def make_input(width=INPUT_W):
            field = QLineEdit()
            field.setStyleSheet(input_style)
            field.setFixedHeight(FIELD_H)
            field.setFixedWidth(width)
            return field
        grid = QGridLayout()
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(8)
        grid.setColumnMinimumWidth(0, LABEL_W)
        grid.setColumnMinimumWidth(3, LABEL_W)
        grid.setColumnStretch(6, 1)
        row = 0
        self.name_input = make_input(WIDE_W)
        self.name_input.textChanged.connect(lambda text: self.on_text_changed(self.name_input, text))
        grid.addWidget(make_label('Product Name *'), row, 0, left_align)
        grid.addWidget(self.name_input, row, 1, 1, 5, left_align)
        row += 1
        self.barcode_input = make_input()
        self.auto_barcode = create_checkbox('Auto', variant='compact')
        self.auto_barcode.setChecked(True)
        self.auto_barcode.toggled.connect(self.auto_barcode_toggle)
        grid.addWidget(make_label('Barcode'), row, 0, left_align)
        grid.addWidget(self.barcode_input, row, 1, left_align)
        grid.addWidget(self.auto_barcode, row, 2, 1, 2, left_align)
        row += 1
        self.hsn_input = make_input()
        self.hsn_input.textChanged.connect(lambda text: self.on_text_changed('hsn', text))
        grid.addWidget(make_label('HSN'), row, 0, left_align)
        grid.addWidget(self.hsn_input, row, 1, left_align)
        row += 1
        self.color_input = make_input()
        self.color_input.textChanged.connect(lambda text: self.on_text_changed(self.color_input, text))
        self.size_input = make_input()
        self.size_input.textChanged.connect(lambda text: self.on_text_changed(self.size_input, text))
        grid.addWidget(make_label('Color'), row, 0, left_align)
        grid.addWidget(self.color_input, row, 1, left_align)
        grid.addWidget(make_label('Size'), row, 3, left_align)
        grid.addWidget(self.size_input, row, 4, left_align)
        row += 1
        self.unit = QComboBox()
        self.unit.addItems(['pcs', 'kg', 'ltr'])
        self.unit.setStyleSheet(input_style)
        self.unit.setFixedHeight(FIELD_H)
        self.unit.setFixedWidth(INPUT_W)
        self.category = QComboBox()
        self.category.setEditable(True)
        self.category.setStyleSheet(input_style)
        self.category.setFixedHeight(FIELD_H)
        self.category.setFixedWidth(INPUT_W)
        category_line_edit = self.category.lineEdit()
        if category_line_edit is not None:
            category_line_edit.textChanged.connect(
                lambda text: self.on_text_changed(category_line_edit, text)
            )
        grid.addWidget(make_label('Unit'), row, 0, left_align)
        grid.addWidget(self.unit, row, 1, left_align)
        grid.addWidget(make_label('Category'), row, 3, left_align)
        grid.addWidget(self.category, row, 4, left_align)
        row += 1
        grid.addWidget(make_section('Rates'), row, 0, 1, 6)
        row += 1
        self.purchase_rate = make_input()
        self.purchase_rate.textChanged.connect(lambda: self.update_percentage_labels())
        self.sale_price = make_input()
        self.sale_price.textChanged.connect(lambda: self.update_percentage_labels())
        self.sale_price_percent = QLabel('')
        self.sale_price_percent.setStyleSheet(percent_label_style)
        self.sale_price_percent.setAlignment(left_align)
        grid.addWidget(make_label('Purchase Rate'), row, 0, left_align)
        grid.addWidget(self.purchase_rate, row, 1, left_align)
        grid.addWidget(make_label('Sale Price'), row, 3, left_align)
        grid.addWidget(self.sale_price, row, 4, left_align)
        grid.addWidget(self.sale_price_percent, row, 5, left_align)
        row += 1
        self.wholesale_rate = make_input()
        self.wholesale_rate.textChanged.connect(lambda: self.update_percentage_labels())
        self.wholesale_rate_percent = QLabel('')
        self.wholesale_rate_percent.setStyleSheet(percent_label_style)
        self.wholesale_rate_percent.setAlignment(left_align)
        self.mrp = make_input()
        self.mrp.textChanged.connect(lambda: self.update_percentage_labels())
        self.mrp_percent = QLabel('')
        self.mrp_percent.setStyleSheet(percent_label_style)
        self.mrp_percent.setAlignment(left_align)
        grid.addWidget(make_label('Wholesale Rate'), row, 0, left_align)
        grid.addWidget(self.wholesale_rate, row, 1, left_align)
        grid.addWidget(self.wholesale_rate_percent, row, 2, left_align)
        grid.addWidget(make_label('MRP'), row, 3, left_align)
        grid.addWidget(self.mrp, row, 4, left_align)
        grid.addWidget(self.mrp_percent, row, 5, left_align)
        row += 1
        grid.addWidget(make_section('Taxes'), row, 0, 1, 6)
        row += 1
        self.cgst = make_input()
        self.sgst = make_input()
        grid.addWidget(make_label('CGST'), row, 0, left_align)
        grid.addWidget(self.cgst, row, 1, left_align)
        grid.addWidget(make_label('SGST'), row, 3, left_align)
        grid.addWidget(self.sgst, row, 4, left_align)
        row += 1
        self.igst = make_input()
        self.cess = make_input()
        grid.addWidget(make_label('IGST'), row, 0, left_align)
        grid.addWidget(self.igst, row, 1, left_align)
        grid.addWidget(make_label('CESS'), row, 3, left_align)
        grid.addWidget(self.cess, row, 4, left_align)
        row += 1
        grid.addWidget(make_section('Stock'), row, 0, 1, 6)
        row += 1
        self.reorder_level = make_input()
        self.qty = make_input()
        grid.addWidget(make_label('Reorder Level'), row, 0, left_align)
        grid.addWidget(self.reorder_level, row, 1, left_align)
        grid.addWidget(make_label('Quantity'), row, 3, left_align)
        grid.addWidget(self.qty, row, 4, left_align)
        row += 1
        self.description = make_input(WIDE_W)
        self.description.textChanged.connect(lambda text: self.on_text_changed(self.description, text))
        grid.addWidget(make_label('Description'), row, 0, left_align)
        grid.addWidget(self.description, row, 1, 1, 5, left_align)
        row += 1
        outer_layout.addLayout(grid)
        outer_layout.addSpacing(6)
        actions_row = QHBoxLayout()
        actions_row.setSpacing(8)
        self.save_btn = QPushButton('Save')
        self.save_btn.setObjectName('save_btn')
        self.save_btn.clicked.connect(self.save)
        self.save_btn.setStyleSheet(theme.master_save_button_style())
        clear_btn = QPushButton('Clear')
        clear_btn.setStyleSheet(theme.master_clear_button_style())
        clear_btn.clicked.connect(self.clear_form)
        actions_row.addWidget(self.save_btn)
        actions_row.addWidget(clear_btn)
        actions_row.addStretch()
        outer_layout.addLayout(actions_row)
        outer_layout.addStretch()
        self._setup_entry_keyboard_navigation()
        return container

    def create_list_page(self):
        container = QFrame()
        container.setObjectName('productListOuterFrame')
        container.setStyleSheet(theme.master_panel_frame_style('productListOuterFrame'))
        layout = QVBoxLayout(container)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        list_title = QLabel('Product List')
        list_title.setStyleSheet(theme.master_page_title_style(18))
        layout.addWidget(list_title)
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(0, 0, 0, 10)
        search_label = QLabel('Search:')
        search_label.setStyleSheet(theme.master_label_style())
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText('Search by product name or barcode...')
        self.search_input.setStyleSheet(theme.sales_compact_input_style())
        self.search_input.textChanged.connect(self.filter_products)
        export_btn = QPushButton('Export')
        export_btn.setStyleSheet(theme.master_primary_action_button_style('8px 16px', 14))
        export_btn.clicked.connect(self.show_export_menu)
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(export_btn)
        search_layout.addStretch()
        layout.addLayout(search_layout)
        table_container = QFrame()
        table_container.setObjectName('productListTableContainer')
        table_container.setStyleSheet(theme.master_panel_frame_style('productListTableContainer'))
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.setSpacing(0)
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(['SL No', 'Product Name', 'Barcode', 'Purchase Rate', 'Sale Rate', 'Wholesale Rate', 'Quantity'])
        apply_read_only_report_table_selection(self.table)
        self.table.itemSelectionChanged.connect(self.on_table_selection_changed)
        self.table.itemDoubleClicked.connect(self.on_table_double_click)
        self.table.setCornerButtonEnabled(False)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_product_list_context_menu)
        self.table.setFrameShape(QFrame.NoFrame)
        self.table.setFrameShadow(QFrame.Plain)
        self.table.setLineWidth(0)
        self.table.setMidLineWidth(0)
        self.table.setContentsMargins(0, 0, 0, 0)
        self.table.setViewportMargins(0, 0, 0, 0)
        self.table.setStyleSheet(theme.master_table_style())
        header = self.table.horizontalHeader()
        header.setVisible(True)
        header.setFixedHeight(36)
        header.setMinimumHeight(36)
        header.setDefaultSectionSize(36)
        header.setHighlightSections(False)
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setStretchLastSection(False)
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setMinimumSectionSize(60)
        self.table.setColumnWidth(0, 90)
        self.table.setColumnWidth(1, 330)
        self.table.setColumnWidth(2, 130)
        self.table.setColumnWidth(3, 170)
        self.table.setColumnWidth(4, 160)
        self.table.setColumnWidth(5, 180)
        self.table.setColumnWidth(6, 130)
        table_layout.addWidget(self.table)
        layout.addWidget(table_container)
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 10, 0, 0)
        edit_btn = QPushButton('Edit Selected')
        edit_btn.setStyleSheet(theme.master_nav_primary_button_style())
        edit_btn.clicked.connect(self.edit_selected_product)
        delete_btn = QPushButton('Delete Selected')
        delete_btn.setStyleSheet(theme.master_danger_action_button_style())
        delete_btn.clicked.connect(self.delete_selected_product)
        button_layout.addWidget(edit_btn)
        button_layout.addWidget(delete_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)
        return container

    def auto_barcode_toggle(self, checked):
        if checked:
            self.generate_sequential_barcode()
            self.barcode_input.setEnabled(False)
            self.barcode_input.setReadOnly(True)
        else:
            self.barcode_input.setEnabled(True)
            self.barcode_input.setReadOnly(False)
            self.barcode_input.clear()

    def generate_sequential_barcode(self):
        """Generate sequential barcode based on existing products for active company."""
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                current_text = self.barcode_input.text()
                if current_text and current_text.isdigit():
                    next_barcode = str(int(current_text) + 1)
                else:
                    next_barcode = '1'
                self.barcode_input.setText(next_barcode)
                return
            next_barcode = self.product_logic.generate_sequential_barcode(active_company['id'])
            self.barcode_input.setText(next_barcode)
        except Exception as e:
            current_text = self.barcode_input.text()
            if current_text and current_text.isdigit():
                next_barcode = str(int(current_text) + 1)
            else:
                next_barcode = str(random.randint(100000, 999999))
            self.barcode_input.setText(next_barcode)

    def get_next_unique_barcode(self, exclude_current_id=None):
        """Get the next unique barcode for the active company."""
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                return '1'
            return self.product_logic.generate_next_barcode(active_company['id'], exclude_current_id)
        except Exception:
            return str(int(time.time() * 1000) % 1000000)

    def capitalize_first_letter(self, text):
        """Capitalize first letter of text without disturbing cursor position."""
        if not text:
            return text
        if len(text) > 0 and text[0].isalpha() and text[0].islower():
            return text[0].upper() + text[1:]
        return text

    def apply_capitalization_to_field(self, field, text):
        """Apply capitalization to field text."""
        if text:
            capitalized_text = self.capitalize_first_letter(text)
            if capitalized_text != text:
                cursor_pos = field.cursorPosition()
                field.blockSignals(True)
                field.setText(capitalized_text)
                field.setCursorPosition(cursor_pos)
                field.blockSignals(False)

    def initialize_field_capitalization(self):
        """Apply capitalization to all relevant fields with existing text."""
        fields_to_capitalize = [
            (self.name_input, 'name'),
            (self.hsn_input, 'hsn'),
            (self.color_input, 'color'),
            (self.size_input, 'size'),
            (self.category.lineEdit(), 'category'),
            (self.description, 'description'),
        ]
        for field, field_name in fields_to_capitalize:
            if field is None:
                continue
            current_text = field.text()
            if current_text:
                capitalized = self.capitalize_first_letter(current_text)
                if capitalized != current_text:
                    if field is self.category.lineEdit():
                        self.category.setCurrentText(capitalized)
                    else:
                        field.setText(capitalized)

    def safe_calculate_expression(self, expression):
        """Safely calculate simple arithmetic expression."""
        try:
            expr = expression.replace(' ', '')
            if not all((c in '0123456789+-*/.' for c in expr)):
                return None
            if not expr or expr[0] in '+-*/' or expr[-1] in '+-*/':
                return None
            for i in range(len(expr) - 1):
                if expr[i] in '+-*/' and expr[i + 1] in '+-*/':
                    return None
            result = self._evaluate_simple_expression(expr)
            return result
        except Exception:
            return None

    def _evaluate_simple_expression(self, expr):
        """Evaluate simple arithmetic expression step by step."""
        try:
            parts = []
            current = ''
            i = 0
            while i < len(expr):
                if expr[i] in '+-*/':
                    if current:
                        parts.append(current)
                    parts.append(expr[i])
                    current = ''
                    i += 1
                else:
                    current += expr[i]
                    i += 1
            if current:
                parts.append(current)
            i = 0
            while i < len(parts):
                if parts[i] == '*':
                    left = float(parts[i - 1])
                    right = float(parts[i + 1])
                    result = left * right
                    parts[i - 1:i + 2] = [str(result)]
                    i -= 1
                elif parts[i] == '/':
                    left = float(parts[i - 1])
                    right = float(parts[i + 1])
                    if right == 0:
                        return None
                    result = left / right
                    parts[i - 1:i + 2] = [str(result)]
                    i -= 1
                i += 1
            result = float(parts[0])
            i = 1
            while i < len(parts):
                if parts[i] == '+':
                    result += float(parts[i + 1])
                elif parts[i] == '-':
                    result -= float(parts[i + 1])
                i += 2
            return result
        except Exception:
            return None

    def format_calculation_result(self, result):
        """Format calculation result for display."""
        if result is None:
            return None
        if result == int(result):
            return str(int(result))
        else:
            return str(round(result, 2)).rstrip('0').rstrip('.')

    def handle_calculator_field(self, field, next_field=None):
        """Handle calculator behavior for a field."""
        text = field.text().strip()
        if not text:
            field.setText('0')
            return True
        result = self.safe_calculate_expression(text)
        if result is not None:
            formatted_result = self.format_calculation_result(result)
            field.setText(formatted_result)
            return True
        elif text.replace('.', '').replace('-', '').isdigit():
            return True
        else:
            return False

    def on_text_changed(self, widget, text):
        """Handle text change to capitalize first letter for relevant fields."""
        if widget and hasattr(widget, 'cursorPosition'):
            cursor_pos = widget.cursorPosition()
            capitalized_text = self.capitalize_first_letter(text)
            if capitalized_text != text:
                widget.blockSignals(True)
                widget.setText(capitalized_text)
                widget.setCursorPosition(cursor_pos)
                widget.blockSignals(False)

    def clear_form(self):
        self.name_input.clear()
        self.hsn_input.clear()
        self.color_input.clear()
        self.size_input.clear()
        self._load_category_options(select_text='')
        self.purchase_rate.clear()
        self.sale_price.clear()
        self.wholesale_rate.clear()
        self.mrp.clear()
        self.cgst.clear()
        self.sgst.clear()
        self.igst.clear()
        self.cess.clear()
        self.reorder_level.clear()
        self.description.clear()
        self.qty.clear()
        self.sale_price_percent.setText('')
        self.wholesale_rate_percent.setText('')
        self.mrp_percent.setText('')
        self.initialize_field_capitalization()
        self.auto_barcode.setChecked(True)
        self.barcode_input.setEnabled(False)
        self.barcode_input.setReadOnly(True)
        self.generate_sequential_barcode()
        self.current_product_id = None
        if hasattr(self, 'save_btn'):
            self.save_btn.setText('Save')

    def save(self):
        active_company = active_company_manager.get_active_company()
        if not active_company:
            QMessageBox.warning(self, 'No Active Company', 'Please open a company first.')
            return
        name = self.name_input.text().strip()
        if not name:
            QMessageBox.warning(self, 'Validation Error', 'Product/Service Name is required.')
            self.name_input.setFocus()
            return
        try:
            auto_barcode = self.auto_barcode.isChecked()
            if auto_barcode:
                if not self.current_product_id:
                    barcode = self.get_next_unique_barcode()
                else:
                    barcode = self.barcode_input.text().strip() or self.get_next_unique_barcode(self.current_product_id)
            else:
                barcode = self.barcode_input.text().strip()
                if not barcode:
                    QMessageBox.warning(self, 'Validation Error', 'Please enter a barcode or enable Auto Barcode.')
                    self.barcode_input.setFocus()
                    return
            product_data = {'name': name, 'barcode': barcode, 'hsn': self.hsn_input.text().strip(), 'color': self.color_input.text().strip(), 'size': self.size_input.text().strip(), 'unit': self.unit.currentText().strip(), 'category': self.category.currentText().strip(), 'purchase_rate': self.purchase_rate.text() or '0', 'sale_price': self.sale_price.text() or '0', 'wholesale_rate': self.wholesale_rate.text() or '0', 'mrp': self.mrp.text() or '0', 'cgst': self.cgst.text() or '0', 'sgst': self.sgst.text() or '0', 'igst': self.igst.text() or '0', 'cess': self.cess.text() or '0', 'reorder_level': self.reorder_level.text() or '0', 'description': self.description.text().strip(), 'quantity': self.qty.text() or '0', 'auto_barcode': auto_barcode}
            validation_result = self.product_logic.validate_product_data(
                product_data,
                auto_barcode,
                self.current_product_id,
                active_company['id'],
                allow_duplicate_name=bool(self.page_settings.get('allow_duplicate')),
            )
            if not validation_result['success']:
                QMessageBox.warning(self, 'Validation Error', validation_result['message'])
                if 'barcode' in validation_result['message'].lower():
                    self.barcode_input.setFocus()
                return
            if self.source_context == 'purchase_entry_edit':
                qty_value = product_data.get('quantity', '0')
                if not qty_value or qty_value == '0':
                    QMessageBox.warning(self, 'Validation Error', 'Please enter a quantity.')
                    self.qty.setFocus()
                    return
            if self.source_context == 'purchase_entry_edit':
                reply = QMessageBox.question(self, 'Save Product', 'Save product?', QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                if reply != QMessageBox.Yes:
                    return
            skip_opening_stock = self.source_context == 'purchase_entry'
            save_result = self.product_logic.save_product(active_company['id'], product_data, self.current_product_id, skip_opening_stock=skip_opening_stock)
            if save_result['success']:
                if self.source_context == 'purchase_entry_edit':
                    saved_product_id = save_result.get('data', {}).get('id', self.current_product_id)
                    self._return_to_purchase_entry(saved_product_id, product_data.get('quantity', '0'))
                    return
                if self.source_context == 'purchase_entry' and self.source_purchase_widget:
                    saved_product_id = save_result.get('data', {}).get('id')
                    if not saved_product_id:
                        saved_product_id = save_result.get('data')
                    if saved_product_id:
                        self._return_to_purchase_entry(saved_product_id, product_data.get('quantity', '0'))
                    return
                QMessageBox.information(self, 'Success', 'Product/Service saved successfully.')
                self.clear_form()
                QTimer.singleShot(0, lambda: self.name_input.setFocus())
                self.load_products()
                search_term = self.search_input.text().strip()
                if search_term:
                    self.filter_products(search_term)
            else:
                QMessageBox.warning(self, 'Error', save_result['message'])
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to save product: {str(e)}')

    def load_products(self):
        """Load all products from database into memory."""
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                self.products_data = []
                self.visible_products = []
                self.render_products([])
                return
            result = self.product_logic.get_products(active_company['id'])
            if result['success']:
                self.products_data = result['data']
                self.visible_products = self.products_data
                self.render_products(self.products_data)
                self._load_category_options()
            else:
                QMessageBox.critical(self, 'Error', result['message'])
                self.products_data = []
                self.visible_products = []
                self.render_products([])
                self._load_category_options()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load products: {str(e)}')
            self.products_data = []
            self.visible_products = []
            self.render_products([])
            self._load_category_options()

    def _load_category_options(self, select_text: str | None = None) -> None:
        """Populate the category dropdown from existing product categories."""
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                categories = []
            else:
                result = self.product_logic.get_product_categories(active_company['id'])
                categories = result.get('data', []) if result.get('success') else []
        except Exception:
            categories = []

        current = select_text if select_text is not None else self.category.currentText().strip()
        self.category.blockSignals(True)
        try:
            self.category.clear()
            self.category.addItem('')
            for value in categories:
                if value and self.category.findText(value, Qt.MatchFixedString) < 0:
                    self.category.addItem(value)
            if current:
                match_index = self.category.findText(current, Qt.MatchFixedString)
                if match_index >= 0:
                    self.category.setCurrentIndex(match_index)
                else:
                    self.category.setEditText(current)
            else:
                self.category.setCurrentIndex(0)
        finally:
            self.category.blockSignals(False)

    def render_products(self, products):
        """Render products in table without heading row."""
        self.table.setRowCount(len(products))
        for row, product in enumerate(products):
            sl_no_item = QTableWidgetItem(str(row + 1))
            sl_no_item.setData(Qt.UserRole, product['id'])
            self.table.setItem(row, 0, sl_no_item)
            name_item = QTableWidgetItem(product['name'])
            name_item.setData(Qt.UserRole, product['id'])
            self.table.setItem(row, 1, name_item)
            barcode_item = QTableWidgetItem(product['barcode'] or '')
            barcode_item.setData(Qt.UserRole, product['id'])
            self.table.setItem(row, 2, barcode_item)
            purchase_item = QTableWidgetItem(f"{float(product['purchase_rate']):.2f}")
            purchase_item.setData(Qt.UserRole, product['id'])
            self.table.setItem(row, 3, purchase_item)
            sale_item = QTableWidgetItem(f"{float(product['sale_price']):.2f}")
            sale_item.setData(Qt.UserRole, product['id'])
            self.table.setItem(row, 4, sale_item)
            wholesale_item = QTableWidgetItem(f"{float(product['wholesale_rate']):.2f}")
            wholesale_item.setData(Qt.UserRole, product['id'])
            self.table.setItem(row, 5, wholesale_item)
            quantity = max(0, float(product['quantity']))
            quantity_item = QTableWidgetItem(f'{quantity:.2f}')
            quantity_item.setData(Qt.UserRole, product['id'])
            self.table.setItem(row, 6, quantity_item)

    def filter_products(self, search_term):
        """Filter products in memory based on search term."""
        search_term = search_term.strip()
        if not search_term:
            self.visible_products = self.products_data
            self.render_products(self.products_data)
            return
        filtered_products = self.product_logic.filter_products(self.products_data, search_term)
        self.visible_products = filtered_products
        self.render_products(filtered_products)

    def on_table_selection_changed(self):
        """Handle table row selection change - normal selection for product rows."""
        pass

    def on_table_double_click(self, item):
        """Handle double-click on table row to edit product."""
        self.edit_selected_product()

    def edit_selected_product(self):
        """Edit the selected product by switching to entry page."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, 'No Selection', 'Please select a product to edit.')
            return
        product_id = selected_items[0].data(Qt.UserRole)
        if not product_id:
            QMessageBox.warning(self, 'Error', 'Unable to identify selected product.')
            return
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                return
            result = self.product_logic.get_product_by_id(active_company['id'], product_id)
            if result['success'] and result['data']:
                self.load_product_to_form(result['data'])
                self.show_entry_page(clear_form=False)
            else:
                QMessageBox.warning(self, 'Error', 'Product not found.')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load product: {str(e)}')

    def delete_selected_product(self):
        """Delete the selected product."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, 'No Selection', 'Please select a product to delete.')
            return
        product_id = selected_items[0].data(Qt.UserRole)
        selected_row = self.table.currentRow()
        product_name_item = self.table.item(selected_row, 1)
        product_name = product_name_item.text() if product_name_item else 'selected product'
        if not product_id:
            QMessageBox.warning(self, 'Error', 'Unable to identify selected product.')
            return
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                return
            reply = QMessageBox.question(self, 'Confirm Delete', f"Are you sure you want to delete '{product_name}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                result = self.product_logic.delete_product(active_company['id'], product_id)
                if result['success']:
                    QMessageBox.information(self, 'Success', 'Product deleted successfully.')
                    self.load_products()
                    search_term = self.search_input.text().strip()
                    if search_term:
                        self.filter_products(search_term)
                else:
                    QMessageBox.warning(self, 'Error', result['message'])
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to delete product: {str(e)}')

    def load_product_to_form(self, product):
        """Load product data into form fields."""
        self.current_product_id = product['id']
        self.name_input.setText(product['name'])
        self.barcode_input.setText(product['barcode'] or '')
        self.hsn_input.setText(product['hsn'] or '')
        self.color_input.setText(product['color'] or '')
        self.size_input.setText(product['size'] or '')
        self.unit.setCurrentText(product['unit'] or 'pcs')
        self._load_category_options(select_text=product['category'] or '')
        self.purchase_rate.setText(str(product['purchase_rate']))
        self.sale_price.setText(str(product['sale_price']))
        self.wholesale_rate.setText(str(product['wholesale_rate']))
        self.mrp.setText(str(product['mrp']))
        self.cgst.setText(str(product['cgst']))
        self.sgst.setText(str(product['sgst']))
        self.igst.setText(str(product['igst']))
        self.cess.setText(str(product['cess']))
        self.reorder_level.setText(str(product['reorder_level']))
        self.description.setText(product['description'] or '')
        self.qty.setText(str(product['quantity']))
        auto_barcode = bool(product['auto_barcode'])
        self.auto_barcode.setChecked(auto_barcode)
        self.barcode_input.setEnabled(not auto_barcode)
        self.update_percentage_labels()
        if hasattr(self, 'save_btn'):
            self.save_btn.setText('Update')

    def load_product_for_edit(self, product_id):
        """Load product by ID for editing from Purchase Entry."""
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                return
            result = self.product_logic.get_product_by_id(active_company['id'], product_id)
            if result['success'] and result['data']:
                product = result['data']
                self.load_product_to_form(product)
                if hasattr(self, 'save_btn'):
                    self.save_btn.setText('Update')
                QTimer.singleShot(300, self._focus_and_select_product_name)
        except Exception as e:
            print(f'Error loading product for edit: {e}')

    def _focus_and_select_product_name(self):
        """Helper to focus and select product name field."""
        self.name_input.setReadOnly(False)
        self.name_input.setFocus()
        QTimer.singleShot(100, self._do_select_all)

    def _do_select_all(self):
        """Actually perform selectAll on product name."""
        self.name_input.setReadOnly(False)
        self.name_input.setFocus()
        text = self.name_input.text()
        if text:
            self.name_input.setSelection(0, len(text))
        else:
            self.name_input.selectAll()

    def open_for_new_from_purchase(self, purchase_widget, row, suggested_name=''):
        """Open Product Entry in new mode from Purchase Entry."""
        self.source_context = 'purchase_entry'
        self.source_purchase_widget = purchase_widget
        self.source_purchase_row = row
        self.clear_form()
        suggested_name = (suggested_name or '').strip()
        if suggested_name:
            self.name_input.setText(suggested_name)
        if hasattr(self, 'save_btn'):
            self.save_btn.setText('Save')
        QTimer.singleShot(0, lambda: self.name_input.setFocus())
        QTimer.singleShot(50, lambda: self.name_input.selectAll())
        self.show()
        self.raise_()
        self.activateWindow()

    def open_for_edit_from_purchase(self, product_id, purchase_widget=None, row=None):
        """Open Product Entry in edit mode from Purchase Entry (without qty)."""
        self.source_context = 'purchase_entry_edit'
        self.source_purchase_widget = purchase_widget
        self.source_purchase_row = row
        self.load_product_for_edit(product_id)
        if hasattr(self, 'qty'):
            self.qty.clear()
        if hasattr(self, 'qty'):
            self.qty.setFocus()
        self.show()
        self.raise_()
        self.activateWindow()

    def _finish_purchase_return(self, product_id, qty):
        """Helper to finish purchase return by switching back to Purchase Entry."""
        if not self.source_purchase_widget or self.source_purchase_row is None:
            return
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                return
            result = self.product_logic.get_product_by_id(active_company['id'], product_id)
            if result['success'] and result['data']:
                product = result['data']
                self.source_purchase_widget.receive_product_from_product_page(product, qty, self.source_purchase_row)
                purchase_window = self.source_purchase_widget.window()
                purchase_window.show()
                purchase_window.raise_()
                purchase_window.activateWindow()
                if self.window() != purchase_window:
                    self.window().close()
            else:
                QMessageBox.warning(self, 'Error', f"Failed to retrieve product (ID: {product_id}): {result.get('message', 'Unknown error')}")
                if self.source_purchase_widget:
                    purchase_window = self.source_purchase_widget.window()
                    purchase_window.show()
                    purchase_window.raise_()
                    purchase_window.activateWindow()
        except Exception as e:
            print(f'Error returning to purchase entry: {e}')
            if self.source_purchase_widget:
                purchase_window = self.source_purchase_widget.window()
                purchase_window.show()
                purchase_window.raise_()
                purchase_window.activateWindow()
        finally:
            self.source_context = None
            self.source_purchase_widget = None
            self.source_purchase_row = None

    def _return_to_purchase_entry(self, product_id, qty):
        """Callback to return to Purchase Entry with saved product data."""
        self._finish_purchase_return(product_id, qty)

    def delete_current_product(self):
        """Delete the current product being edited."""
        if not self.current_product_id:
            QMessageBox.warning(self, 'No Product', 'No product selected for deletion.')
            return
        reply = QMessageBox.question(self, 'Confirm Delete', 'Are you sure you want to delete this product?', QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                active_company = active_company_manager.get_active_company()
                if not active_company:
                    return
                result = self.product_logic.delete_product(active_company['id'], self.current_product_id)
                if result['success']:
                    QMessageBox.information(self, 'Success', 'Product deleted successfully.')
                    self.clear_form()
                    self.load_products()
                    search_term = self.search_input.text().strip()
                    if search_term:
                        self.filter_products(search_term)
                else:
                    QMessageBox.warning(self, 'Error', result['message'])
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Failed to delete product: {str(e)}')

    def _is_entry_page_active(self) -> bool:
        """Return True when the product entry form is the visible page."""
        return self.stack_widget.currentWidget() == self.entry_page

    def _setup_entry_keyboard_navigation(self) -> None:
        """Install event filters so Enter/Esc work while child fields hold focus."""
        widgets = set(self._get_field_widget_map().values())
        widgets.add(self.save_btn)
        self._entry_nav_widgets = widgets
        for widget in widgets:
            widget.installEventFilter(self)
        category_line_edit = self.category.lineEdit()
        if category_line_edit is not None:
            category_line_edit.installEventFilter(self)
            self._entry_nav_widgets.add(category_line_edit)

    def _normalize_entry_focus_widget(self, focus_widget):
        """Map editable combo line edits back to their parent combo widget."""
        category_line_edit = self.category.lineEdit()
        if category_line_edit is not None and focus_widget == category_line_edit:
            return self.category
        return focus_widget

    def _process_entry_navigation_key(self, focus_widget, key) -> bool:
        """Handle Enter, Esc, and Down navigation for the product entry form."""
        if not self._is_entry_page_active():
            return False
        if QApplication.activeModalWidget():
            return False

        focus_widget = self._normalize_entry_focus_widget(focus_widget)

        if key in (Qt.Key_Return, Qt.Key_Enter):
            if focus_widget == self.name_input and self._name_completer is not None:
                popup = self._name_completer.popup()
                if popup is not None and popup.isVisible():
                    return False
            field_order = self._get_enter_field_order()
            calculator_fields = [self.purchase_rate, self.sale_price, self.wholesale_rate, self.mrp, self.qty]
            if focus_widget in calculator_fields:
                if self.handle_calculator_field(focus_widget):
                    current_index = field_order.index(focus_widget)
                    if current_index < len(field_order) - 1:
                        next_field = field_order[current_index + 1]
                        self.focus_and_force_select(next_field)
                    else:
                        self.save()
                    return True
                return False
            if focus_widget in field_order:
                current_index = field_order.index(focus_widget)
                if current_index < len(field_order) - 1:
                    next_field = field_order[current_index + 1]
                    self.focus_and_force_select(next_field)
                else:
                    self.save()
                return True
            if focus_widget == self.save_btn:
                self.save()
                return True
            return False

        if key == Qt.Key_Down:
            percentage_fields = [self.sale_price, self.wholesale_rate, self.mrp]
            if focus_widget in percentage_fields:
                if self.handle_percentage_field(focus_widget):
                    self.focus_and_force_select(focus_widget)
                return True
            return False

        if key == Qt.Key_Escape:
            if isinstance(focus_widget, QComboBox):
                view = focus_widget.view()
                if view is not None and view.isVisible():
                    return False
            if focus_widget == self.save_btn:
                field_order = self._get_enter_field_order()
                if field_order:
                    self.focus_and_force_select(field_order[-1])
                return True
            field_order = self._get_enter_field_order()
            if focus_widget in field_order:
                current_index = field_order.index(focus_widget)
                if current_index > 0:
                    prev_field = field_order[current_index - 1]
                    if isinstance(prev_field, QTextEdit):
                        prev_field.setFocus()
                        prev_field.moveCursor(QTextCursor.End)
                    else:
                        self.focus_and_force_select(prev_field)
                    return True
            return False

        return False

    def eventFilter(self, obj, event):
        """Route Enter/Esc navigation from focused entry fields."""
        if (
            event.type() == QEvent.KeyPress
            and hasattr(self, '_entry_nav_widgets')
            and obj in self._entry_nav_widgets
            and self._process_entry_navigation_key(obj, event.key())
        ):
            return True
        return super().eventFilter(obj, event)

    def focus_and_force_select(self, widget):
        """Set focus and force select all text with proper timing."""
        widget.setFocus()
        if isinstance(widget, QComboBox):
            if widget.lineEdit():
                QTimer.singleShot(0, widget.lineEdit().selectAll)
        else:
            QTimer.singleShot(0, lambda: widget.setSelection(0, len(widget.text())))

    def focus_and_select(self, widget):
        """Set focus and select all text with proper timing."""
        widget.setFocus()
        QTimer.singleShot(0, widget.selectAll)

    def apply_percentage(self, base_value, percentage_text):
        """Apply percentage to base value."""
        try:
            if base_value:
                base = float(base_value)
            else:
                base = 0.0
            percentage = float(percentage_text)
            result = base + base * percentage / 100
            return result
        except (ValueError, TypeError):
            return None

    def handle_percentage_field(self, field):
        """Handle percentage calculation for rate fields."""
        text = field.text().strip()
        if not text:
            return False
        base_text = self.purchase_rate.text().strip()
        if not base_text:
            base_text = '0'
        result = self.apply_percentage(base_text, text)
        if result is not None:
            formatted_result = self.format_calculation_result(result)
            field.setText(formatted_result)
            if field == self.sale_price:
                self.sale_price_percent.setText(f'{text}%')
            elif field == self.wholesale_rate:
                self.wholesale_rate_percent.setText(f'{text}%')
            elif field == self.mrp:
                self.mrp_percent.setText(f'{text}%')
            return True
        else:
            return False

    def update_percentage_labels(self):
        """Update percentage labels based on current field values."""
        try:
            purchase_rate = float(self.purchase_rate.text() or '0')
            sale_price = float(self.sale_price.text() or '0')
            if purchase_rate > 0:
                sale_percent = (sale_price - purchase_rate) / purchase_rate * 100
                self.sale_price_percent.setText(f'{sale_percent:.1f}%')
            else:
                self.sale_price_percent.setText('')
            wholesale_rate = float(self.wholesale_rate.text() or '0')
            if purchase_rate > 0:
                wholesale_percent = (wholesale_rate - purchase_rate) / purchase_rate * 100
                self.wholesale_rate_percent.setText(f'{wholesale_percent:.1f}%')
            else:
                self.wholesale_rate_percent.setText('')
            mrp = float(self.mrp.text() or '0')
            if purchase_rate > 0:
                mrp_percent = (mrp - purchase_rate) / purchase_rate * 100
                self.mrp_percent.setText(f'{mrp_percent:.1f}%')
            else:
                self.mrp_percent.setText('')
        except (ValueError, TypeError):
            self.sale_price_percent.setText('')
            self.wholesale_rate_percent.setText('')
            self.mrp_percent.setText('')

    def keyPressEvent(self, event):
        """Handle key press events for Enter and Esc navigation."""
        focus_widget = self.focusWidget()
        if self._process_entry_navigation_key(focus_widget, event.key()):
            return
        super().keyPressEvent(event)

    def show_product_list_context_menu(self, position):
        """Show context menu for product list table."""
        menu = QMenu(self)
        select_all_action = menu.addAction('Select All')
        select_all_action.triggered.connect(self.select_all_visible_rows)
        clear_selection_action = menu.addAction('Clear Selection')
        clear_selection_action.triggered.connect(self.clear_selection)
        menu.exec(self.table.mapToGlobal(position))

    def select_all_visible_rows(self):
        """Select all visible rows in the table."""
        self.table.selectAll()

    def clear_selection(self):
        """Clear all selections in the table."""
        self.table.clearSelection()

    def get_visible_table_data(self):
        """Extract visible table data for export."""
        data = []
        headers = []
        for col in range(self.table.columnCount()):
            headers.append(self.table.horizontalHeaderItem(col).text())
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else '')
            data.append(row_data)
        return (headers, data)

    def show_export_menu(self):
        """Show export format selection dialog."""
        if not self.visible_products:
            QMessageBox.warning(self, 'No Data', 'No products to export.')
            return
        dialog = QDialog(self)
        dialog.setWindowTitle('Export Product List')
        dialog.setStyleSheet(report_detail_dialog_style())
        dialog.setFixedSize(350, 150)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        title_label = QLabel('Select Export Format:')
        title_label.setStyleSheet(theme.master_dialog_heading_style())
        layout.addWidget(title_label)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        pdf_btn = QPushButton('PDF')
        pdf_btn.setCheckable(True)
        pdf_btn.setEnabled(self.pdf_available)
        pdf_btn.clicked.connect(lambda: self.export_to_pdf(dialog))
        excel_btn = QPushButton('Excel')
        excel_btn.setCheckable(True)
        excel_btn.setEnabled(self.excel_available)
        excel_btn.clicked.connect(lambda: self.export_to_excel(dialog))
        cancel_btn = QPushButton('Cancel')
        cancel_btn.clicked.connect(dialog.reject)
        cancel_btn.setStyleSheet(theme.master_clear_button_style())
        button_layout.addWidget(pdf_btn)
        button_layout.addWidget(excel_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        dialog.exec()

    def export_to_pdf(self, dialog=None):
        """Export visible products to PDF."""
        if not self.pdf_available:
            QMessageBox.warning(self, 'Library Not Installed', 'Required library not installed.\n\nTo enable PDF export, install:\npip install reportlab')
            return
        try:
            headers, data = self.get_visible_table_data()
            if not data:
                QMessageBox.warning(self, 'No Data', 'No data available to export.')
                return
            file_path, _ = QFileDialog.getSaveFileName(self, 'Save PDF', 'product_list.pdf', 'PDF Files (*.pdf)')
            if not file_path:
                return
            table_data = [headers] + data
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            table = Table(table_data)
            table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -1), colors.white), ('TEXTCOLOR', (0, 1), (-1, -1), colors.black), ('GRID', (0, 0), (-1, -1), 1, colors.black), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])]))
            doc.build([table])
            QMessageBox.information(self, 'Success', 'PDF exported successfully.')
            if dialog:
                dialog.accept()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to export PDF: {str(e)}')

    def export_to_excel(self, dialog=None):
        """Export visible products to Excel."""
        if not self.excel_available:
            QMessageBox.warning(self, 'Library Not Installed', 'Required library not installed.\n\nTo enable Excel export, install:\npip install openpyxl')
            return
        try:
            file_path, _ = QFileDialog.getSaveFileName(self, 'Save Excel', 'product_list.xlsx', 'Excel Files (*.xlsx)')
            if not file_path:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = 'Product List'
            headers = ['SL No', 'Product Name', 'Barcode', 'Purchase Rate', 'Sale Rate', 'Wholesale Rate', 'Quantity']
            ws.append(headers)
            header_fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            border = Border(left=Side(style='thin', color='4b5563'), right=Side(style='thin', color='4b5563'), top=Side(style='thin', color='4b5563'), bottom=Side(style='thin', color='4b5563'))
            for col in range(1, 8):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            for idx, product in enumerate(self.visible_products):
                row = idx + 2
                ws.cell(row=row, column=1, value=idx + 1)
                ws.cell(row=row, column=2, value=product['name'] or '')
                ws.cell(row=row, column=3, value=product['barcode'] or '')
                ws.cell(row=row, column=4, value=float(product['purchase_rate'] or 0))
                ws.cell(row=row, column=5, value=float(product['sale_price'] or 0))
                ws.cell(row=row, column=6, value=float(product['wholesale_rate'] or 0))
                ws.cell(row=row, column=7, value=float(product['quantity'] or 0))
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if col in [4, 5, 6, 7]:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 12
            wb.save(file_path)
            QMessageBox.information(self, 'Success', 'Excel file exported successfully.')
            if dialog:
                dialog.accept()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to export Excel: {str(e)}')