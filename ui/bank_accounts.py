"""
Bank Account widget for the Accounting Desktop Application.
Manages bank accounts with company-wise data storage.
"""
from PySide6.QtWidgets import *
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QTextCursor
from config import active_company_manager
from db import Database
from bizora_core.bank_account_logic import BankAccountLogic
from ui import theme
from ui.book_report_common import report_detail_dialog_style
PDF_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
try:
    from docx import Document
    from docx.shared import Pt
    WORD_AVAILABLE = True
except ImportError:
    WORD_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
import datetime
from ui.date_formats import configure_qdate_edit, format_display_date, qdate_to_db, qdate_to_display, db_to_qdate
from ui.ui_memory import UiMemoryMixin

class BankAccountWidget(UiMemoryMixin, QWidget):

    def __init__(self, db=None):
        super().__init__()
        self.db = db or Database()
        self.bank_account_logic = BankAccountLogic(self.db)
        self.current_bank_account_id = None
        self.bank_accounts_data = []
        self.visible_bank_accounts = []
        self.pdf_available = PDF_AVAILABLE
        self.word_available = WORD_AVAILABLE
        self.excel_available = EXCEL_AVAILABLE
        self.setup_ui()
        self.load_bank_accounts()
        self.clear_form()
        self._init_ui_memory()

    def setup_ui(self):
        self.setStyleSheet(theme.master_page_background_style())
        layout = QVBoxLayout(self)
        title = QLabel('Bank Account')
        title.setStyleSheet(theme.master_page_title_style(24))
        layout.addWidget(title)
        nav_layout = QHBoxLayout()
        nav_layout.setContentsMargins(0, 10, 0, 10)
        self.entry_btn = QPushButton('Bank Account Entry')
        self.entry_btn.setStyleSheet(theme.master_nav_primary_button_style())
        self.entry_btn.clicked.connect(self.show_entry_page)
        self.list_btn = QPushButton('Bank Account List')
        self.list_btn.setStyleSheet(theme.master_nav_secondary_button_style())
        self.list_btn.clicked.connect(self.show_list_page)
        nav_layout.addWidget(self.entry_btn)
        nav_layout.addWidget(self.list_btn)
        nav_layout.addStretch()
        layout.addLayout(nav_layout)
        self.stack_widget = QStackedWidget()
        self.entry_page = self.create_entry_page()
        self.list_page = self.create_list_page()
        self.stack_widget.addWidget(self.entry_page)
        self.stack_widget.addWidget(self.list_page)
        layout.addWidget(self.stack_widget)

    def _apply_nav_styles(self, active: str) -> None:
        self.entry_btn.setStyleSheet(theme.master_nav_primary_button_style() if active == 'entry' else theme.master_nav_secondary_button_style())
        self.list_btn.setStyleSheet(theme.master_nav_primary_button_style() if active == 'list' else theme.master_nav_secondary_button_style())

    def show_entry_page(self, clear_form=True):
        """Switch to Bank Account Entry page."""
        self.stack_widget.setCurrentWidget(self.entry_page)
        self._apply_nav_styles('entry')
        if clear_form:
            self.clear_form()

    def show_list_page(self):
        """Switch to Bank Account List page."""
        self.stack_widget.setCurrentWidget(self.list_page)
        self._apply_nav_styles('list')
        self.load_bank_accounts()

    def label(self, text):
        lbl = QLabel(text)
        lbl.setStyleSheet(theme.master_label_style())
        return lbl

    def create_entry_page(self):
        container = QFrame()
        container.setStyleSheet(theme.master_panel_frame_style())
        layout = QVBoxLayout(container)
        layout.setContentsMargins(10, 5, 10, 10)
        layout.setSpacing(3)
        compact_input_style = theme.master_input_style()
        wide_input_style = theme.master_wide_input_style()
        extra_wide_input_style = theme.master_extra_wide_input_style()
        compact_label_style = theme.master_label_style()
        account_name_label = QLabel('Account Name *')
        account_name_label.setStyleSheet(compact_label_style)
        account_name_label.setFixedWidth(100)
        self.account_name_input = QLineEdit()
        self.account_name_input.setStyleSheet(extra_wide_input_style)
        self.account_name_input.textChanged.connect(lambda text: self.on_text_changed(self.account_name_input, text))
        layout.addWidget(account_name_label)
        layout.addWidget(self.account_name_input)
        layout.addSpacing(3)
        bank_name_label = QLabel('Bank Name *')
        bank_name_label.setStyleSheet(compact_label_style)
        bank_name_label.setFixedWidth(100)
        self.bank_name_input = QLineEdit()
        self.bank_name_input.setStyleSheet(extra_wide_input_style)
        self.bank_name_input.textChanged.connect(lambda text: self.on_text_changed(self.bank_name_input, text))
        layout.addWidget(bank_name_label)
        layout.addWidget(self.bank_name_input)
        layout.addSpacing(3)
        account_ifsc_row = QHBoxLayout()
        account_ifsc_row.setSpacing(8)
        account_number_label = QLabel('Account Number *')
        account_number_label.setStyleSheet(compact_label_style)
        account_number_label.setFixedWidth(100)
        self.account_number_input = QLineEdit()
        self.account_number_input.setStyleSheet(wide_input_style)
        ifsc_label = QLabel('IFSC Code')
        ifsc_label.setStyleSheet(compact_label_style)
        ifsc_label.setFixedWidth(80)
        self.ifsc_input = QLineEdit()
        self.ifsc_input.setStyleSheet(wide_input_style)
        self.ifsc_input.textChanged.connect(self.on_ifsc_changed)
        account_ifsc_row.addWidget(account_number_label)
        account_ifsc_row.addWidget(self.account_number_input)
        account_ifsc_row.addSpacing(10)
        account_ifsc_row.addWidget(ifsc_label)
        account_ifsc_row.addWidget(self.ifsc_input)
        account_ifsc_row.addStretch()
        layout.addLayout(account_ifsc_row)
        layout.addSpacing(3)
        branch_balance_row = QHBoxLayout()
        branch_balance_row.setSpacing(8)
        branch_label = QLabel('Branch Name')
        branch_label.setStyleSheet(compact_label_style)
        branch_label.setFixedWidth(80)
        self.branch_input = QLineEdit()
        self.branch_input.setStyleSheet(wide_input_style)
        self.branch_input.textChanged.connect(lambda text: self.on_text_changed(self.branch_input, text))
        opening_balance_label = QLabel('Opening Balance')
        opening_balance_label.setStyleSheet(compact_label_style)
        opening_balance_label.setFixedWidth(100)
        self.opening_balance_input = QLineEdit()
        self.opening_balance_input.setStyleSheet(compact_input_style)
        self.opening_balance_input.setFixedWidth(120)
        branch_balance_row.addWidget(branch_label)
        branch_balance_row.addWidget(self.branch_input)
        branch_balance_row.addSpacing(10)
        branch_balance_row.addWidget(opening_balance_label)
        branch_balance_row.addWidget(self.opening_balance_input)
        branch_balance_row.addStretch()
        layout.addLayout(branch_balance_row)
        layout.addSpacing(3)
        notes_label = QLabel('Notes')
        notes_label.setStyleSheet(compact_label_style)
        notes_label.setFixedWidth(50)
        self.notes_input = QLineEdit()
        self.notes_input.setStyleSheet(extra_wide_input_style)
        self.notes_input.textChanged.connect(lambda text: self.on_text_changed(self.notes_input, text))
        layout.addWidget(notes_label)
        layout.addWidget(self.notes_input)
        layout.addSpacing(3)
        actions_row = QHBoxLayout()
        actions_row.setSpacing(5)
        self.save_btn = QPushButton('Save')
        self.save_btn.setObjectName('save_btn')
        self.save_btn.clicked.connect(self.save)
        self.save_btn.setStyleSheet(theme.master_save_button_style())
        clear_btn = QPushButton('Clear')
        clear_btn.setStyleSheet(theme.master_clear_button_style())
        clear_btn.clicked.connect(self.clear_form)
        actions_row.addWidget(self.save_btn)
        actions_row.addWidget(clear_btn)
        actions_row.addStretch()
        layout.addLayout(actions_row)
        layout.addStretch()
        return container

    def create_list_page(self):
        container = QFrame()
        container.setObjectName('bankAccountListOuterFrame')
        container.setStyleSheet(theme.master_panel_frame_style('bankAccountListOuterFrame'))
        layout = QVBoxLayout(container)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        list_title = QLabel('Bank Account List')
        list_title.setStyleSheet(theme.master_page_title_style(18))
        layout.addWidget(list_title)
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(0, 0, 0, 10)
        search_label = QLabel('Search:')
        search_label.setStyleSheet(theme.master_label_style())
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText('Search by account name, bank name, account number, or IFSC code...')
        self.search_input.setStyleSheet(theme.sales_compact_input_style())
        self.search_input.textChanged.connect(self.apply_search)
        export_btn = QPushButton('Export')
        export_btn.setStyleSheet(theme.master_primary_action_button_style('8px 16px', 14))
        export_btn.clicked.connect(self.show_export_menu)
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(export_btn)
        search_layout.addStretch()
        layout.addLayout(search_layout)
        table_container = QFrame()
        table_container.setObjectName('bankAccountListTableContainer')
        table_container.setStyleSheet(theme.master_panel_frame_style('bankAccountListTableContainer'))
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.setSpacing(0)
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(['SL No', 'Account Name', 'Bank Name', 'Account Number', 'IFSC Code', 'Opening Balance'])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.itemSelectionChanged.connect(self.on_table_selection_changed)
        self.table.itemDoubleClicked.connect(self.on_table_double_click)
        self.table.setCornerButtonEnabled(False)
        self.table.verticalHeader().setVisible(False)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_bank_account_list_context_menu)
        self.table.setFrameShape(QFrame.NoFrame)
        self.table.setFrameShadow(QFrame.Plain)
        self.table.setLineWidth(0)
        self.table.setMidLineWidth(0)
        self.table.setContentsMargins(0, 0, 0, 0)
        self.table.setViewportMargins(0, 0, 0, 0)
        self.table.setStyleSheet(theme.master_table_style())
        header = self.table.horizontalHeader()
        header.setVisible(True)
        header.setFixedHeight(36)
        header.setMinimumHeight(36)
        header.setDefaultSectionSize(36)
        header.setHighlightSections(False)
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setStretchLastSection(False)
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setMinimumSectionSize(60)
        self.table.setColumnWidth(0, 80)
        self.table.setColumnWidth(1, 220)
        self.table.setColumnWidth(2, 220)
        self.table.setColumnWidth(3, 180)
        self.table.setColumnWidth(4, 140)
        self.table.setColumnWidth(5, 140)
        table_layout.addWidget(self.table)
        layout.addWidget(table_container)
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 10, 0, 0)
        edit_btn = QPushButton('Edit Selected')
        edit_btn.setStyleSheet(theme.master_nav_primary_button_style())
        edit_btn.clicked.connect(self.edit_selected_bank_account)
        delete_btn = QPushButton('Delete Selected')
        delete_btn.setStyleSheet(theme.master_danger_action_button_style())
        delete_btn.clicked.connect(self.delete_selected_bank_account)
        button_layout.addWidget(edit_btn)
        button_layout.addWidget(delete_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)
        return container

    def on_text_changed(self, widget, text):
        """Handle text change to capitalize first letter for relevant fields."""
        if widget and hasattr(widget, 'cursorPosition'):
            cursor_pos = widget.cursorPosition()
            capitalized_text = self.capitalize_first_letter(text)
            if capitalized_text != text:
                widget.blockSignals(True)
                widget.setText(capitalized_text)
                widget.setCursorPosition(cursor_pos)
                widget.blockSignals(False)

    def on_ifsc_changed(self, text):
        """Handle IFSC text change to convert to uppercase."""
        if self.ifsc_input and hasattr(self.ifsc_input, 'cursorPosition'):
            cursor_pos = self.ifsc_input.cursorPosition()
            uppercase_text = text.upper()
            if uppercase_text != text:
                self.ifsc_input.blockSignals(True)
                self.ifsc_input.setText(uppercase_text)
                self.ifsc_input.setCursorPosition(cursor_pos)
                self.ifsc_input.blockSignals(False)

    def capitalize_first_letter(self, text):
        """Capitalize the first letter of the text."""
        if not text:
            return text
        return text[0].upper() + text[1:]

    def clear_form(self):
        self.account_name_input.clear()
        self.bank_name_input.clear()
        self.account_number_input.clear()
        self.ifsc_input.clear()
        self.branch_input.clear()
        self.opening_balance_input.clear()
        self.notes_input.clear()
        self.current_bank_account_id = None
        if hasattr(self, 'save_btn'):
            self.save_btn.setText('Save')

    def save(self):
        active_company = active_company_manager.get_active_company()
        if not active_company:
            QMessageBox.warning(self, 'No Active Company', 'Please open a company first.')
            return
        account_name = self.account_name_input.text().strip()
        if not account_name:
            QMessageBox.warning(self, 'Validation Error', 'Account Name is required.')
            self.account_name_input.setFocus()
            return
        bank_name = self.bank_name_input.text().strip()
        if not bank_name:
            QMessageBox.warning(self, 'Validation Error', 'Bank Name is required.')
            self.bank_name_input.setFocus()
            return
        account_number = self.account_number_input.text().strip()
        if not account_number:
            QMessageBox.warning(self, 'Validation Error', 'Account Number is required.')
            self.account_number_input.setFocus()
            return
        try:
            bank_account_data = {'account_name': account_name, 'bank_name': bank_name, 'account_number': account_number, 'ifsc_code': self.ifsc_input.text().strip(), 'branch_name': self.branch_input.text().strip(), 'opening_balance': self.opening_balance_input.text() or '0', 'notes': self.notes_input.text().strip()}
            save_result = self.bank_account_logic.save_bank_account(active_company['id'], bank_account_data, self.current_bank_account_id)
            if save_result['success']:
                is_update = self.current_bank_account_id is not None
                success_message = (
                    'Bank account updated successfully.'
                    if is_update
                    else 'Bank account saved successfully.'
                )
                QMessageBox.information(self, 'Success', success_message)
                self.clear_form()
                QTimer.singleShot(0, lambda: self.account_name_input.setFocus())
                self.load_bank_accounts()
            else:
                QMessageBox.warning(self, 'Error', save_result['message'])
                if save_result.get('errors'):
                    QMessageBox.warning(self, 'Validation Errors', '\n'.join(save_result['errors']))
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to save bank account: {str(e)}')

    def load_bank_accounts(self):
        """Load all bank accounts from database into memory."""
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                self.bank_accounts_data = []
                self.visible_bank_accounts = []
                self.render_bank_accounts([])
                return
            result = self.bank_account_logic.get_bank_accounts(active_company['id'])
            self.bank_accounts_data = result
            self.apply_search()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load bank accounts: {str(e)}')
            self.bank_accounts_data = []
            self.visible_bank_accounts = []
            self.render_bank_accounts([])

    def render_bank_accounts(self, bank_accounts):
        """Render bank accounts in table."""
        self.table.setRowCount(len(bank_accounts))
        for row, bank_account in enumerate(bank_accounts):
            sl_no_item = QTableWidgetItem(str(row + 1))
            sl_no_item.setData(Qt.UserRole, bank_account['id'])
            self.table.setItem(row, 0, sl_no_item)
            name_item = QTableWidgetItem(bank_account['account_name'])
            name_item.setData(Qt.UserRole, bank_account['id'])
            self.table.setItem(row, 1, name_item)
            bank_name_item = QTableWidgetItem(bank_account['bank_name'])
            bank_name_item.setData(Qt.UserRole, bank_account['id'])
            self.table.setItem(row, 2, bank_name_item)
            account_number_item = QTableWidgetItem(bank_account['account_number'] or '')
            account_number_item.setData(Qt.UserRole, bank_account['id'])
            self.table.setItem(row, 3, account_number_item)
            ifsc_item = QTableWidgetItem(bank_account['ifsc_code'] or '')
            ifsc_item.setData(Qt.UserRole, bank_account['id'])
            self.table.setItem(row, 4, ifsc_item)
            opening_balance = float(bank_account['opening_balance'] or 0)
            balance_item = QTableWidgetItem(f'{opening_balance:.2f}')
            balance_item.setData(Qt.UserRole, bank_account['id'])
            balance_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row, 5, balance_item)

    def apply_search(self):
        """Apply search filter to bank accounts."""
        search_term = self.search_input.text().strip()
        if not search_term:
            self.visible_bank_accounts = self.bank_accounts_data
        else:
            self.visible_bank_accounts = self.bank_account_logic.filter_bank_accounts(self.bank_accounts_data, search_term)
        self.render_bank_accounts(self.visible_bank_accounts)

    def on_table_selection_changed(self):
        """Handle table row selection change."""
        pass

    def on_table_double_click(self, item):
        """Handle double-click on table row to edit bank account."""
        self.edit_selected_bank_account()

    def edit_selected_bank_account(self):
        """Edit the selected bank account by switching to entry page."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, 'No Selection', 'Please select a bank account to edit.')
            return
        bank_account_id = selected_items[0].data(Qt.UserRole)
        if not bank_account_id:
            QMessageBox.warning(self, 'Error', 'Unable to identify selected bank account.')
            return
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                return
            bank_account = self.bank_account_logic.get_bank_account(active_company['id'], bank_account_id)
            if bank_account:
                self.load_bank_account_to_form(bank_account)
                self.show_entry_page(clear_form=False)
            else:
                QMessageBox.warning(self, 'Error', 'Bank account not found.')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load bank account: {str(e)}')

    def delete_selected_bank_account(self):
        """Delete the selected bank account."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, 'No Selection', 'Please select a bank account to delete.')
            return
        bank_account_id = selected_items[0].data(Qt.UserRole)
        selected_row = self.table.currentRow()
        account_name_item = self.table.item(selected_row, 1)
        account_name = account_name_item.text() if account_name_item else 'selected bank account'
        if not bank_account_id:
            QMessageBox.warning(self, 'Error', 'Unable to identify selected bank account.')
            return
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                return
            reply = QMessageBox.question(self, 'Confirm Delete', f"Are you sure you want to delete '{account_name}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                result = self.bank_account_logic.delete_bank_account(active_company['id'], bank_account_id)
                if result['success']:
                    QMessageBox.information(self, 'Success', 'Bank account deleted successfully.')
                    self.load_bank_accounts()
                else:
                    QMessageBox.warning(self, 'Error', result['message'])
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to delete bank account: {str(e)}')

    def load_bank_account_to_form(self, bank_account):
        """Load bank account data into form fields."""
        self.current_bank_account_id = bank_account['id']
        self.account_name_input.setText(bank_account['account_name'])
        self.bank_name_input.setText(bank_account['bank_name'])
        self.account_number_input.setText(bank_account['account_number'])
        self.ifsc_input.setText(bank_account['ifsc_code'] or '')
        self.branch_input.setText(bank_account['branch_name'] or '')
        self.opening_balance_input.setText(str(bank_account['opening_balance']))
        self.notes_input.setText(bank_account['notes'] or '')
        if hasattr(self, 'save_btn'):
            self.save_btn.setText('Update')

    def keyPressEvent(self, event):
        """Handle key press events for Enter and Esc navigation."""
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            focus_widget = self.focusWidget()
            field_order = [self.account_name_input, self.bank_name_input, self.account_number_input, self.ifsc_input, self.branch_input, self.opening_balance_input, self.notes_input]
            if focus_widget in field_order:
                current_index = field_order.index(focus_widget)
                if current_index < len(field_order) - 1:
                    next_field = field_order[current_index + 1]
                    self.focus_and_force_select(next_field)
                else:
                    self.save()
            elif focus_widget == self.save_btn:
                self.save()
        elif event.key() == Qt.Key_Escape:
            focus_widget = self.focusWidget()
            if focus_widget == self.save_btn:
                self.focus_and_force_select(self.notes_input)
                return
            field_order = [self.account_name_input, self.bank_name_input, self.account_number_input, self.ifsc_input, self.branch_input, self.opening_balance_input, self.notes_input]
            if focus_widget in field_order:
                current_index = field_order.index(focus_widget)
                if current_index > 0:
                    prev_field = field_order[current_index - 1]
                    self.focus_and_force_select(prev_field)
        else:
            super().keyPressEvent(event)

    def focus_and_force_select(self, widget):
        """Set focus and select all text with proper timing."""
        widget.setFocus()
        QTimer.singleShot(0, widget.selectAll)

    def focus_and_select(self, widget):
        """Set focus and select all text with proper timing."""
        widget.setFocus()
        QTimer.singleShot(0, widget.selectAll)

    def show_bank_account_list_context_menu(self, position):
        """Show context menu for bank account list table."""
        menu = QMenu(self)
        select_all_action = menu.addAction('Select All')
        select_all_action.triggered.connect(self.select_all_visible_rows)
        clear_selection_action = menu.addAction('Clear Selection')
        clear_selection_action.triggered.connect(self.clear_selection)
        menu.exec(self.table.mapToGlobal(position))

    def select_all_visible_rows(self):
        """Select all visible rows in the table."""
        self.table.selectAll()

    def clear_selection(self):
        """Clear all selections in the table."""
        self.table.clearSelection()

    def get_visible_table_data(self):
        """Extract visible table data for export."""
        data = []
        headers = []
        for col in range(self.table.columnCount()):
            headers.append(self.table.horizontalHeaderItem(col).text())
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else '')
            data.append(row_data)
        return (headers, data)

    def show_export_menu(self):
        """Show export format selection dialog."""
        if not self.visible_bank_accounts:
            QMessageBox.warning(self, 'No Data', 'No bank accounts to export.')
            return
        dialog = QDialog(self)
        dialog.setWindowTitle('Export Bank Account List')
        dialog.setStyleSheet(report_detail_dialog_style())
        dialog.setFixedSize(350, 150)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        title_label = QLabel('Select Export Format:')
        title_label.setStyleSheet(theme.master_dialog_heading_style())
        layout.addWidget(title_label)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        pdf_btn = QPushButton('PDF')
        pdf_btn.setCheckable(True)
        pdf_btn.setEnabled(self.pdf_available)
        pdf_btn.clicked.connect(lambda: self.export_to_pdf(dialog))
        excel_btn = QPushButton('Excel')
        excel_btn.setCheckable(True)
        excel_btn.setEnabled(self.excel_available)
        excel_btn.clicked.connect(lambda: self.export_to_excel(dialog))
        cancel_btn = QPushButton('Cancel')
        cancel_btn.clicked.connect(dialog.reject)
        cancel_btn.setStyleSheet(theme.master_clear_button_style())
        button_layout.addWidget(pdf_btn)
        button_layout.addWidget(excel_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        dialog.exec()

    def export_to_pdf(self, dialog=None):
        """Export visible bank accounts to PDF."""
        if not self.pdf_available:
            QMessageBox.warning(self, 'Library Not Installed', 'Required library not installed.\n\nTo enable PDF export, install:\npip install reportlab')
            return
        try:
            headers, data = self.get_visible_table_data()
            if not data:
                QMessageBox.warning(self, 'No Data', 'No data available to export.')
                return
            file_path, _ = QFileDialog.getSaveFileName(self, 'Save PDF', 'bank_account_list.pdf', 'PDF Files (*.pdf)')
            if not file_path:
                return
            table_data = [headers] + data
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            table = Table(table_data)
            table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -1), colors.white), ('TEXTCOLOR', (0, 1), (-1, -1), colors.black), ('GRID', (0, 0), (-1, -1), 1, colors.black), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])]))
            doc.build([table])
            QMessageBox.information(self, 'Success', 'PDF exported successfully.')
            if dialog:
                dialog.accept()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to export PDF: {str(e)}')

    def export_to_excel(self, dialog=None):
        """Export visible bank accounts to Excel."""
        if not self.excel_available:
            QMessageBox.warning(self, 'Library Not Installed', 'Required library not installed.\n\nTo enable Excel export, install:\npip install openpyxl')
            return
        try:
            file_path, _ = QFileDialog.getSaveFileName(self, 'Save Excel', 'bank_account_list.xlsx', 'Excel Files (*.xlsx)')
            if not file_path:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = 'Bank Account List'
            headers = ['SL No', 'Account Name', 'Bank Name', 'Account Number', 'IFSC Code', 'Opening Balance']
            ws.append(headers)
            header_fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            border = Border(left=Side(style='thin', color='4b5563'), right=Side(style='thin', color='4b5563'), top=Side(style='thin', color='4b5563'), bottom=Side(style='thin', color='4b5563'))
            for col in range(1, 7):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            for idx, bank_account in enumerate(self.visible_bank_accounts):
                row = idx + 2
                ws.cell(row=row, column=1, value=idx + 1)
                ws.cell(row=row, column=2, value=bank_account['account_name'] or '')
                ws.cell(row=row, column=3, value=bank_account['bank_name'] or '')
                ws.cell(row=row, column=4, value=bank_account['account_number'] or '')
                ws.cell(row=row, column=5, value=bank_account['ifsc_code'] or '')
                ws.cell(row=row, column=6, value=float(bank_account['opening_balance'] or 0))
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if col == 6:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15
            wb.save(file_path)
            QMessageBox.information(self, 'Success', 'Excel file exported successfully.')
            if dialog:
                dialog.accept()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to export Excel: {str(e)}')