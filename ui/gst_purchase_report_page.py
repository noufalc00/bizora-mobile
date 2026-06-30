"""
GST Purchase Report Page for the Accounting Desktop Application.
Provides GSTIN-wise, Rate-wise, and Invoice-wise purchase reporting for GST compliance.
"""
from PySide6.QtWidgets import *
from PySide6.QtCore import Qt, QDate, QObject, QThread, Signal
from config import active_company_manager
from db import Database
from bizora_core.gst_compliance import normalized_gstin, gst_slab_rate_from_totals
from ui.table_header_utils import apply_adjustable_table_columns, apply_read_only_report_table_selection, apply_compact_report_table_rows
from ui import theme
from ui.book_report_common import compact_date_style, page_background_style, report_filter_frame_style, report_detail_dialog_style, report_compound_entry_page_style, report_page_shell_style, page_heading_style, report_status_label_style, report_data_table_style, footer_title_style, footer_value_style, report_summary_label_style
from ui.date_formats import configure_qdate_edit, format_display_date, prepare_report_date_edit, qdate_to_db, qdate_to_display
from ui.message_boxes import critical as show_critical, information as show_information, warning as show_warning
from ui.ui_memory import UiMemoryMixin, memory_table_attr_slug

class GSTPurchaseReportWorker(QObject):
    """Load GST purchase report vouchers outside the GUI thread."""
    data_ready = Signal(list)
    error = Signal(str)
    finished = Signal()

    def __init__(self, db_type, db_path, company_id, company_state, from_date, to_date, search):
        """Store worker query inputs for background report generation."""
        super().__init__()
        self.db_type = db_type
        self.db_path = db_path
        self.company_id = company_id
        self.company_state = company_state
        self.from_date = from_date
        self.to_date = to_date
        self.search = search

    def run(self):
        """Fetch and normalize purchase report rows in a worker-owned connection."""
        worker_db = None
        try:
            worker_db = Database(db_type=self.db_type, db_path=self.db_path)
            purchases = worker_db.get_purchases_for_gst_report(self.company_id, self.from_date, self.to_date, self.search)
            for purchase in purchases:
                self._normalise_purchase_tax_values(purchase)
                purchase['rate'] = self._get_tax_rate(purchase)
                purchase['place_of_supply'] = self._place_of_supply(purchase)
                purchase['supplier_classification'] = self._supplier_classification(purchase)
            self.data_ready.emit(purchases)
        except Exception as exc:
            self.error.emit(str(exc))
        finally:
            if worker_db is not None:
                try:
                    worker_db.force_disconnect()
                except Exception:
                    pass
            self.finished.emit()

    def _normalise_purchase_tax_values(self, purchase):
        """Normalize split taxes without changing header invoice value."""
        taxable = self._to_float(purchase.get('taxable_value'))
        tax_total = self._to_float(purchase.get('tax_total') or purchase.get('item_tax_total'))
        cgst = self._to_float(purchase.get('cgst'))
        sgst = self._to_float(purchase.get('sgst'))
        igst = self._to_float(purchase.get('igst'))
        cess = self._to_float(purchase.get('cess'))
        if tax_total and (not (cgst or sgst or igst or cess)):
            nature = (purchase.get('nature') or '').casefold()
            state = self._place_of_supply(purchase)
            is_interstate = 'inter' in nature or (state and self.company_state and (state.casefold() != self.company_state.casefold()))
            if is_interstate:
                igst = tax_total
            else:
                cgst = round(tax_total / 2, 2)
                sgst = round(tax_total - cgst, 2)
        purchase['taxable_value'] = taxable
        purchase['cgst'] = cgst
        purchase['sgst'] = sgst
        purchase['igst'] = igst
        purchase['cess'] = cess
        purchase['tax_total'] = tax_total
        purchase['itc_eligible'] = taxable + cgst + sgst + igst + cess

    def _supplier_gstin(self, purchase):
        """Return structurally valid supplier GSTIN or blank."""
        return normalized_gstin(purchase.get('supplier_gstin') or purchase.get('party_gstin') or purchase.get('gstin'))

    def _supplier_classification(self, purchase):
        """Classify purchase supplier for GST purchase reporting."""
        if self._supplier_gstin(purchase):
            return 'B2B (Registered)'
        return 'B2BUR (Unregistered)'

    def _place_of_supply(self, purchase):
        """Resolve supplier place of supply from party or voucher state."""
        return (purchase.get('supplier_state') or purchase.get('party_state') or purchase.get('state') or '').strip()

    def _get_tax_rate(self, purchase):
        """Return GST slab percentage excluding CESS from the rate calculation."""
        return gst_slab_rate_from_totals(purchase)

    def _to_float(self, value):
        """Convert worker calculation inputs to float."""
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0

class GSTPurchaseReportPage(UiMemoryMixin, QWidget):
    """GST Purchase Report page with GSTIN-wise, Rate-wise, and Invoice-wise views."""
    MONEY_KEYS = {'Invoice Value', 'Total Invoice Value', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'CESS', 'ITC Eligible', 'Total Value', 'Footer Adjustment'}
    COLUMN_MIN_WIDTHS = {
        'Document Type': 110,
        'Classification': 130,
        'Supplier GSTIN': 160,
        'Supplier Name': 180,
        'Supplier Invoice No': 150,
        'Voucher No': 120,
        'Voucher Date': 110,
        'Total Invoice Value': 140,
        'Place of Supply': 130,
        'Rate': 70,
        'Taxable Value': 110,
        'CGST': 90,
        'SGST': 90,
        'IGST': 90,
        'CESS': 90,
        'Footer Adjustment': 130,
        'ITC Eligible': 110,
        'Invoice Count': 100,
    }

    def __init__(self, db=None):
        super().__init__()
        self.db = db or Database()
        self.company_id = None
        self.company_state = ''
        self._loading = False
        self._report_thread = None
        self._report_worker = None
        self.current_report_data = []
        self.setup_ui()
        self.load_company()
        self._init_ui_memory(table_attrs=("results_table",))
        report_type = self.report_type_combo.currentText()
        memory_attr = f"results_table_{memory_table_attr_slug(report_type)}"
        self._ui_memory_active_table_attr = memory_attr
        self._ui_memory_active_table = self.results_table
        self._restore_memory_table(self.results_table, memory_attr)

    def setup_ui(self):
        self.setStyleSheet(report_compound_entry_page_style())
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(5)
        self.title_label = QLabel('GST Purchase Report')
        self.title_label.setStyleSheet(page_heading_style(18))
        layout.addWidget(self.title_label)
        self.filter_frame = QFrame()
        self.filter_frame.setStyleSheet(report_filter_frame_style())
        filter_layout = QHBoxLayout(self.filter_frame)
        filter_layout.setSpacing(10)
        self.from_date = QDateEdit()
        self.from_date.setDate(QDate.currentDate().addDays(-30))
        prepare_report_date_edit(self.from_date, style_sheet=compact_date_style())
        self.to_date = QDateEdit()
        self.to_date.setDate(QDate.currentDate())
        prepare_report_date_edit(self.to_date, style_sheet=compact_date_style())
        self.report_type_combo = QComboBox()
        self.report_type_combo.addItems(['Invoice-wise', 'GSTIN-wise', 'Rate-wise'])
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText('Supplier Name / GSTIN')
        for label_text, widget in (('From Date:', self.from_date), ('To Date:', self.to_date), ('Report Type:', self.report_type_combo), ('Supplier:', self.search_input)):
            filter_layout.addWidget(QLabel(label_text))
            filter_layout.addWidget(widget)
        generate_btn = QPushButton('Generate')
        generate_btn.clicked.connect(self.generate_report)
        export_btn = QPushButton('Export Excel')
        export_btn.clicked.connect(self.export_excel)
        self.generate_btn = generate_btn
        self.export_btn = export_btn
        filter_layout.addWidget(generate_btn)
        filter_layout.addWidget(export_btn)
        filter_layout.addStretch()
        layout.addWidget(self.filter_frame)
        self.status_label = QLabel('Ready')
        self.status_label.setStyleSheet(report_status_label_style())
        layout.addWidget(self.status_label)
        self.results_table = QTableWidget()
        self.results_table.setStyleSheet(report_data_table_style())
        self.results_table.setSortingEnabled(False)
        apply_read_only_report_table_selection(self.results_table)
        self.results_table.setTextElideMode(Qt.ElideNone)
        self.results_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.results_table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.results_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.results_table.itemDoubleClicked.connect(self.on_row_double_clicked)
        apply_compact_report_table_rows(self.results_table)
        layout.addWidget(self.results_table, 1)
        layout.addWidget(self._build_footer_frame())
        self.report_type_combo.currentTextChanged.connect(self._sync_results_table_headers)
        self._sync_results_table_headers(self.report_type_combo.currentText())

    def _invoice_wise_headers(self):
        return [
            'Document Type', 'Classification', 'Supplier GSTIN', 'Supplier Name',
            'Supplier Invoice No', 'Voucher No', 'Voucher Date', 'Total Invoice Value',
            'Place of Supply', 'Rate', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'CESS',
            'Footer Adjustment', 'ITC Eligible',
        ]

    def _gstin_wise_headers(self):
        return [
            'Classification', 'Supplier GSTIN', 'Supplier Name', 'Total Invoice Value',
            'Taxable Value', 'CGST', 'SGST', 'IGST', 'CESS', 'Footer Adjustment', 'Invoice Count',
        ]

    def _rate_wise_headers(self):
        return [
            'Classification', 'Rate', 'Total Invoice Value', 'Taxable Value', 'CGST', 'SGST',
            'IGST', 'CESS', 'Footer Adjustment', 'Invoice Count',
        ]

    def _sync_results_table_headers(self, report_type: str) -> None:
        """Pre-fill GST purchase table headers for the selected report layout."""
        headers_map = {
            'Invoice-wise': self._invoice_wise_headers(),
            'GSTIN-wise': self._gstin_wise_headers(),
            'Rate-wise': self._rate_wise_headers(),
        }
        headers = headers_map.get(report_type, self._invoice_wise_headers())
        self.results_table.setColumnCount(len(headers))
        self.results_table.setHorizontalHeaderLabels(headers)
        apply_read_only_report_table_selection(self.results_table)
        memory_attr = f"results_table_{memory_table_attr_slug(report_type)}"
        self._ui_memory_active_table_attr = memory_attr
        self._ui_memory_active_table = self.results_table
        if hasattr(self, "settings"):
            self._restore_memory_table(self.results_table, memory_attr)
        self._apply_results_table_column_widths()

    def load_company(self):
        active = active_company_manager.get_active_company()
        if active:
            self.company_id = active.get('id')
            self.company_state = (active.get('state') or '').strip()

    def generate_report(self):
        """Start the GST purchase report worker."""
        if self._loading:
            return
        if not self.company_id:
            show_warning(self, 'No Company', 'Please open a company first.')
            return
        from_date = qdate_to_db(self.from_date.date())
        to_date = qdate_to_db(self.to_date.date())
        if from_date > to_date:
            show_warning(
                self,
                'Invalid Date Range',
                (
                    f'From Date ({qdate_to_display(self.from_date.date())}) cannot be later than '
                    f'To Date ({qdate_to_display(self.to_date.date())}).'
                ),
            )
            return
        report_type = self.report_type_combo.currentText()
        search = self.search_input.text().strip()
        db_type = getattr(self.db, 'db_type', None)
        db_path = getattr(self.db, 'db_path', None)
        thread = QThread(self)
        worker = GSTPurchaseReportWorker(db_type, db_path, self.company_id, self.company_state, from_date, to_date, search)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.data_ready.connect(lambda data, selected=report_type: self._on_report_ready(data, selected))
        worker.error.connect(self._on_report_error)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(self._on_report_finished)
        self._report_thread = thread
        self._report_worker = worker
        self._set_loading_state(True)
        thread.start()

    def _set_loading_state(self, is_loading):
        """Disable report controls while the worker is running."""
        self._loading = is_loading
        for widget in (
            self.generate_btn,
            self.export_btn,
            self.report_type_combo,
            self.from_date,
            self.to_date,
            self.search_input,
        ):
            widget.setEnabled(not is_loading)
        if is_loading:
            self.status_label.setText('Calculating...')
            self._reset_footer_totals()

    def _show_empty_report(self, message: str) -> None:
        """Render an empty report state without clearing table headers."""
        report_type = self.report_type_combo.currentText()
        self._sync_results_table_headers(report_type)
        column_count = max(self.results_table.columnCount(), 1)
        self.results_table.setRowCount(1)
        item = QTableWidgetItem(message)
        item.setFlags(Qt.ItemIsEnabled)
        item.setTextAlignment(Qt.AlignCenter)
        self.results_table.setItem(0, 0, item)
        self.results_table.setSpan(0, 0, 1, column_count)
        self._reset_footer_totals()
        self.status_label.setText(message)

    def _on_report_ready(self, purchases, report_type):
        """Display worker-calculated purchase data on the GUI thread."""
        self.current_report_data = purchases
        if not purchases:
            self._show_empty_report('No purchase data found for the selected criteria.')
            return
        if report_type == 'Invoice-wise':
            self._display_invoice_wise(purchases)
        elif report_type == 'GSTIN-wise':
            self._display_gstin_wise(purchases)
        elif report_type == 'Rate-wise':
            self._display_rate_wise(purchases)
        self._update_footer_totals(purchases)
        self.status_label.setText('GST purchase report generated successfully.')

    def _on_report_error(self, message):
        """Show worker errors without blocking the GUI thread."""
        show_critical(self, 'Error', f'Failed to generate report: {message}')
        self._show_empty_report('GST purchase report generation failed.')

    def _on_report_finished(self):
        """Clear worker references after the thread exits."""
        self._report_thread = None
        self._report_worker = None
        self._set_loading_state(False)

    def _build_footer_frame(self):
        """Create the dark themed GST totals footer."""
        frame = QFrame()
        frame.setStyleSheet(theme.section_panel_frame_style() + '\n            QLabel {\n                background: transparent;\n                border: none;\n            }\n        ')
        footer_layout = QHBoxLayout(frame)
        footer_layout.setContentsMargins(5, 5, 5, 5)
        footer_layout.setSpacing(12)
        self.footer_total_labels = {}
        self.footer_total_titles = {}
        footer_fields = [('Total Invoice Value', 'invoice_value'), ('Total Taxable Value', 'taxable_value'), ('Total CGST', 'cgst'), ('Total SGST', 'sgst'), ('Total IGST', 'igst'), ('Total Cess', 'cess'), ('Total Tax', 'total_tax')]
        for title, key in footer_fields:
            metric_label = QLabel()
            metric_label.setTextFormat(Qt.RichText)
            metric_label.setWordWrap(False)
            metric_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
            metric_label.setStyleSheet('background: transparent; border: none;')
            self._set_footer_metric(metric_label, title, '0.00')
            footer_layout.addWidget(metric_label)
            self.footer_total_labels[key] = metric_label
            self.footer_total_titles[key] = title
        footer_layout.addStretch()
        return frame

    def _footer_metric_html(self, title, value):
        """Return rich footer HTML that keeps a metric title and value together."""
        clean_title = str(title or '').strip()
        if not clean_title.endswith(':'):
            clean_title = f'{clean_title}:'
        return f"<span style='{footer_title_style()}'>{clean_title} </span><span style='{footer_value_style()}'>{value}</span>"

    def _set_footer_metric(self, label, title, value):
        """Update a combined footer metric label without splitting widgets."""
        label.setText(self._footer_metric_html(title, value))

    def _set_footer_total(self, key, value):
        """Set one footer total while preserving its title text."""
        label = self.footer_total_labels.get(key)
        title = self.footer_total_titles.get(key, key)
        if label is not None:
            self._set_footer_metric(label, title, value)

    def _reset_footer_totals(self):
        """Reset footer totals while loading or after errors."""
        if not hasattr(self, 'footer_total_labels'):
            return
        for key in self.footer_total_labels:
            self._set_footer_total(key, '0.00')

    def _update_footer_totals(self, purchases):
        """Calculate footer totals from currently loaded purchase report data."""
        totals = {'invoice_value': 0.0, 'taxable_value': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0, 'cess': 0.0, 'total_tax': 0.0}
        for purchase in purchases:
            cgst = self._to_float(purchase.get('cgst'))
            sgst = self._to_float(purchase.get('sgst'))
            igst = self._to_float(purchase.get('igst'))
            cess = self._to_float(purchase.get('cess'))
            totals['invoice_value'] += self._to_float(purchase.get('grand_total'))
            totals['taxable_value'] += self._to_float(purchase.get('taxable_value'))
            totals['cgst'] += cgst
            totals['sgst'] += sgst
            totals['igst'] += igst
            totals['cess'] += cess
            totals['total_tax'] += cgst + sgst + igst + cess
        for key, value in totals.items():
            self._set_footer_total(key, f'{value:.2f}')

    def _display_invoice_wise(self, purchases):
        """Display one row per purchase invoice or debit note."""
        headers = ['Document Type', 'Classification', 'Supplier GSTIN', 'Supplier Name', 'Supplier Invoice No', 'Voucher No', 'Voucher Date', 'Total Invoice Value', 'Place of Supply', 'Rate', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'CESS', 'Footer Adjustment', 'ITC Eligible']
        self._reset_table(headers)
        for purchase in purchases:
            row = self.results_table.rowCount()
            self.results_table.insertRow(row)
            values = [purchase.get('document_type', 'Purchase'), purchase.get('supplier_classification', ''), self._supplier_gstin(purchase), purchase.get('party_name', ''), purchase.get('supplier_invoice_no', ''), purchase.get('purchase_number', ''), format_display_date(purchase.get('purchase_date', '')), purchase.get('grand_total', 0.0), purchase.get('place_of_supply', ''), purchase.get('rate', 0.0), purchase.get('taxable_value', 0.0), purchase.get('cgst', 0.0), purchase.get('sgst', 0.0), purchase.get('igst', 0.0), purchase.get('cess', 0.0), purchase.get('footer_adjustment', 0.0), purchase.get('itc_eligible', 0.0)]
            for col, value in enumerate(values):
                item = self._make_item(value, headers[col])
                item.setData(Qt.UserRole, self._voucher_metadata(purchase))
                self.results_table.setItem(row, col, item)
        self._finish_table()

    def _display_gstin_wise(self, purchases):
        """Display GSTIN and supplier-classification grouped totals."""
        gstin_data = {}
        for purchase in purchases:
            gstin = self._supplier_gstin(purchase) or 'Unregistered'
            classification = purchase.get('supplier_classification', self._supplier_classification(purchase))
            key = (classification, gstin, purchase.get('party_name', ''))
            if key not in gstin_data:
                gstin_data[key] = {'classification': classification, 'gstin': gstin, 'party_name': purchase.get('party_name', ''), 'invoice_value': 0.0, 'taxable_value': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0, 'cess': 0.0, 'footer_adjustment': 0.0, 'total_value': 0.0, 'invoice_count': 0}
            gstin_data[key]['invoice_value'] += self._to_float(purchase.get('grand_total'))
            gstin_data[key]['taxable_value'] += self._to_float(purchase.get('taxable_value'))
            gstin_data[key]['cgst'] += self._to_float(purchase.get('cgst'))
            gstin_data[key]['sgst'] += self._to_float(purchase.get('sgst'))
            gstin_data[key]['igst'] += self._to_float(purchase.get('igst'))
            gstin_data[key]['cess'] += self._to_float(purchase.get('cess'))
            gstin_data[key]['footer_adjustment'] += self._to_float(purchase.get('footer_adjustment'))
            gstin_data[key]['total_value'] += self._to_float(purchase.get('grand_total'))
            gstin_data[key]['invoice_count'] += 1
        headers = ['Classification', 'Supplier GSTIN', 'Supplier Name', 'Total Invoice Value', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'CESS', 'Footer Adjustment', 'Invoice Count']
        self._reset_table(headers)
        for data in gstin_data.values():
            row = self.results_table.rowCount()
            self.results_table.insertRow(row)
            values = [data['classification'], data['gstin'], data['party_name'], data['invoice_value'], data['taxable_value'], data['cgst'], data['sgst'], data['igst'], data['cess'], data['footer_adjustment'], data['invoice_count']]
            for col, value in enumerate(values):
                self.results_table.setItem(row, col, self._make_item(value, headers[col]))
        self._finish_table()

    def _display_rate_wise(self, purchases):
        """Display tax-rate grouped totals split by supplier classification."""
        rate_data = {}
        for purchase in purchases:
            rate = purchase.get('rate', 0.0)
            classification = purchase.get('supplier_classification', self._supplier_classification(purchase))
            key = (classification, rate)
            if key not in rate_data:
                rate_data[key] = {'classification': classification, 'rate': rate, 'invoice_value': 0.0, 'taxable_value': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'igst': 0.0, 'cess': 0.0, 'footer_adjustment': 0.0, 'total_value': 0.0, 'invoice_count': 0}
            rate_data[key]['invoice_value'] += self._to_float(purchase.get('grand_total'))
            rate_data[key]['taxable_value'] += self._to_float(purchase.get('taxable_value'))
            rate_data[key]['cgst'] += self._to_float(purchase.get('cgst'))
            rate_data[key]['sgst'] += self._to_float(purchase.get('sgst'))
            rate_data[key]['igst'] += self._to_float(purchase.get('igst'))
            rate_data[key]['cess'] += self._to_float(purchase.get('cess'))
            rate_data[key]['footer_adjustment'] += self._to_float(purchase.get('footer_adjustment'))
            rate_data[key]['total_value'] += self._to_float(purchase.get('grand_total'))
            rate_data[key]['invoice_count'] += 1
        headers = ['Classification', 'Rate', 'Total Invoice Value', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'CESS', 'Footer Adjustment', 'Invoice Count']
        self._reset_table(headers)
        for data in sorted(rate_data.values(), key=lambda item: (item['classification'], item['rate'])):
            row = self.results_table.rowCount()
            self.results_table.insertRow(row)
            values = [data['classification'], data['rate'], data['invoice_value'], data['taxable_value'], data['cgst'], data['sgst'], data['igst'], data['cess'], data['footer_adjustment'], data['invoice_count']]
            for col, value in enumerate(values):
                self.results_table.setItem(row, col, self._make_item(value, headers[col]))
        self._finish_table()

    def _reset_table(self, headers):
        self.results_table.clear()
        self.results_table.setRowCount(0)
        self.results_table.setColumnCount(len(headers))
        self.results_table.setHorizontalHeaderLabels(headers)

    def _column_minimum_width(self, header_text: str) -> int:
        """Return the preferred minimum width for one report column header."""
        return self.COLUMN_MIN_WIDTHS.get(header_text, 90)

    def _apply_results_table_column_widths(self) -> None:
        """Apply readable default widths so the table can scroll horizontally."""
        header = self.results_table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setMinimumSectionSize(48)
        for column_index in range(self.results_table.columnCount()):
            header.setSectionResizeMode(column_index, QHeaderView.ResizeMode.Interactive)
            header_text = ''
            header_item = self.results_table.horizontalHeaderItem(column_index)
            if header_item is not None:
                header_text = header_item.text()
            minimum_width = self._column_minimum_width(header_text)
            current_width = self.results_table.columnWidth(column_index)
            self.results_table.setColumnWidth(column_index, max(current_width, minimum_width))

    def _finish_table(self):
        apply_adjustable_table_columns(self.results_table, auto_size=False)
        apply_compact_report_table_rows(self.results_table)
        report_type = self.report_type_combo.currentText()
        memory_attr = f"results_table_{memory_table_attr_slug(report_type)}"
        self._ui_memory_active_table_attr = memory_attr
        self._ui_memory_active_table = self.results_table
        if hasattr(self, "settings"):
            self._restore_memory_table(self.results_table, memory_attr)
        self._apply_results_table_column_widths()

    def _make_item(self, value, header):
        """Create a non-editable table item with monetary formatting."""
        if header in self.MONEY_KEYS or isinstance(value, float):
            text = f'{self._to_float(value):.2f}'
        else:
            text = str(value if value is not None else '')
        item = QTableWidgetItem(text)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        return item

    def _voucher_metadata(self, purchase):
        """Return metadata used by double-click handlers."""
        return {'voucher_type': purchase.get('voucher_type', 'purchase'), 'id': purchase.get('id')}

    def _normalise_purchase_tax_values(self, purchase):
        taxable = self._to_float(purchase.get('taxable_value'))
        tax_total = self._to_float(purchase.get('tax_total'))
        cgst = self._to_float(purchase.get('cgst'))
        sgst = self._to_float(purchase.get('sgst'))
        igst = self._to_float(purchase.get('igst'))
        cess = self._to_float(purchase.get('cess'))
        if tax_total and (not (cgst or sgst or igst or cess)):
            nature = (purchase.get('nature') or '').casefold()
            state = self._place_of_supply(purchase)
            is_interstate = 'inter' in nature or (state and self.company_state and (state.casefold() != self.company_state.casefold()))
            if is_interstate:
                igst = tax_total
            else:
                cgst = round(tax_total / 2, 2)
                sgst = round(tax_total - cgst, 2)
        purchase['taxable_value'] = taxable
        purchase['cgst'] = cgst
        purchase['sgst'] = sgst
        purchase['igst'] = igst
        purchase['cess'] = cess
        purchase['tax_total'] = tax_total
        purchase['itc_eligible'] = taxable + cgst + sgst + igst + cess

    def _supplier_gstin(self, purchase):
        return normalized_gstin(purchase.get('supplier_gstin') or purchase.get('party_gstin') or purchase.get('gstin'))

    def _supplier_classification(self, purchase):
        """Return GST purchase supplier registration classification."""
        if self._supplier_gstin(purchase):
            return 'B2B (Registered)'
        return 'B2BUR (Unregistered)'

    def _place_of_supply(self, purchase):
        return (purchase.get('supplier_state') or purchase.get('party_state') or purchase.get('state') or '').strip()

    def _get_tax_rate(self, purchase):
        """Return GST slab percentage excluding CESS from the rate calculation."""
        return gst_slab_rate_from_totals(purchase)

    def _to_float(self, value):
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0

    def on_row_double_clicked(self, item):
        metadata = item.data(Qt.UserRole)
        if not metadata:
            return
        if isinstance(metadata, dict):
            if metadata.get('voucher_type') == 'purchase':
                self._show_purchase_details(metadata.get('id'))
            else:
                QMessageBox.information(self, 'Debit Note', 'Debit note detail view is not available from this report.')
        else:
            self._show_purchase_details(metadata)

    def _show_purchase_details(self, purchase_id):
        try:
            ph = self.db._get_placeholder()
            query = f'\n                SELECT p.*, pr.name as party_name, pr.gstin as party_gstin, pr.state as party_state\n                FROM purchases p\n                LEFT JOIN parties pr ON p.party_id = pr.id\n                WHERE p.id = {ph}\n            '
            result = self.db.execute_query(query, (purchase_id,))
            if not result:
                QMessageBox.information(self, 'Not Found', 'Purchase not found.')
                return
            purchase = result[0]
            items_query = f'\n                SELECT pi.*, prod.name as product_name, prod.hsn\n                FROM purchase_items pi\n                LEFT JOIN products prod ON pi.product_id = prod.id\n                WHERE pi.purchase_id = {ph}\n                ORDER BY pi.sl_no\n            '
            items = self.db.execute_query(items_query, (purchase_id,))
            self._show_details_dialog(purchase, items)
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load purchase details: {e}')

    def _show_details_dialog(self, purchase, items):
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Purchase Details - {purchase.get('purchase_number', '')}")
        dialog.setMinimumSize(1000, 650)
        dialog.setStyleSheet(self.results_table.styleSheet() + report_detail_dialog_style())
        layout = QVBoxLayout(dialog)
        header_label = QLabel(f"<b>Purchase No:</b> {purchase.get('purchase_number', '')} &nbsp;&nbsp; <b>Date:</b> {format_display_date(purchase.get('purchase_date', ''))} &nbsp;&nbsp; <b>Party:</b> {purchase.get('party_name', '')} &nbsp;&nbsp; <b>GSTIN:</b> {purchase.get('party_gstin') or purchase.get('gstin') or ''} &nbsp;&nbsp; <b>State:</b> {purchase.get('party_state') or purchase.get('state') or ''}")
        header_label.setWordWrap(True)
        header_label.setStyleSheet(page_heading_style(14))
        layout.addWidget(header_label)
        table = QTableWidget()
        apply_read_only_report_table_selection(table)
        headers = ['SL', 'Product', 'HSN', 'Qty', 'Rate', 'Taxable', 'Tax', 'Total']
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(items))
        for i, row in enumerate(items):
            values = [row.get('sl_no', ''), row.get('product_name', ''), row.get('hsn', ''), row.get('quantity', 0), row.get('rate', 0), row.get('net_value', 0), row.get('tax_amount', 0), row.get('grand_total', 0)]
            for col, value in enumerate(values):
                table.setItem(i, col, self._make_item(value, headers[col]))
        apply_adjustable_table_columns(table)
        apply_compact_report_table_rows(table)
        layout.addWidget(table)
        totals_label = QLabel(f"<b>Taxable:</b> {self._to_float(purchase.get('sub_total')):.2f} &nbsp;&nbsp; <b>Tax:</b> {self._to_float(purchase.get('tax_total')):.2f} &nbsp;&nbsp; <b>Grand Total:</b> {self._to_float(purchase.get('grand_total')):.2f}")
        totals_label.setStyleSheet(report_summary_label_style() + ' padding: 10px;')
        layout.addWidget(totals_label)
        close_btn = QPushButton('Close')
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignRight)
        dialog.exec()

    def export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            wb = Workbook()
            ws = wb.active
            ws.title = 'GST Purchase Report'
            from_date = qdate_to_db(self.from_date.date())
            to_date = qdate_to_db(self.to_date.date())
            ws['A1'] = f'GST Purchase Report - {from_date} to {to_date}'
            ws['A1'].font = Font(bold=True, size=14)
            for col in range(self.results_table.columnCount()):
                header = self.results_table.horizontalHeaderItem(col)
                if header:
                    ws.cell(row=3, column=col + 1, value=header.text())
            for row in range(self.results_table.rowCount()):
                for col in range(self.results_table.columnCount()):
                    item = self.results_table.item(row, col)
                    if item:
                        ws.cell(row=row + 4, column=col + 1, value=item.text())
            filename = f'GST_Purchase_Report_{from_date}_to_{to_date}.xlsx'
            wb.save(filename)
            QMessageBox.information(self, 'Export Success', f'Report exported to {filename}')
        except ImportError:
            QMessageBox.warning(self, 'Export Failed', 'openpyxl is not installed. Please install it first.')
        except Exception as e:
            QMessageBox.critical(self, 'Export Failed', f'Failed to export: {e}')

    def refresh_theme(self) -> None:
        """Re-apply theme-aware styles after a global theme change."""
        self.setStyleSheet(report_compound_entry_page_style())
        if hasattr(self, 'title_label'):
            self.title_label.setStyleSheet(page_heading_style(18))
        if hasattr(self, 'filter_frame'):
            self.filter_frame.setStyleSheet(report_filter_frame_style())
        if hasattr(self, 'status_label'):
            self.status_label.setStyleSheet(report_status_label_style())
        if hasattr(self, 'results_table'):
            self.results_table.setStyleSheet(report_data_table_style())
        prepare_report_date_edit(self.from_date, style_sheet=compact_date_style())
        prepare_report_date_edit(self.to_date, style_sheet=compact_date_style())