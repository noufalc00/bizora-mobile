"""
Debitor/Creditor widget for the Accounting Desktop Application.
Manages debtors, creditors, and parties with company-wise data storage.
"""
from PySide6.QtWidgets import *
from PySide6.QtCore import Qt, QTimer, Signal, QEvent
from PySide6.QtGui import QTextCursor
from config import active_company_manager
from db import Database
from bizora_core.party_logic import PartyLogic
from ui import theme
from ui.book_report_common import report_detail_dialog_style
from ui.party_display import normalise_party_code, party_display_name
from ui.table_header_utils import apply_read_only_report_table_selection
from ui.date_formats import configure_qdate_edit, format_display_date, qdate_to_db, qdate_to_display, db_to_qdate
from ui.ui_memory import UiMemoryMixin
PDF_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
try:
    from docx import Document
    from docx.shared import Pt
    WORD_AVAILABLE = True
except ImportError:
    WORD_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class DebitorCreditorWidget(UiMemoryMixin, QWidget):
    party_saved = Signal()

    def __init__(self, db=None):
        super().__init__()
        self.db = db or Database()
        self.party_logic = PartyLogic(self.db)
        self.current_party_id = None
        self.parties_data = []
        self.current_filter = 'All'
        self.visible_parties = []
        self.pdf_available = PDF_AVAILABLE
        self.word_available = WORD_AVAILABLE
        self.excel_available = EXCEL_AVAILABLE
        self.setup_ui()
        self.load_parties()
        self.clear_form()
        self._init_ui_memory()

    @staticmethod
    def _display_party_type(party_type):
        """Return the user-facing spelling for a stored party type."""
        return 'Debtor' if str(party_type or '') == 'Debitor' else str(party_type or '')

    @staticmethod
    def _stored_party_type(party_type):
        """Return the database spelling for a user-facing party type."""
        return 'Debitor' if str(party_type or '') == 'Debtor' else str(party_type or '')

    def setup_ui(self):
        self.setStyleSheet(theme.master_page_background_style())
        layout = QVBoxLayout(self)
        title = QLabel('Debtor / Creditor')
        title.setStyleSheet(theme.master_page_title_style(24))
        layout.addWidget(title)
        nav_layout = QHBoxLayout()
        nav_layout.setContentsMargins(0, 10, 0, 10)
        self.entry_btn = QPushButton('Party Entry')
        self.entry_btn.setStyleSheet(theme.master_nav_primary_button_style())
        self.entry_btn.clicked.connect(self.show_entry_page)
        self.list_btn = QPushButton('Party List')
        self.list_btn.setStyleSheet(theme.master_nav_secondary_button_style())
        self.list_btn.clicked.connect(self.show_list_page)
        nav_layout.addWidget(self.entry_btn)
        nav_layout.addWidget(self.list_btn)
        nav_layout.addStretch()
        layout.addLayout(nav_layout)
        self.stack_widget = QStackedWidget()
        self.entry_page = self.create_entry_page()
        self.list_page = self.create_list_page()
        self.stack_widget.addWidget(self.entry_page)
        self.stack_widget.addWidget(self.list_page)
        self.party_tab_widget = self.stack_widget
        layout.addWidget(self.stack_widget)

    def showEvent(self, event):
        """Always reopen the master on the Party Entry view."""
        super().showEvent(event)
        if hasattr(self, 'party_tab_widget') and hasattr(self, 'entry_page'):
            self.party_tab_widget.setCurrentIndex(0)
            self.show_entry_page(clear_form=False)

    def _apply_nav_styles(self, active: str) -> None:
        self.entry_btn.setStyleSheet(theme.master_nav_primary_button_style() if active == 'entry' else theme.master_nav_secondary_button_style())
        self.list_btn.setStyleSheet(theme.master_nav_primary_button_style() if active == 'list' else theme.master_nav_secondary_button_style())

    def show_entry_page(self, clear_form=True):
        """Switch to Party Entry page."""
        self.stack_widget.setCurrentWidget(self.entry_page)
        self._apply_nav_styles('entry')
        if clear_form:
            self.clear_form()

    def show_list_page(self):
        """Switch to Party List page."""
        self.stack_widget.setCurrentWidget(self.list_page)
        self._apply_nav_styles('list')
        self.load_parties()

    def label(self, text):
        lbl = QLabel(text)
        lbl.setStyleSheet(theme.master_label_style())
        return lbl

    def create_entry_page(self):
        container = QFrame()
        container.setStyleSheet(theme.master_panel_frame_style())
        outer_layout = QVBoxLayout(container)
        outer_layout.setContentsMargins(16, 14, 16, 14)
        outer_layout.setSpacing(12)
        input_style = theme.sales_compact_input_style()
        label_style = theme.sales_micro_label_style()
        self._code_manually_edited = False
        STD_WIDTH = 210
        WIDE_WIDTH = 542
        CODE_WIDTH = 95
        FIELD_HEIGHT = 26
        LABEL_COL_WIDTH = 110

        def make_label(text):
            lbl = QLabel(text)
            lbl.setStyleSheet(label_style)
            return lbl
        left_align = Qt.AlignLeft | Qt.AlignVCenter
        form_grid = QGridLayout()
        form_grid.setContentsMargins(0, 0, 0, 0)
        form_grid.setHorizontalSpacing(12)
        form_grid.setVerticalSpacing(10)
        form_grid.setColumnMinimumWidth(0, LABEL_COL_WIDTH)
        form_grid.setColumnMinimumWidth(2, LABEL_COL_WIDTH)
        form_grid.setColumnStretch(4, 1)
        self.party_name_input = QLineEdit()
        self.party_name_input.setStyleSheet(input_style)
        self.party_name_input.setFixedWidth(STD_WIDTH)
        self.party_name_input.setFixedHeight(FIELD_HEIGHT)
        self.party_name_input.textChanged.connect(self.on_party_name_text_changed)
        self.code_field = QLineEdit()
        self.code_field.setStyleSheet(input_style + '\n            QLineEdit {\n                font-weight: bold;\n                text-transform: uppercase;\n            }\n        ')
        self.code_field.setFixedWidth(CODE_WIDTH)
        self.code_field.setFixedHeight(FIELD_HEIGHT)
        self.code_field.setMaxLength(7)
        self.code_field.textEdited.connect(self.on_code_text_edited)
        form_grid.addWidget(make_label('Party Name *'), 0, 0, left_align)
        form_grid.addWidget(self.party_name_input, 0, 1, left_align)
        form_grid.addWidget(make_label('Code'), 0, 2, left_align)
        form_grid.addWidget(self.code_field, 0, 3, left_align)
        self.party_type_combo = QComboBox()
        self.party_type_combo.addItems(['Debtor', 'Creditor', 'Both'])
        self.party_type_box = self.party_type_combo
        self.party_type_combo.setStyleSheet(input_style)
        self.party_type_combo.setFixedWidth(STD_WIDTH)
        self.party_type_combo.setFixedHeight(FIELD_HEIGHT)
        form_grid.addWidget(make_label('Party Type *'), 1, 0, left_align)
        form_grid.addWidget(self.party_type_combo, 1, 1, left_align)
        self.opening_balance_input = QLineEdit()
        self.opening_balance_input.setStyleSheet(input_style)
        self.opening_balance_input.setFixedWidth(STD_WIDTH)
        self.opening_balance_input.setFixedHeight(FIELD_HEIGHT)
        self.mobile_input = QLineEdit()
        self.mobile_input.setStyleSheet(input_style)
        self.mobile_input.setFixedWidth(STD_WIDTH)
        self.mobile_input.setFixedHeight(FIELD_HEIGHT)
        form_grid.addWidget(make_label('Opening Balance'), 2, 0, left_align)
        form_grid.addWidget(self.opening_balance_input, 2, 1, left_align)
        form_grid.addWidget(make_label('Mobile Number'), 2, 2, left_align)
        form_grid.addWidget(self.mobile_input, 2, 3, left_align)
        self.email_input = QLineEdit()
        self.email_input.setStyleSheet(input_style)
        self.email_input.setFixedWidth(STD_WIDTH)
        self.email_input.setFixedHeight(FIELD_HEIGHT)
        self.credit_limit_input = QLineEdit()
        self.credit_limit_input.setStyleSheet(input_style)
        self.credit_limit_input.setFixedWidth(STD_WIDTH)
        self.credit_limit_input.setFixedHeight(FIELD_HEIGHT)
        form_grid.addWidget(make_label('Email'), 3, 0, left_align)
        form_grid.addWidget(self.email_input, 3, 1, left_align)
        form_grid.addWidget(make_label('Credit Limit'), 3, 2, left_align)
        form_grid.addWidget(self.credit_limit_input, 3, 3, left_align)
        self.gstin_input = QLineEdit()
        self.gstin_input.setStyleSheet(input_style)
        self.gstin_input.setFixedWidth(STD_WIDTH)
        self.gstin_input.setFixedHeight(FIELD_HEIGHT)
        self.gstin_input.setMaxLength(15)
        self.gstin_input.textChanged.connect(self.on_gstin_changed)
        self.state_combo = QComboBox()
        self.state_combo.setEditable(True)
        self.state_dropdown = self.state_combo
        self.state_combo.setStyleSheet(input_style)
        self.state_combo.setFixedWidth(STD_WIDTH)
        self.state_combo.setFixedHeight(FIELD_HEIGHT)
        self.state_combo.addItems([''])
        from .theme import GST_STATE_CODES
        for state in sorted(GST_STATE_CODES.values()):
            self.state_combo.addItem(state)
        form_grid.addWidget(make_label('GSTIN'), 4, 0, left_align)
        form_grid.addWidget(self.gstin_input, 4, 1, left_align)
        form_grid.addWidget(make_label('State'), 4, 2, left_align)
        form_grid.addWidget(self.state_combo, 4, 3, left_align)
        self.contact_person_input = QLineEdit()
        self.contact_person_input.setStyleSheet(input_style)
        self.contact_person_input.setFixedWidth(WIDE_WIDTH)
        self.contact_person_input.setFixedHeight(FIELD_HEIGHT)
        self.contact_person_input.textChanged.connect(lambda text: self._apply_title_case(self.contact_person_input, text))
        form_grid.addWidget(make_label('Contact Person'), 5, 0, left_align)
        form_grid.addWidget(self.contact_person_input, 5, 1, 1, 3, left_align)
        self.address_input = QLineEdit()
        self.address_input.setStyleSheet(input_style)
        self.address_input.setFixedWidth(WIDE_WIDTH)
        self.address_input.setFixedHeight(FIELD_HEIGHT)
        self.address_input.textChanged.connect(lambda text: self._apply_title_case(self.address_input, text))
        form_grid.addWidget(make_label('Address'), 6, 0, left_align)
        form_grid.addWidget(self.address_input, 6, 1, 1, 3, left_align)
        self.notes_input = QLineEdit()
        self.notes_input.setStyleSheet(input_style)
        self.notes_input.setFixedWidth(WIDE_WIDTH)
        self.notes_input.setFixedHeight(FIELD_HEIGHT)
        self.notes_input.textChanged.connect(lambda text: self._apply_title_case(self.notes_input, text))
        form_grid.addWidget(make_label('Notes'), 7, 0, left_align)
        form_grid.addWidget(self.notes_input, 7, 1, 1, 3, left_align)
        outer_layout.addLayout(form_grid)
        outer_layout.addSpacing(4)
        actions_row = QHBoxLayout()
        actions_row.setSpacing(8)
        self.save_btn = QPushButton('Save')
        self.save_btn.setObjectName('save_btn')
        self.save_btn.clicked.connect(self.save)
        self.save_btn.setStyleSheet(theme.master_save_button_style())
        clear_btn = QPushButton('Clear')
        self.clear_btn = clear_btn
        clear_btn.setStyleSheet(theme.master_clear_button_style())
        clear_btn.clicked.connect(self.clear_form)
        actions_row.addWidget(self.save_btn)
        actions_row.addWidget(clear_btn)
        actions_row.addStretch()
        outer_layout.addLayout(actions_row)
        outer_layout.addStretch()
        self._install_party_entry_event_filters()
        return container

    def _party_entry_focus_chain(self):
        """Ordered field chain for Enter/Escape navigation."""
        return [self.party_name_input, self.code_field, self.party_type_combo, self.opening_balance_input, self.mobile_input, self.email_input, self.credit_limit_input, self.gstin_input, self.state_combo, self.contact_person_input, self.address_input, self.notes_input, self.save_btn]

    def _install_party_entry_event_filters(self):
        self._party_type_popup_open = False
        for widget in self._party_entry_focus_chain():
            widget.installEventFilter(self)
        if hasattr(self.party_type_combo, 'view') and self.party_type_combo.view():
            self.party_type_combo.view().installEventFilter(self)
            self.party_type_combo.view().viewport().installEventFilter(self)
        if self.state_combo.isEditable() and self.state_combo.lineEdit():
            self.state_combo.lineEdit().installEventFilter(self)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.KeyPress:
            key = event.key()
            popup_view = self.party_type_combo.view()
            popup_viewport = popup_view.viewport() if popup_view else None
            if obj in (popup_view, popup_viewport):
                if key in (Qt.Key_Return, Qt.Key_Enter):
                    index = self.party_type_combo.view().currentIndex()
                    if index.isValid():
                        self.party_type_combo.setCurrentIndex(index.row())
                    self.party_type_combo.hidePopup()
                    self._party_type_popup_open = False
                    if self._check_duplicate_party_name_realtime():
                        return True
                    self.focus_and_force_select(self.opening_balance_input)
                    return True
                if key == Qt.Key_Escape:
                    self.party_type_combo.hidePopup()
                    self._party_type_popup_open = False
                    self.focus_and_force_select(self.code_field)
                    return True
            if key in (Qt.Key_Return, Qt.Key_Enter):
                return self._handle_party_entry_enter(obj)
            if key == Qt.Key_Escape:
                return self._handle_party_entry_escape(obj)
        return super().eventFilter(obj, event)

    def _normalise_focus_widget(self, obj):
        if self.state_combo.isEditable() and obj == self.state_combo.lineEdit():
            return self.state_combo
        return obj

    def _handle_party_entry_enter(self, obj):
        widget = self._normalise_focus_widget(obj)
        if widget == self.code_field:
            if self._check_duplicate_party_code_realtime():
                return True
            self.focus_and_force_select(self.party_type_combo)
            return True
        if widget == self.party_type_combo:
            if self._party_type_popup_open or self.party_type_combo.view().isVisible():
                index = self.party_type_combo.view().currentIndex()
                if index.isValid():
                    self.party_type_combo.setCurrentIndex(index.row())
                self.party_type_combo.hidePopup()
                self._party_type_popup_open = False
                if self._check_duplicate_party_name_realtime():
                    return True
                self.focus_and_force_select(self.opening_balance_input)
            else:
                self._party_type_popup_open = True
                self.party_type_box.showPopup()
            return True
        chain = self._party_entry_focus_chain()
        if widget in chain:
            current_index = chain.index(widget)
            if current_index < len(chain) - 1:
                self.focus_and_force_select(chain[current_index + 1])
            else:
                self.save()
            return True
        return False

    def _handle_party_entry_escape(self, obj):
        widget = self._normalise_focus_widget(obj)
        if widget == self.party_type_combo and self.party_type_combo.view().isVisible():
            self.party_type_combo.hidePopup()
            self._party_type_popup_open = False
        chain = self._party_entry_focus_chain()
        if widget in chain:
            current_index = chain.index(widget)
            if current_index > 0:
                self.focus_and_force_select(chain[current_index - 1])
            return True
        return False

    def create_list_page(self):
        container = QFrame()
        container.setObjectName('partyListOuterFrame')
        container.setStyleSheet(theme.master_panel_frame_style('partyListOuterFrame'))
        layout = QVBoxLayout(container)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        list_title = QLabel('Party List')
        list_title.setStyleSheet(theme.master_page_title_style(18))
        layout.addWidget(list_title)
        search_filter_layout = QHBoxLayout()
        search_filter_layout.setContentsMargins(0, 0, 0, 10)
        search_label = QLabel('Search:')
        search_label.setStyleSheet(theme.master_label_style())
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText('Search by party name or mobile number...')
        self.search_input.setStyleSheet(theme.sales_compact_input_style())
        self.search_input.textChanged.connect(self.apply_filters)
        filter_label = QLabel('Filter:')
        filter_label.setStyleSheet(theme.master_label_style())
        self.filter_combo = QComboBox()
        self.filter_combo.addItems(['All', 'Debtor', 'Creditor'])
        self.filter_combo.setStyleSheet(theme.master_combo_style())
        self.filter_combo.currentTextChanged.connect(self.on_filter_changed)
        export_btn = QPushButton('Export')
        export_btn.setStyleSheet(theme.master_primary_action_button_style('8px 16px', 14))
        export_btn.clicked.connect(self.show_export_menu)
        search_filter_layout.addWidget(search_label)
        search_filter_layout.addWidget(self.search_input)
        search_filter_layout.addWidget(filter_label)
        search_filter_layout.addWidget(self.filter_combo)
        search_filter_layout.addWidget(export_btn)
        search_filter_layout.addStretch()
        layout.addLayout(search_filter_layout)
        table_container = QFrame()
        table_container.setObjectName('partyListTableContainer')
        table_container.setStyleSheet(theme.master_panel_frame_style('partyListTableContainer'))
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.setSpacing(0)
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(['SL No', 'Party Name', 'Code', 'Party Type', 'Opening Balance', 'Mobile Number', 'Email'])
        apply_read_only_report_table_selection(self.table)
        self.table.itemSelectionChanged.connect(self.on_table_selection_changed)
        self.table.itemDoubleClicked.connect(self.on_table_double_click)
        self.table.setCornerButtonEnabled(False)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_party_list_context_menu)
        self.table.setFrameShape(QFrame.NoFrame)
        self.table.setFrameShadow(QFrame.Plain)
        self.table.setLineWidth(0)
        self.table.setMidLineWidth(0)
        self.table.setContentsMargins(0, 0, 0, 0)
        self.table.setViewportMargins(0, 0, 0, 0)
        header = self.table.horizontalHeader()
        header.setVisible(True)
        header.setFixedHeight(36)
        header.setMinimumHeight(36)
        header.setDefaultSectionSize(36)
        header.setHighlightSections(False)
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setStretchLastSection(False)
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setMinimumSectionSize(60)
        self.table.setColumnWidth(0, 80)
        self.table.setColumnWidth(1, 240)
        self.table.setColumnWidth(2, 110)
        self.table.setColumnWidth(3, 140)
        self.table.setColumnWidth(4, 140)
        self.table.setColumnWidth(5, 150)
        self.table.setColumnWidth(6, 220)
        table_layout.addWidget(self.table)
        layout.addWidget(table_container)
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 10, 0, 0)
        edit_btn = QPushButton('Edit Selected')
        edit_btn.setStyleSheet(theme.master_nav_primary_button_style())
        edit_btn.clicked.connect(self.edit_selected_party)
        delete_btn = QPushButton('Delete Selected')
        delete_btn.setStyleSheet(theme.master_danger_action_button_style())
        delete_btn.clicked.connect(self.delete_selected_party)
        button_layout.addWidget(edit_btn)
        button_layout.addWidget(delete_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)
        return container

    def on_text_changed(self, widget, text):
        """Handle text change to capitalize first letter for relevant fields."""
        if widget and hasattr(widget, 'cursorPosition'):
            cursor_pos = widget.cursorPosition()
            capitalized_text = self.capitalize_first_letter(text)
            if capitalized_text != text:
                widget.blockSignals(True)
                widget.setText(capitalized_text)
                widget.setCursorPosition(cursor_pos)
                widget.blockSignals(False)

    def on_party_name_text_changed(self, text):
        self._apply_title_case(self.party_name_input, text)
        if not self._code_manually_edited:
            generated_code = normalise_party_code(self.party_name_input.text())
            self.code_field.blockSignals(True)
            self.code_field.setText(generated_code)
            self.code_field.blockSignals(False)

    def on_code_text_edited(self, text):
        self._code_manually_edited = True
        clean_code = normalise_party_code(text)
        if clean_code != text:
            cursor_pos = self.code_field.cursorPosition()
            self.code_field.blockSignals(True)
            self.code_field.setText(clean_code)
            self.code_field.setCursorPosition(min(cursor_pos, len(clean_code)))
            self.code_field.blockSignals(False)

    def on_gstin_changed(self, text):
        """Handle GSTIN text change to convert to uppercase and auto-fill state."""
        if self.gstin_input and hasattr(self.gstin_input, 'cursorPosition'):
            cursor_pos = self.gstin_input.cursorPosition()
            uppercase_text = ''.join((char for char in text if char.isalnum()))[:15].upper()
            if uppercase_text != text:
                self.gstin_input.blockSignals(True)
                self.gstin_input.setText(uppercase_text)
                self.gstin_input.setCursorPosition(min(cursor_pos, len(uppercase_text)))
                self.gstin_input.blockSignals(False)
            if not uppercase_text.strip() and hasattr(self, 'state_dropdown'):
                self.state_dropdown.blockSignals(True)
                self.state_dropdown.setCurrentIndex(-1)
                if self.state_dropdown.isEditable() and self.state_dropdown.lineEdit():
                    self.state_dropdown.lineEdit().clear()
                self.state_dropdown.blockSignals(False)
                return
            if len(uppercase_text) >= 2 and hasattr(self, 'state_combo'):
                from .theme import GST_STATE_CODES
                state_code = uppercase_text[:2]
                if state_code in GST_STATE_CODES:
                    derived_state = GST_STATE_CODES[state_code]
                    current_state = self.state_combo.currentText()
                    if not current_state:
                        self.state_combo.blockSignals(True)
                        self.state_combo.setCurrentText(derived_state)
                        self.state_combo.blockSignals(False)

    def capitalize_first_letter(self, text):
        """Capitalize the first letter of the text."""
        if not text:
            return text
        return text[0].upper() + text[1:]

    def _to_title_case(self, text):
        """Return text with the first letter of every word upper-cased.

        Only the leading character of each whitespace-separated word is forced
        to uppercase; every other character is preserved exactly as typed. The
        string length never changes, which keeps cursor restoration trivial.
        """
        result = []
        capitalize_next = True
        for char in text:
            if char.isspace():
                capitalize_next = True
                result.append(char)
            elif capitalize_next:
                result.append(char.upper())
                capitalize_next = False
            else:
                result.append(char)
        return ''.join(result)

    def _apply_title_case(self, widget, text):
        """Auto-capitalize the first letter of each word on the fly.

        Signals are blocked around the in-place rewrite so the textChanged
        handler cannot recurse, and the caret position is saved and restored so
        the user's typing rhythm is never disrupted.
        """
        if widget is None:
            return
        titled = self._to_title_case(text)
        if titled == text:
            return
        cursor_pos = widget.cursorPosition() if hasattr(widget, 'cursorPosition') else None
        widget.blockSignals(True)
        widget.setText(titled)
        if cursor_pos is not None:
            widget.setCursorPosition(min(cursor_pos, len(titled)))
        widget.blockSignals(False)

    def _show_deferred_warning(self, title, message, focus_widget=None):
        """Show modal warnings after the active text event has unwound."""

        def show_warning():
            QMessageBox.warning(self, title, message)
            if focus_widget:
                focus_widget.setFocus()
                if hasattr(focus_widget, 'selectAll'):
                    focus_widget.selectAll()
        QTimer.singleShot(0, show_warning)

    def _check_duplicate_party_name_realtime(self):
        """Fire an immediate duplicate Party Name check on the Party Type Enter key.

        Returns True when a duplicate is found so the caller can block focus
        from advancing to the Opening Balance field. Focus is forced back into
        the Party Name field so the user can correct it instantly.
        """
        active_company = active_company_manager.get_active_company()
        if not active_company:
            return False
        party_name = self.party_name_input.text().strip()
        if not party_name:
            return False
        party_type = self._stored_party_type(self.party_type_combo.currentText().strip())
        try:
            exists = self.db.party_name_exists(active_company['id'], party_name, self.current_party_id, party_type)
        except Exception:
            return False
        if not exists:
            return False
        display_type = self._display_party_type(party_type)
        self.party_name_input.blockSignals(True)
        self.party_type_combo.blockSignals(True)
        try:
            QMessageBox.warning(self, 'Duplicate Party Name', f'This Party Name is already saved as a {display_type}! Please use a unique name.')
        finally:
            self.party_name_input.blockSignals(False)
            self.party_type_combo.blockSignals(False)
        self.party_name_input.setFocus()
        self.party_name_input.selectAll()
        return True

    def _check_duplicate_party_code_realtime(self):
        """Fire an immediate duplicate Short Code check on the Code Enter key.

        Returns True when the code is already assigned so the caller can halt
        focus from advancing to the Party Type field. Focus stays locked in the
        Code field with all text selected for a quick overwrite.
        """
        active_company = active_company_manager.get_active_company()
        if not active_company:
            return False
        party_code = normalise_party_code(self.code_field.text())
        if not party_code:
            return False
        try:
            exists = self.db.party_code_exists(active_company['id'], party_code, self.current_party_id)
        except Exception:
            return False
        if not exists:
            return False
        self.code_field.blockSignals(True)
        self.party_type_combo.blockSignals(True)
        try:
            QMessageBox.warning(self, 'Duplicate Short Code', 'This Short Code is already assigned to another party! You must manually enter a different unique code to save.')
        finally:
            self.code_field.blockSignals(False)
            self.party_type_combo.blockSignals(False)
        self.code_field.setFocus()
        self.code_field.selectAll()
        return True

    def clear_form(self):
        for widget in [self.party_name_input, self.code_field, self.opening_balance_input, self.mobile_input, self.email_input, self.gstin_input, self.credit_limit_input, self.contact_person_input, self.address_input, self.notes_input]:
            widget.blockSignals(True)
            widget.clear()
            widget.blockSignals(False)
        self.party_type_combo.blockSignals(True)
        self.party_type_combo.setCurrentIndex(0)
        self.party_type_combo.blockSignals(False)
        self._code_manually_edited = False
        self.state_combo.blockSignals(True)
        self.state_combo.setCurrentIndex(-1)
        if self.state_combo.isEditable() and self.state_combo.lineEdit():
            self.state_combo.lineEdit().clear()
        self.state_combo.blockSignals(False)
        self.current_party_id = None
        if hasattr(self, 'save_btn'):
            self.save_btn.setText('Save')

    def save(self):
        active_company = active_company_manager.get_active_company()
        if not active_company:
            self._show_deferred_warning('No Active Company', 'Please open a company first.')
            return
        party_name = self.party_name_input.text().strip()
        if not party_name:
            self._show_deferred_warning('Validation Error', 'Party Name is required.', self.party_name_input)
            return
        party_type = self._stored_party_type(self.party_type_combo.currentText())
        if not party_type:
            self._show_deferred_warning('Validation Error', 'Party Type is required.', self.party_type_combo)
            return
        try:
            party_data = {'name': party_name, 'party_code': normalise_party_code(self.code_field.text()), 'party_type': party_type, 'opening_balance': self.opening_balance_input.text() or '0', 'mobile_number': self.mobile_input.text().strip(), 'email': self.email_input.text().strip(), 'gstin': self.gstin_input.text().strip(), 'state': self.state_combo.currentText().strip(), 'credit_limit': self.credit_limit_input.text() or '0', 'contact_person': self.contact_person_input.text().strip(), 'address': self.address_input.text().strip(), 'notes': self.notes_input.text().strip()}
            validation_result = self.party_logic.validate_party_data(party_data, self.current_party_id, active_company['id'])
            if not validation_result['success']:
                focus_widget = self.code_field if validation_result.get('field') == 'party_code' else None
                self._show_deferred_warning('Validation Error', validation_result['message'], focus_widget)
                return
            save_result = self.party_logic.save_party(active_company['id'], party_data, self.current_party_id)
            if save_result['success']:
                QMessageBox.information(self, 'Success', 'Party saved successfully.')
                self.clear_form()
                QTimer.singleShot(0, lambda: self.party_name_input.setFocus())
                self.load_parties()
                self.party_saved.emit()
            else:
                self._show_deferred_warning('Error', save_result['message'])
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to save party: {str(e)}')

    def load_parties(self):
        """Load all parties from database into memory."""
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                self.parties_data = []
                self.visible_parties = []
                self.render_parties([])
                return
            result = self.party_logic.get_parties(active_company['id'])
            if result['success']:
                self.parties_data = result['data']
                self.apply_filters()
            else:
                QMessageBox.critical(self, 'Error', result['message'])
                self.parties_data = []
                self.visible_parties = []
                self.render_parties([])
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load parties: {str(e)}')
            self.parties_data = []
            self.visible_parties = []
            self.render_parties([])

    def render_parties(self, parties):
        """Render parties in table."""
        self.table.setRowCount(len(parties))
        for row, party in enumerate(parties):
            sl_no_item = QTableWidgetItem(str(row + 1))
            sl_no_item.setData(Qt.UserRole, party['id'])
            self.table.setItem(row, 0, sl_no_item)
            name_item = QTableWidgetItem(party_display_name(party))
            name_item.setData(Qt.UserRole, party['id'])
            self.table.setItem(row, 1, name_item)
            code_item = QTableWidgetItem(party.get('party_code') or '')
            code_item.setData(Qt.UserRole, party['id'])
            code_font = code_item.font()
            code_font.setBold(True)
            code_item.setFont(code_font)
            self.table.setItem(row, 2, code_item)
            party_type_item = QTableWidgetItem(self._display_party_type(party['party_type']))
            party_type_item.setData(Qt.UserRole, party['id'])
            self.table.setItem(row, 3, party_type_item)
            opening_balance = float(party['opening_balance'] or 0)
            balance_item = QTableWidgetItem(f'{opening_balance:.2f}')
            balance_item.setData(Qt.UserRole, party['id'])
            balance_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row, 4, balance_item)
            mobile_item = QTableWidgetItem(party['mobile_number'] or '')
            mobile_item.setData(Qt.UserRole, party['id'])
            self.table.setItem(row, 5, mobile_item)
            email_item = QTableWidgetItem(party['email'] or '')
            email_item.setData(Qt.UserRole, party['id'])
            self.table.setItem(row, 6, email_item)

    def on_filter_changed(self, filter_value):
        """Handle filter dropdown change."""
        self.current_filter = filter_value
        self.apply_filters()

    def apply_filters(self):
        """Apply both search and filter to parties."""
        search_term = self.search_input.text().strip()
        filter_value = self.current_filter
        filtered_parties = []
        for party in self.parties_data:
            if filter_value != 'All':
                if party['party_type'] != self._stored_party_type(filter_value):
                    continue
            if search_term:
                name_match = search_term.lower() in (party['name'] or '').lower()
                code_match = search_term.lower() in (party.get('party_code') or '').lower()
                mobile_match = search_term.lower() in (party['mobile_number'] or '').lower()
                if not (name_match or code_match or mobile_match):
                    continue
            filtered_parties.append(party)
        self.visible_parties = filtered_parties
        self.render_parties(filtered_parties)

    def on_table_selection_changed(self):
        """Handle table row selection change."""
        pass

    def on_table_double_click(self, item):
        """Handle double-click on table row to edit party."""
        self.edit_selected_party()

    def edit_selected_party(self):
        """Edit the selected party by switching to entry page."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, 'No Selection', 'Please select a party to edit.')
            return
        party_id = selected_items[0].data(Qt.UserRole)
        if not party_id:
            QMessageBox.warning(self, 'Error', 'Unable to identify selected party.')
            return
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                return
            result = self.party_logic.get_party_by_id(active_company['id'], party_id)
            if result['success'] and result['data']:
                self.load_party_to_form(result['data'])
                self.show_entry_page(clear_form=False)
            else:
                QMessageBox.warning(self, 'Error', 'Party not found.')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load party: {str(e)}')

    def delete_selected_party(self):
        """Delete the selected party."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, 'No Selection', 'Please select a party to delete.')
            return
        party_id = selected_items[0].data(Qt.UserRole)
        selected_row = self.table.currentRow()
        party_name_item = self.table.item(selected_row, 1)
        party_name = party_name_item.text() if party_name_item else 'selected party'
        if not party_id:
            QMessageBox.warning(self, 'Error', 'Unable to identify selected party.')
            return
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                return
            reply = QMessageBox.question(self, 'Confirm Delete', f"Are you sure you want to delete '{party_name}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                result = self.party_logic.delete_party(active_company['id'], party_id)
                if result['success']:
                    QMessageBox.information(self, 'Success', 'Party deleted successfully.')
                    self.load_parties()
                else:
                    QMessageBox.warning(self, 'Error', result['message'])
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to delete party: {str(e)}')

    def load_party_to_form(self, party):
        """Load party data into form fields."""
        self.current_party_id = party['id']
        text_values = [(self.party_name_input, party['name']), (self.code_field, party.get('party_code') or normalise_party_code(party.get('name'))), (self.opening_balance_input, str(party['opening_balance'])), (self.mobile_input, party['mobile_number'] or ''), (self.email_input, party['email'] or ''), (self.gstin_input, party['gstin'] or ''), (self.credit_limit_input, str(party['credit_limit'])), (self.contact_person_input, party['contact_person'] or ''), (self.address_input, party['address'] or ''), (self.notes_input, party['notes'] or '')]
        for widget, value in text_values:
            widget.blockSignals(True)
            widget.setText(value)
            widget.blockSignals(False)
        self._code_manually_edited = bool(self.code_field.text().strip())
        self.party_type_combo.blockSignals(True)
        self.party_type_combo.setCurrentText(self._display_party_type(party['party_type']))
        self.party_type_combo.blockSignals(False)
        self.state_combo.blockSignals(True)
        state = party.get('state', '') or ''
        if state:
            self.state_combo.setCurrentText(state)
        else:
            self.state_combo.setCurrentIndex(-1)
            if self.state_combo.isEditable() and self.state_combo.lineEdit():
                self.state_combo.lineEdit().clear()
        self.state_combo.blockSignals(False)
        if hasattr(self, 'save_btn'):
            self.save_btn.setText('Update')

    def load_party_for_edit(self, party_id):
        """Load party by ID for editing from Purchase Entry."""
        try:
            active_company = active_company_manager.get_active_company()
            if not active_company:
                return
            result = self.party_logic.get_party_by_id(active_company['id'], party_id)
            if result['success'] and result['data']:
                party = result['data']
                self.load_party_to_form(party)
                self.show_entry_page(clear_form=False)
        except Exception as e:
            print(f'Error loading party for edit: {e}')

    def keyPressEvent(self, event):
        """Handle key press events for Enter and Esc navigation."""
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            focus_widget = self.focusWidget()
            field_order = [self.party_name_input, self.party_type_combo, self.opening_balance_input, self.mobile_input, self.email_input, self.credit_limit_input, self.gstin_input, self.contact_person_input, self.address_input, self.notes_input]
            if focus_widget in field_order:
                current_index = field_order.index(focus_widget)
                if current_index < len(field_order) - 1:
                    next_field = field_order[current_index + 1]
                    self.focus_and_force_select(next_field)
                else:
                    self.save()
            elif focus_widget == self.save_btn:
                self.save()
        elif event.key() == Qt.Key_Escape:
            focus_widget = self.focusWidget()
            if focus_widget == self.save_btn:
                self.focus_and_force_select(self.notes_input)
                return
            field_order = [self.party_name_input, self.party_type_combo, self.opening_balance_input, self.mobile_input, self.email_input, self.credit_limit_input, self.gstin_input, self.contact_person_input, self.address_input, self.notes_input]
            if focus_widget in field_order:
                current_index = field_order.index(focus_widget)
                if current_index > 0:
                    prev_field = field_order[current_index - 1]
                    self.focus_and_force_select(prev_field)
        else:
            super().keyPressEvent(event)

    def focus_and_force_select(self, widget):
        """Set focus and select all text with proper timing."""
        widget.setFocus()
        if isinstance(widget, QLineEdit):
            QTimer.singleShot(0, widget.selectAll)

    def focus_and_select(self, widget):
        """Set focus and select all text with proper timing."""
        widget.setFocus()
        if isinstance(widget, QLineEdit):
            QTimer.singleShot(0, widget.selectAll)

    def show_party_list_context_menu(self, position):
        """Show context menu for party list table."""
        menu = QMenu(self)
        select_all_action = menu.addAction('Select All')
        select_all_action.triggered.connect(self.select_all_visible_rows)
        clear_selection_action = menu.addAction('Clear Selection')
        clear_selection_action.triggered.connect(self.clear_selection)
        menu.exec(self.table.mapToGlobal(position))

    def select_all_visible_rows(self):
        """Select all visible rows in the table."""
        self.table.selectAll()

    def clear_selection(self):
        """Clear all selections in the table."""
        self.table.clearSelection()

    def get_visible_table_data(self):
        """Extract visible table data for export."""
        data = []
        headers = []
        for col in range(self.table.columnCount()):
            headers.append(self.table.horizontalHeaderItem(col).text())
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else '')
            data.append(row_data)
        return (headers, data)

    def show_export_menu(self):
        """Show export format selection dialog."""
        if not self.visible_parties:
            QMessageBox.warning(self, 'No Data', 'No parties to export.')
            return
        dialog = QDialog(self)
        dialog.setWindowTitle('Export Party List')
        dialog.setStyleSheet(report_detail_dialog_style())
        dialog.setFixedSize(350, 150)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        title_label = QLabel('Select Export Format:')
        title_label.setStyleSheet(theme.master_dialog_heading_style())
        layout.addWidget(title_label)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        pdf_btn = QPushButton('PDF')
        pdf_btn.setCheckable(True)
        pdf_btn.setEnabled(self.pdf_available)
        pdf_btn.clicked.connect(lambda: self.export_to_pdf(dialog))
        excel_btn = QPushButton('Excel')
        excel_btn.setCheckable(True)
        excel_btn.setEnabled(self.excel_available)
        excel_btn.clicked.connect(lambda: self.export_to_excel(dialog))
        cancel_btn = QPushButton('Cancel')
        cancel_btn.clicked.connect(dialog.reject)
        cancel_btn.setStyleSheet(theme.master_clear_button_style())
        button_layout.addWidget(pdf_btn)
        button_layout.addWidget(excel_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        dialog.exec()

    def export_to_pdf(self, dialog=None):
        """Export visible parties to PDF."""
        if not self.pdf_available:
            QMessageBox.warning(self, 'Library Not Installed', 'Required library not installed.\n\nTo enable PDF export, install:\npip install reportlab')
            return
        try:
            headers, data = self.get_visible_table_data()
            if not data:
                QMessageBox.warning(self, 'No Data', 'No data available to export.')
                return
            file_path, _ = QFileDialog.getSaveFileName(self, 'Save PDF', 'party_list.pdf', 'PDF Files (*.pdf)')
            if not file_path:
                return
            table_data = [headers] + data
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            table = Table(table_data)
            table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -1), colors.white), ('TEXTCOLOR', (0, 1), (-1, -1), colors.black), ('GRID', (0, 0), (-1, -1), 1, colors.black), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])]))
            doc.build([table])
            QMessageBox.information(self, 'Success', 'PDF exported successfully.')
            if dialog:
                dialog.accept()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to export PDF: {str(e)}')

    def export_to_excel(self, dialog=None):
        """Export visible parties to Excel."""
        if not self.excel_available:
            QMessageBox.warning(self, 'Library Not Installed', 'Required library not installed.\n\nTo enable Excel export, install:\npip install openpyxl')
            return
        try:
            file_path, _ = QFileDialog.getSaveFileName(self, 'Save Excel', 'party_list.xlsx', 'Excel Files (*.xlsx)')
            if not file_path:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = 'Party List'
            headers = ['SL No', 'Party Name', 'Code', 'Party Type', 'Opening Balance', 'Mobile Number', 'Email']
            ws.append(headers)
            header_fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            border = Border(left=Side(style='thin', color='4b5563'), right=Side(style='thin', color='4b5563'), top=Side(style='thin', color='4b5563'), bottom=Side(style='thin', color='4b5563'))
            for col in range(1, 8):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            bold_data_font = Font(bold=True)
            for idx, party in enumerate(self.visible_parties):
                row = idx + 2
                ws.cell(row=row, column=1, value=idx + 1)
                ws.cell(row=row, column=2, value=party['name'] or '')
                code_cell = ws.cell(row=row, column=3, value=party.get('party_code') or '')
                code_cell.font = bold_data_font
                ws.cell(row=row, column=4, value=party['party_type'] or '')
                ws.cell(row=row, column=5, value=float(party['opening_balance'] or 0))
                ws.cell(row=row, column=6, value=party['mobile_number'] or '')
                ws.cell(row=row, column=7, value=party['email'] or '')
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if col == 5:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 25
            wb.save(file_path)
            QMessageBox.information(self, 'Success', 'Excel file exported successfully.')
            if dialog:
                dialog.accept()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to export Excel: {str(e)}')